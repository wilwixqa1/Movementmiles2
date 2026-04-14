from fastapi import FastAPI, HTTPException, Request, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse, Response
from pydantic import BaseModel
import httpx
import os
import json
import asyncpg
import stripe
import re
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
import gzip
import asyncio
import base64
import csv
import io
import hashlib
import time
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import jwt as pyjwt

app = FastAPI(title="Movement & Miles")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Config ---
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
DATABASE_URL = os.environ.get("DATABASE_URL", "")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "mmadmin2026")
STRIPE_SECRET_KEY = os.environ.get("STRIPE_SECRET_KEY", "")
STRIPE_WEBHOOK_SECRET = os.environ.get("STRIPE_WEBHOOK_SECRET", "")
RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")
DIGEST_RECIPIENTS = os.environ.get("DIGEST_RECIPIENTS", "")
DIGEST_FROM_EMAIL = os.environ.get("DIGEST_FROM_EMAIL", "onboarding@resend.dev")

stripe.api_key = STRIPE_SECRET_KEY

# Apple App Store Connect API
APPLE_KEY_ID = os.environ.get("APPLE_KEY_ID", "")
APPLE_ISSUER_ID = os.environ.get("APPLE_ISSUER_ID", "")
APPLE_KEY_CONTENT = os.environ.get("APPLE_KEY_CONTENT", "")
APPLE_VENDOR_NUMBER = os.environ.get("APPLE_VENDOR_NUMBER", "")

# ymove API (Session 17)
YMOVE_API_KEY = os.environ.get("YMOVE_API_KEY", "")
YMOVE_API_BASE = "https://v6-beta-api.ymove.app"
YMOVE_SITE_ID = "75"  # Movement & Miles site ID (from Tosh, S19)

# --- Database ---
db_pool = None

# --- Scheduler ---
scheduler = None

@app.on_event("startup")
async def startup():
    global db_pool, scheduler
    if DATABASE_URL:
        try:
            db_pool = await asyncpg.create_pool(DATABASE_URL, min_size=1, max_size=5)

            # Block 1: Core tables (leads, page_views, chat_sessions)
            async with db_pool.acquire() as conn:
                await conn.execute("""
                    CREATE TABLE IF NOT EXISTS leads (
                        id SERIAL PRIMARY KEY,
                        first_name TEXT,
                        email TEXT,
                        experience_level TEXT,
                        goals TEXT,
                        referral_source TEXT,
                        recommended_plan TEXT,
                        extra TEXT,
                        created_at TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                await conn.execute("""
                    CREATE TABLE IF NOT EXISTS page_views (
                        id SERIAL PRIMARY KEY,
                        page TEXT,
                        path TEXT,
                        referrer TEXT,
                        user_agent TEXT,
                        created_at TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                await conn.execute("""
                    CREATE TABLE IF NOT EXISTS chat_sessions (
                        id SERIAL PRIMARY KEY,
                        session_type TEXT,
                        message_count INTEGER DEFAULT 1,
                        created_at TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                # UTM tracking columns on leads
                for col in ['utm_source', 'utm_medium', 'utm_campaign', 'utm_term', 'utm_content', 'ym_source']:
                    await conn.execute(f"ALTER TABLE leads ADD COLUMN IF NOT EXISTS {col} TEXT DEFAULT ''")
                # S22: UTM tracking columns on page_views
                for col in ['utm_source', 'utm_medium', 'utm_campaign']:
                    await conn.execute(f"ALTER TABLE page_views ADD COLUMN IF NOT EXISTS {col} TEXT DEFAULT ''")
            print("[Startup] Block 1: Core tables ready")

            # Block 2: Subscription tables + ALTER columns
            async with db_pool.acquire() as conn:
                await conn.execute("""
                    CREATE TABLE IF NOT EXISTS subscriptions (
                        id SERIAL PRIMARY KEY,
                        stripe_customer_id TEXT,
                        stripe_subscription_id TEXT UNIQUE,
                        email TEXT,
                        status TEXT,
                        plan_interval TEXT,
                        plan_amount INTEGER,
                        currency TEXT DEFAULT 'usd',
                        source TEXT DEFAULT 'stripe',
                        trial_start TIMESTAMPTZ,
                        trial_end TIMESTAMPTZ,
                        current_period_start TIMESTAMPTZ,
                        current_period_end TIMESTAMPTZ,
                        canceled_at TIMESTAMPTZ,
                        created_at TIMESTAMPTZ DEFAULT NOW(),
                        updated_at TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                await conn.execute("ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS source TEXT DEFAULT 'stripe'")
                await conn.execute("ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS converted_at TIMESTAMPTZ")
                await conn.execute("ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS readable_id TEXT")
                await conn.execute("ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS renewal_count INTEGER DEFAULT 0")
                await conn.execute("ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS last_renewed_at TIMESTAMPTZ")
                # S16: import batch tracking for safe revert
                await conn.execute("ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS import_batch TEXT")
                # S17: store names directly on subscriptions
                await conn.execute("ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS first_name TEXT")
                await conn.execute("ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS last_name TEXT")
                # S21: UTM attribution from ymove meta parameters
                for col in ['utm_source', 'utm_medium', 'utm_campaign', 'utm_term', 'utm_content', 'ym_source']:
                    await conn.execute(f"ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS {col} TEXT DEFAULT ''")
                await conn.execute("ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS utm_meta_raw JSONB")
                await conn.execute("""
                    CREATE TABLE IF NOT EXISTS subscription_events (
                        id SERIAL PRIMARY KEY,
                        stripe_event_id TEXT UNIQUE,
                        event_type TEXT,
                        stripe_customer_id TEXT,
                        stripe_subscription_id TEXT,
                        source TEXT DEFAULT 'stripe',
                        data JSONB,
                        created_at TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                await conn.execute("ALTER TABLE subscription_events ADD COLUMN IF NOT EXISTS source TEXT DEFAULT 'stripe'")
            print("[Startup] Block 2: Subscription tables ready")

            # Block 3: Analytics tables (ad_spend, platform_metrics, ymove_webhook_log)
            async with db_pool.acquire() as conn:
                await conn.execute("""
                    CREATE TABLE IF NOT EXISTS ad_spend (
                        id SERIAL PRIMARY KEY,
                        month TEXT NOT NULL,
                        channel TEXT NOT NULL,
                        amount_cents INTEGER NOT NULL DEFAULT 0,
                        created_at TIMESTAMPTZ DEFAULT NOW(),
                        updated_at TIMESTAMPTZ DEFAULT NOW(),
                        UNIQUE(month, channel)
                    )
                """)
                await conn.execute("""
                    CREATE TABLE IF NOT EXISTS platform_metrics (
                        id SERIAL PRIMARY KEY,
                        date DATE NOT NULL,
                        source TEXT NOT NULL,
                        metric_type TEXT NOT NULL,
                        active_subscriptions INTEGER DEFAULT 0,
                        active_free_trials INTEGER DEFAULT 0,
                        new_subscriptions INTEGER DEFAULT 0,
                        renewals INTEGER DEFAULT 0,
                        conversions INTEGER DEFAULT 0,
                        cancellations INTEGER DEFAULT 0,
                        reactivations INTEGER DEFAULT 0,
                        revenue_cents INTEGER DEFAULT 0,
                        proceeds_cents INTEGER DEFAULT 0,
                        report_data JSONB,
                        created_at TIMESTAMPTZ DEFAULT NOW(),
                        updated_at TIMESTAMPTZ DEFAULT NOW(),
                        UNIQUE(date, source, metric_type)
                    )
                """)
                await conn.execute("""
                    CREATE TABLE IF NOT EXISTS ymove_webhook_log (
                        id SERIAL PRIMARY KEY,
                        event_type TEXT,
                        payload JSONB,
                        created_at TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                await conn.execute("""
                    CREATE TABLE IF NOT EXISTS ymove_sync_runs (
                        id SERIAL PRIMARY KEY,
                        status TEXT DEFAULT 'running',
                        started_at TIMESTAMPTZ DEFAULT NOW(),
                        completed_at TIMESTAMPTZ,
                        phase TEXT DEFAULT 'init',
                        progress_current INTEGER DEFAULT 0,
                        progress_total INTEGER DEFAULT 0,
                        our_active_count INTEGER DEFAULT 0,
                        ymove_active_count INTEGER DEFAULT 0,
                        results JSONB,
                        error TEXT
                    )
                """)
            print("[Startup] Block 3: Analytics tables ready")

            # Block 3b: UTM links table (S23)
            async with db_pool.acquire() as conn:
                await conn.execute("""
                    CREATE TABLE IF NOT EXISTS utm_links (
                        id SERIAL PRIMARY KEY,
                        label TEXT DEFAULT '',
                        base_url TEXT NOT NULL,
                        utm_source TEXT DEFAULT '',
                        utm_medium TEXT DEFAULT '',
                        utm_campaign TEXT DEFAULT '',
                        utm_term TEXT DEFAULT '',
                        utm_content TEXT DEFAULT '',
                        ym_source TEXT DEFAULT '',
                        full_url TEXT NOT NULL,
                        created_at TIMESTAMPTZ DEFAULT NOW()
                    )
                """)

            # Block 3c: reactivated_at column + backfill (S23)
            async with db_pool.acquire() as conn:
                await conn.execute("ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS reactivated_at TIMESTAMPTZ")
                # S26: Period-end cancellation semantics
                # cancel_state: NULL = not cancelled, 'pending' = cancel requested but still in paid period,
                #               'expired' = period ended, fully cancelled
                # pending_cancel_at: when paid period ends (status stays active until then)
                # cancel_requested_at: when user actually clicked cancel (for analytics)
                await conn.execute("ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS pending_cancel_at TIMESTAMPTZ")
                await conn.execute("ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS cancel_requested_at TIMESTAMPTZ")
                await conn.execute("ALTER TABLE subscriptions ADD COLUMN IF NOT EXISTS cancel_state TEXT")
                # Backfill: tag subs reactivated by ymove verification batches
                backfilled = await conn.execute(
                    """UPDATE subscriptions SET reactivated_at = updated_at
                       WHERE import_batch LIKE 'ymove_react_%'
                       AND reactivated_at IS NULL
                       AND status = 'active'"""
                )
                count = int(backfilled.split(" ")[-1]) if backfilled else 0
                if count > 0:
                    print(f"[Startup] Block 3c: Backfilled reactivated_at on {count} subs")

            # Block 4: Indexes (non-blocking, failure will not kill startup)
            try:
                async with db_pool.acquire() as conn:
                    await conn.execute("CREATE INDEX IF NOT EXISTS idx_leads_email_lower ON leads (lower(email))")
                    await conn.execute("CREATE INDEX IF NOT EXISTS idx_subs_email_lower ON subscriptions (lower(email))")
                    await conn.execute("CREATE INDEX IF NOT EXISTS idx_subs_trial_start ON subscriptions (trial_start) WHERE trial_start IS NOT NULL")
                    await conn.execute("CREATE INDEX IF NOT EXISTS idx_subs_import_batch ON subscriptions (import_batch) WHERE import_batch IS NOT NULL")
                print("[Startup] Block 4: Indexes ready")
            except Exception as e:
                print(f"[Startup] Index creation deferred (non-fatal): {e}")
        except Exception as e:
            print(f"Database connection failed: {e}")
            db_pool = None
    else:
        print("No DATABASE_URL â running without database")

    # Start daily digest scheduler
    if RESEND_API_KEY and DIGEST_RECIPIENTS:
        try:
            from apscheduler.schedulers.asyncio import AsyncIOScheduler
            from apscheduler.triggers.cron import CronTrigger
            scheduler = AsyncIOScheduler()
            et = ZoneInfo("America/New_York")
            scheduler.add_job(
                run_daily_digest,
                CronTrigger(hour=9, minute=0, timezone=et),
                id="daily_digest",
                replace_existing=True,
            )
            # S20: Daily shadow sync at 8:00 AM ET (1 hour before digest)
            if YMOVE_API_KEY:
                scheduler.add_job(
                    run_daily_shadow_sync,
                    CronTrigger(hour=8, minute=0, timezone=et),
                    id="daily_shadow_sync",
                    replace_existing=True,
                )
            scheduler.start()
            ymove_sync_msg = " + shadow sync 8:00 AM ET" if YMOVE_API_KEY else ""
            print(f"Daily digest scheduler started (9:00 AM ET -> {DIGEST_RECIPIENTS}{ymove_sync_msg})")
        except Exception as e:
            print(f"Scheduler startup error: {e}")
    else:
        missing = []
        if not RESEND_API_KEY:
            missing.append("RESEND_API_KEY")
        if not DIGEST_RECIPIENTS:
            missing.append("DIGEST_RECIPIENTS")
        print(f"Daily digest disabled â missing: {', '.join(missing)}")

@app.on_event("shutdown")
async def shutdown():
    global db_pool, scheduler
    if scheduler:
        scheduler.shutdown(wait=False)
    if db_pool:
        await db_pool.close()


# --- System Prompts ---

NELLY_SYSTEM_PROMPT = """You are Nelly, the AI coaching assistant for Movement & Miles (M&M), a holistic running and fitness app created by coach Meg.

PERSONALITY: Warm, encouraging, conversational. You talk like a friend who happens to be a running coach. Keep responses SHORT (2-4 sentences max). Never dump walls of text.

CRITICAL CONVERSATION RULE: Ask ONE question at a time. Never list multiple questions. Have a natural back-and-forth conversation. Guide them step by step.

LINK FORMAT: When recommending a program or answering about a page, include a clickable link using this format:
[[page:PageName]]
Available pages: [[page:Training Programs]], [[page:Race Plans]], [[page:Store]]

BUTTON FORMAT: When you want to give the user options to choose from, end your message with options in this exact format on a new line:
[Option A | Option B | Option C]

ONLY use options when there are clear choices. For open-ended questions, just ask normally without options.

CRITICAL: You may ONLY recommend programs that exist in the app. NEVER invent program names.

COMPLETE PROGRAM LIST:

RUNNING + STRENGTH MONTHLY PLANS:
Beginner: Walk to Run Part 1, Walk to Run Part 2, Miles + Bodyweight Strength, Building Endurance & Strength, Beginners: Total Package
Intermediate: Strides + Calisthenics, Outdoor Miles + Weights, Balanced Strides & Strength, Endurance & Strength
Advanced: Run + Lift, Endurance Speed & Strength, Peak Endurance & Power, 7 Weeks to 10 Miles

STRENGTH-ONLY PLANS:
Beginner: Bodyweight & Bands (4wk), Strength Starts Here (2wk), Pure Strength (6wk)
Intermediate: Stronger Strides (6wk), Total Body Power (6wk), Well Built
Advanced: Total Power & Strength (6wk), Cross-Training Power

MOBILITY & PREHAB:
All Levels: Prehab: Knee/ITBS + Mobility (4wk), Mobility Master (4wk), Trail/Road Running Prehab Essentials (6wk), Ankle Foot and Calf Strength (10 modules), ITBS/Knee Pain Workout Collection (10 modules), The Ultimate Mobility Plan (4wk)

RACE PLANS:
Beginner: 8-Week Beginner 5K Treadmill & Outdoor, Beginner 5K Plan (Outdoor), Beginner 10K Plan (Tread & Outdoor), Beginner 10K Plan (Outdoor), Beginner Half Marathon Plan (20wk), Beginner Marathon Plan (20wk)
Intermediate: Intermediate 5K Plan (Tread & Outdoor), Intermediate 5K Plan (Outdoor), Intermediate 10K Plan (Tread & Outdoor), Intermediate 10K Plan (Outdoor), Intermediate Half Marathon Plan (10wk), Intermediate 16-Week Marathon Plan
Advanced: Advanced Half Marathon (12wk), Advanced Marathon Plan, 50K Race Plan (16wk)

DETRAINING PLANS:
Beginner: Detrain Protocol
Intermediate: Recover, Restore & Reset
Advanced: The Adaptation Block

NUTRITION PLANS: Endurance Nutrition, Strength Nutrition, Weight Loss Nutrition

PROGRESSIONS:
BEGINNER RUNNING: Walk to Run Part 1 > Walk to Run Part 2 > Miles + Bodyweight Strength > Building Endurance & Strength > Beginners: Total Package
INTERMEDIATE RUNNING: Strides + Calisthenics > Outdoor Miles + Weights > Balanced Strides & Strength > Endurance & Strength
ADVANCED RUNNING: Run + Lift > Endurance Speed & Strength > Peak Endurance & Power > 7 Weeks to 10 Miles
BEGINNER STRENGTH: Bodyweight & Bands > Strength Starts Here > Pure Strength
INTERMEDIATE STRENGTH: Stronger Strides > Total Body Power > Well Built
ADVANCED STRENGTH: Total Power & Strength > Cross-Training Power
RACE ORDER: 5K > 10K > Half Marathon > Marathon > 50K (never skip)

DETRAINING RULES:
- After running/race program AND person has NOT taken 3+ weeks off: recommend detraining first
- After running/race program AND person HAS taken 3+ weeks off: skip detraining
- After strength-only program: NO detraining needed

EQUIPMENT: Ask about weights (kettlebells or dumbbells; add barbell for advanced). Ask about treadmill preference.

CONVERSATION FLOW for plan recommendations:
Set expectations first, then ask ONE AT A TIME:
1. Running+strength, strength only, or train for a race?
2. What level?
3. Can you run 3 miles without stopping?
4. Any pain?
5. Access to weights?
6. Treadmill preference?
7. If race: when is it and what distance?
Max 7 questions, then give 3 OPTIONS.

FAQs:
CANCEL: Apple/Google > subscription settings. Website > app profile > Info > Manage Subscription
PRICING: Monthly $19.99, Annual $179.99
INCLUDED: Everything - all programs, plans, nutrition
GARMIN/ANNUAL SWITCH: email support@movementandmiles.com
PAYMENT: https://movementandmiles.ymove.app/account
MISSED WORKOUTS: 1-2 = continue. 3-5 = resume easier. Week+ = repeat previous week.

FINAL RECOMMENDATIONS: Always present exactly 3 options with one-sentence explanations. Use button format."""


DIGEST_SYSTEM_PROMPT = """You are a concise analytics advisor for Movement & Miles, a fitness subscription app ($19.99/mo after 1-month free trial). Generate 3-5 bullet points for the business owner.

KEY DISTINCTIONS:
- trial_starts = new free trials, $0 revenue. Frame as "pipeline."
- conversions = trials that became paying. This is real revenue growth.
- net_paid_growth and net_trial_growth are pre-computed. Lead with these.
- Only flag concern when net_paid_growth is negative. Trial cancellations are normal attrition.

RULES:
- One short overview paragraph, then 3-5 bullets. Under 200 words total.
- Lead with net paid growth, then trial pipeline.
- If marketing UTM data is present, note top channel in one bullet. If empty, skip.
- Flag anything unusual. Suggest ONE action if warranted.
- Plain language, no jargon. Don't over-analyze zero-activity areas."""


# --- Shared Helpers ---

async def call_anthropic(system_prompt: str, messages: list, max_tokens: int = 800, cache_system: bool = False) -> str:
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="API key not configured")
    headers = {
        "x-api-key": ANTHROPIC_API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    if cache_system:
        system_val = [{"type": "text", "text": system_prompt, "cache_control": {"type": "ephemeral"}}]
        headers["anthropic-beta"] = "prompt-caching-2024-07-31"
    else:
        system_val = system_prompt
    async with httpx.AsyncClient(timeout=30.0) as client:
        try:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers=headers,
                json={
                    "model": "claude-sonnet-4-20250514",
                    "max_tokens": max_tokens,
                    "system": system_val,
                    "messages": messages,
                },
            )
            resp.raise_for_status()
            data = resp.json()
            return data["content"][0]["text"]
        except httpx.HTTPStatusError as e:
            raise HTTPException(status_code=502, detail=f"Anthropic API error: {e.response.status_code}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))


async def call_anthropic_raw(system_prompt: str, messages: list, max_tokens: int = 800, cache_system: bool = False) -> str:
    """Like call_anthropic but doesn't raise HTTPException â returns error string instead."""
    if not ANTHROPIC_API_KEY:
        return "[Error: ANTHROPIC_API_KEY not configured]"
    headers = {
        "x-api-key": ANTHROPIC_API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    if cache_system:
        system_val = [{"type": "text", "text": system_prompt, "cache_control": {"type": "ephemeral"}}]
        headers["anthropic-beta"] = "prompt-caching-2024-07-31"
    else:
        system_val = system_prompt
    async with httpx.AsyncClient(timeout=30.0) as client:
        try:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers=headers,
                json={
                    "model": "claude-sonnet-4-20250514",
                    "max_tokens": max_tokens,
                    "system": system_val,
                    "messages": messages,
                },
            )
            resp.raise_for_status()
            data = resp.json()
            return data["content"][0]["text"]
        except httpx.HTTPStatusError as e:
            body = e.response.text if hasattr(e.response, 'text') else 'no body'
            print(f"[Anthropic API error] Status: {e.response.status_code}, Body: {body}")
            return f"[AI insights unavailable: {e.response.status_code} - {body[:200]}]"
        except Exception as e:
            print(f"[Anthropic API error] {type(e).__name__}: {str(e)}")
            return f"[AI insights unavailable: {str(e)}]"


def require_admin(password: str):
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Unauthorized")


# --- Readable ID Assignment ---

async def assign_readable_id(conn, source: str) -> str:
    """Generate next persistent readable ID for a source (APPLE-0001, GOOGLE-0001, STRIPE-0001)."""
    prefix = source.upper()  # APPLE, GOOGLE, STRIPE
    row = await conn.fetchrow(
        "SELECT readable_id FROM subscriptions WHERE source = $1 AND readable_id IS NOT NULL ORDER BY readable_id DESC LIMIT 1",
        source
    )
    if row and row["readable_id"]:
        # Extract number from e.g. APPLE-0042
        try:
            num = int(row["readable_id"].split("-")[-1]) + 1
        except (ValueError, IndexError):
            num = 1
    else:
        num = 1
    return f"{prefix}-{num:04d}"


# --- Shared Parse/Extract Helpers (S22: deduplicated from inline definitions) ---

def _ts(v):
    """Convert Unix timestamp to UTC datetime."""
    if v:
        return datetime.fromtimestamp(v, tz=timezone.utc)
    return None


def _ms_to_dt(ms):
    """Convert millisecond timestamp to UTC datetime."""
    if ms:
        return datetime.fromtimestamp(ms / 1000, tz=timezone.utc)
    return None


def _parse_iso(s):
    """Parse ISO 8601 date string to datetime."""
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return None


def _find_col(row, options):
    """Find a column value by trying multiple header name variants."""
    for opt in options:
        for key in row:
            if key.strip().lower() == opt.lower():
                return (row[key] or "").strip()
    return ""


def _get_email(row):
    """Extract and normalize email from a row dict."""
    for key in ["email", "e-mail", "email address"]:
        if key in row and row[key] and "@" in row[key]:
            return row[key].strip().lower()
    return ""


def _get_source(row, default="unknown"):
    """Extract subscription source from a row dict."""
    for key in ["source", "platform"]:
        if key in row:
            val = row[key].strip().lower()
            if val in ("apple", "google", "stripe"):
                return val
    return default


def _get_date(row):
    """Parse a date from common column name variants."""
    for key in ["date", "sign up date", "signup_date", "signup date", "created_at"]:
        if key in row and row[key]:
            val = row[key].strip()
            for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"]:
                try:
                    return datetime.strptime(val[:19], fmt).replace(tzinfo=timezone.utc)
                except Exception:
                    continue
    return None


# S22: Shared test account filter (used by ymove-import-new and daily shadow sync)
_TEST_EMAIL_DOMAINS = ["ymove.app"]
_TEST_EMAIL_PATTERNS = ["test", "dsfg", "asdf", "qwer", "dkek", "sjwj", "ffas", "fasd"]


def _is_test_email(email: str) -> bool:
    """Check if an email looks like a test account."""
    if not email or "@" not in email:
        return True
    em = email.strip().lower()
    domain = em.split("@")[-1]
    local = em.split("@")[0]
    if domain in _TEST_EMAIL_DOMAINS:
        return True
    for pat in _TEST_EMAIL_PATTERNS:
        if pat in local:
            return True
    return False


# --- Pydantic Models ---

class ChatRequest(BaseModel):
    message: str
    history: list = []
    source: str = "widget"

class ChatResponse(BaseModel):
    reply: str

class LeadRequest(BaseModel):
    first_name: str = ""
    email: str = ""
    experience_level: str = ""
    goals: str = ""
    referral_source: str = ""
    recommended_plan: str = ""
    extra: str = ""
    utm_source: str = ""
    utm_medium: str = ""
    utm_campaign: str = ""
    utm_term: str = ""
    utm_content: str = ""
    ym_source: str = ""

class PageViewRequest(BaseModel):
    page: str = ""
    path: str = ""
    referrer: str = ""
    utm_source: str = ""
    utm_medium: str = ""
    utm_campaign: str = ""

class LoginRequest(BaseModel):
    password: str

class AdSpendRequest(BaseModel):
    month: str = ""
    channel: str = ""
    amount_dollars: float = 0.0


# --- API Routes ---

# Chat endpoints
@app.post("/api/chat", response_model=ChatResponse)
async def chat(req: ChatRequest):
    if not req.message.strip():
        raise HTTPException(status_code=400, detail="Message cannot be empty")
    messages = []
    for msg in req.history[-20:]:
        if msg.get("role") in ("user", "assistant") and msg.get("content"):
            messages.append({"role": msg["role"], "content": msg["content"]})
    messages.append({"role": "user", "content": req.message})
    reply = await call_anthropic(NELLY_SYSTEM_PROMPT, messages, cache_system=True)
    # Track chat session
    if db_pool:
        try:
            async with db_pool.acquire() as conn:
                await conn.execute(
                    "INSERT INTO chat_sessions (session_type, message_count) VALUES ($1, $2)",
                    req.source or "widget", 1
                )
        except Exception:
            pass
    return ChatResponse(reply=reply)


@app.post("/api/lead")
async def save_lead(lead: LeadRequest):
    if not db_pool:
        raise HTTPException(status_code=500, detail="Database not connected")
    try:
        async with db_pool.acquire() as conn:
            await conn.execute(
                """INSERT INTO leads (first_name, email, experience_level, goals, referral_source, recommended_plan, extra, utm_source, utm_medium, utm_campaign, utm_term, utm_content, ym_source)
                   VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13)""",
                lead.first_name, lead.email, lead.experience_level,
                lead.goals, lead.referral_source, lead.recommended_plan, lead.extra,
                lead.utm_source, lead.utm_medium, lead.utm_campaign,
                lead.utm_term, lead.utm_content, lead.ym_source
            )
        return {"status": "saved", "storage": "postgres"}
    except Exception as e:
        print(f"DB lead save error: {e}")
        raise HTTPException(status_code=500, detail=f"Lead save failed: {str(e)}")


# Page view tracking
@app.post("/api/page-view")
async def track_page_view(pv: PageViewRequest, request: Request):
    if db_pool:
        try:
            ua = request.headers.get("user-agent", "")
            async with db_pool.acquire() as conn:
                await conn.execute(
                    "INSERT INTO page_views (page, path, referrer, user_agent, utm_source, utm_medium, utm_campaign) VALUES ($1, $2, $3, $4, $5, $6, $7)",
                    pv.page, pv.path, pv.referrer, ua, pv.utm_source, pv.utm_medium, pv.utm_campaign
                )
        except Exception as e:
            print(f"Page view error: {e}")
    return {"status": "ok"}


# --- Stripe Webhook ---

@app.post("/webhooks/stripe")
async def stripe_webhook(request: Request):
    payload = await request.body()
    sig_header = request.headers.get("stripe-signature", "")

    # Verify webhook signature if secret is configured
    if STRIPE_WEBHOOK_SECRET:
        try:
            event = stripe.Webhook.construct_event(payload, sig_header, STRIPE_WEBHOOK_SECRET)
        except stripe.error.SignatureVerificationError:
            raise HTTPException(status_code=400, detail="Invalid signature")
        except Exception as e:
            raise HTTPException(status_code=400, detail=str(e))
    else:
        # No webhook secret configured â parse raw (dev/testing only)
        try:
            event = json.loads(payload)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid payload")

    event_type = event.get("type", "")
    event_id = event.get("id", "")
    data_obj = event.get("data", {}).get("object", {})

    if not db_pool:
        return {"status": "ok", "note": "no database"}

    async with db_pool.acquire() as conn:
        # Store raw event (idempotent via unique stripe_event_id)
        try:
            await conn.execute(
                """INSERT INTO subscription_events (stripe_event_id, event_type, stripe_customer_id, stripe_subscription_id, data)
                   VALUES ($1, $2, $3, $4, $5)
                   ON CONFLICT (stripe_event_id) DO NOTHING""",
                event_id,
                event_type,
                data_obj.get("customer", ""),
                data_obj.get("id", "") if "sub_" in data_obj.get("id", "") else data_obj.get("subscription", ""),
                json.dumps(data_obj)
            )
        except Exception as e:
            print(f"Event store error: {e}")

        # Handle subscription events
        if event_type in (
            "customer.subscription.created",
            "customer.subscription.updated",
            "customer.subscription.deleted",
            "customer.subscription.trial_will_end",
        ):
            sub = data_obj
            sub_id = sub.get("id", "")
            customer_id = sub.get("customer", "")
            status = sub.get("status", "")

            # Extract plan info from items
            plan_amount = 0
            plan_interval = ""
            items = sub.get("items", {}).get("data", [])
            if items:
                price = items[0].get("price", {}) or items[0].get("plan", {})
                plan_amount = price.get("unit_amount", 0)
                plan_interval = price.get("recurring", {}).get("interval", "") if price.get("recurring") else price.get("interval", "")

            # Get customer email
            email = ""
            if customer_id and STRIPE_SECRET_KEY:
                try:
                    cust = stripe.Customer.retrieve(customer_id)
                    email = cust.get("email", "")
                except Exception:
                    pass

            # Convert timestamps

            try:
                await conn.execute("""
                    INSERT INTO subscriptions (
                        stripe_customer_id, stripe_subscription_id, email, status,
                        plan_interval, plan_amount, currency, source,
                        trial_start, trial_end,
                        current_period_start, current_period_end,
                        canceled_at, updated_at
                    ) VALUES ($1,$2,$3,$4,$5,$6,$7,'stripe',$8,$9,$10,$11,$12,NOW())
                    ON CONFLICT (stripe_subscription_id) DO UPDATE SET
                        status = EXCLUDED.status,
                        plan_interval = EXCLUDED.plan_interval,
                        plan_amount = EXCLUDED.plan_amount,
                        email = EXCLUDED.email,
                        trial_start = EXCLUDED.trial_start,
                        trial_end = EXCLUDED.trial_end,
                        current_period_start = EXCLUDED.current_period_start,
                        current_period_end = EXCLUDED.current_period_end,
                        canceled_at = EXCLUDED.canceled_at,
                        updated_at = NOW()
                """,
                    customer_id, sub_id, email, status,
                    plan_interval, plan_amount, sub.get("currency", "usd"),
                    _ts(sub.get("trial_start")), _ts(sub.get("trial_end")),
                    _ts(sub.get("current_period_start")), _ts(sub.get("current_period_end")),
                    _ts(sub.get("canceled_at"))
                )
            except Exception as e:
                print(f"Subscription upsert error: {e}")

            # S26 Phase 2: Period-end cancellation semantics
            # Stripe model: cancel click sets cancel_at_period_end=true, status stays active.
            # subscription.deleted only fires when the period actually ends (or immediate cancel).
            # We mirror that: track cancel intent in cancel_state='pending', only mark canceled when expired.
            try:
                cape = sub.get("cancel_at_period_end", False)
                period_end_ts = sub.get("current_period_end")
                now_utc_ts = int(datetime.now(timezone.utc).timestamp())

                if event_type == "customer.subscription.deleted":
                    # If period_end is still in the future, this is a click-cancel. Hold in pending.
                    if period_end_ts and period_end_ts > now_utc_ts:
                        await conn.execute(
                            """UPDATE subscriptions
                               SET status = CASE WHEN status = 'canceled' THEN 'active' ELSE status END,
                                   canceled_at = NULL,
                                   cancel_state = 'pending',
                                   pending_cancel_at = $1,
                                   cancel_requested_at = COALESCE(cancel_requested_at, NOW()),
                                   updated_at = NOW()
                               WHERE stripe_subscription_id = $2""",
                            datetime.fromtimestamp(period_end_ts, tz=timezone.utc), sub_id
                        )
                        print(f"[S26] Stripe cancel held in grace until {period_end_ts}: {sub_id}")
                    else:
                        # Period truly expired (or no period_end) — full cancel.
                        await conn.execute(
                            """UPDATE subscriptions
                               SET cancel_state = 'expired', updated_at = NOW()
                               WHERE stripe_subscription_id = $1""",
                            sub_id
                        )
                elif event_type == "customer.subscription.updated":
                    if cape and status in ("active", "trialing"):
                        # User clicked cancel; sub still active until period_end.
                        await conn.execute(
                            """UPDATE subscriptions
                               SET cancel_state = 'pending',
                                   pending_cancel_at = $1,
                                   cancel_requested_at = COALESCE(cancel_requested_at, NOW()),
                                   updated_at = NOW()
                               WHERE stripe_subscription_id = $2""",
                            datetime.fromtimestamp(period_end_ts, tz=timezone.utc) if period_end_ts else None, sub_id
                        )
                    elif (not cape) and status in ("active", "trialing"):
                        # Uncancel: user reversed their cancellation while still in grace.
                        await conn.execute(
                            """UPDATE subscriptions
                               SET cancel_state = NULL,
                                   pending_cancel_at = NULL,
                                   cancel_requested_at = NULL,
                                   updated_at = NOW()
                               WHERE stripe_subscription_id = $1
                                 AND cancel_state = 'pending'""",
                            sub_id
                        )
            except Exception as e:
                print(f"[S26] Period-end cancel handler error: {e}")

            # Session 11: Assign readable_id if not yet set
            try:
                existing = await conn.fetchrow(
                    "SELECT readable_id FROM subscriptions WHERE stripe_subscription_id = $1",
                    sub_id
                )
                if existing and not existing["readable_id"]:
                    rid = await assign_readable_id(conn, "stripe")
                    await conn.execute(
                        "UPDATE subscriptions SET readable_id = $1 WHERE stripe_subscription_id = $2",
                        rid, sub_id
                    )
            except Exception as e:
                print(f"Readable ID assign error: {e}")

            # Session 11: Track renewals (invoice.paid-like events)
            if event_type == "customer.subscription.updated" and status == "active":
                prev_attrs = event.get("data", {}).get("previous_attributes", {})
                prev_period = prev_attrs.get("current_period_start")
                if prev_period and prev_period != sub.get("current_period_start"):
                    try:
                        await conn.execute(
                            """UPDATE subscriptions SET
                               renewal_count = COALESCE(renewal_count, 0) + 1,
                               last_renewed_at = NOW()
                               WHERE stripe_subscription_id = $1""",
                            sub_id
                        )
                    except Exception:
                        pass

            # Session 11: Detect trial-to-paid conversion (trialing -> active)
            if event_type == "customer.subscription.updated":
                prev_attrs = event.get("data", {}).get("previous_attributes", {})
                if prev_attrs.get("status") == "trialing" and status == "active":
                    try:
                        await conn.execute(
                            "UPDATE subscriptions SET converted_at = NOW() WHERE stripe_subscription_id = $1 AND converted_at IS NULL",
                            sub_id
                        )
                        print(f"Trial conversion stamped for {sub_id}")
                    except Exception as e:
                        print(f"Conversion stamp error: {e}")

        # Handle one-time payments (checkout.session.completed for initial signup)
        elif event_type == "checkout.session.completed":
            session = data_obj
            # If this checkout created a subscription, it'll be handled by subscription.created
            # Log it for analytics
            print(f"Checkout completed: {session.get('id')}, customer: {session.get('customer')}")

    return {"status": "ok"}


# --- Apple App Store Server Notifications v2 (STUBBED S22) ---
# Full handler preserved in git history (commit d51fae9, pre-S22 cleanup).
# ymove receives Apple webhooks and forwards to /webhooks/ymove.
# Endpoint kept alive so Apple doesn't 404 if it ever sends here directly.

@app.post("/webhooks/apple")
async def apple_webhook(request: Request):
    """Stub: Apple webhooks are received by ymove, not directly."""
    try:
        body = await request.json()
    except Exception:
        body = {}
    if db_pool:
        try:
            async with db_pool.acquire() as conn:
                await conn.execute(
                    "INSERT INTO ymove_webhook_log (event_type, payload) VALUES ($1, $2)",
                    "apple.direct_stub", json.dumps(body, default=str)
                )
        except Exception:
            pass
    print(f"[Apple webhook] Received (stubbed): {json.dumps(body, default=str)[:200]}")
    return {"status": "ok", "note": "stubbed - ymove handles Apple webhooks"}


# --- Google Play Real-Time Developer Notifications (STUBBED S22) ---
# Full handler preserved in git history (commit d51fae9, pre-S22 cleanup).
# ymove receives Google webhooks and forwards to /webhooks/ymove.
# Endpoint kept alive so Google doesn't 404 if it ever sends here directly.

@app.post("/webhooks/google")
async def google_webhook(request: Request):
    """Stub: Google webhooks are received by ymove, not directly."""
    try:
        body = await request.json()
    except Exception:
        body = {}
    if db_pool:
        try:
            async with db_pool.acquire() as conn:
                await conn.execute(
                    "INSERT INTO ymove_webhook_log (event_type, payload) VALUES ($1, $2)",
                    "google.direct_stub", json.dumps(body, default=str)
                )
        except Exception:
            pass
    print(f"[Google webhook] Received (stubbed): {json.dumps(body, default=str)[:200]}")
    return {"status": "ok", "note": "stubbed - ymove handles Google webhooks"}


# --- ymove Webhook (Session 16 â Phase 2 Processor) ---

@app.post("/webhooks/ymove")
async def ymove_webhook(request: Request):
    """Receive subscription events from ymove Actions system.
    Phase 2: Process into subscriptions table with email attribution."""
    try:
        body = await request.json()
    except Exception:
        body = {"raw": (await request.body()).decode("utf-8", errors="replace")[:5000]}

    # Extract event data
    event_data = body.get("event", {})
    user_data = body.get("user", {})

    event_type = event_data.get("type", "unknown")
    category = event_data.get("category", "")

    # User info â available on ALL events (key advantage over direct Apple/Google webhooks)
    email = user_data.get("email", "")
    first_name = user_data.get("firstName", "")
    last_name = user_data.get("lastName", "")
    ymove_user_id = str(user_data.get("id", ""))

    provider = event_data.get("subscriptionPaymentProvider", "")
    print(f"[ymove] {event_type} | {email} | provider={provider or 'n/a'}")

    # Always log raw payload
    if db_pool:
        try:
            async with db_pool.acquire() as conn:
                await conn.execute(
                    "INSERT INTO ymove_webhook_log (event_type, payload) VALUES ($1, $2)",
                    event_type, json.dumps(body, default=str)
                )
        except Exception as e:
            print(f"[ymove] Log store error: {e}")

    if not db_pool or category != "subscription":
        return {"status": "ok", "received": event_type}

    async with db_pool.acquire() as conn:
        if event_type == "subscriptionCreated":
            await _ymove_handle_created(conn, event_data, email, ymove_user_id, provider, first_name, last_name)
        elif event_type == "subscriptionCancelled":
            await _ymove_handle_cancelled(conn, event_data, email, ymove_user_id)
        else:
            print(f"[ymove] Unhandled event type: {event_type}")

    return {"status": "ok", "processed": event_type, "email": email}


async def _ymove_lookup_utm(email: str, ymove_user_id: str = "") -> dict:
    """S21: Query ymove Member Lookup API for a user's meta parameters (UTM attribution).
    Returns dict with extracted UTM fields + raw meta response.
    Tosh confirmed: new checkout (ymove.app/join/movementandmiles) stores UTM params
    in the user's meta field. Only applies to Stripe signups (Apple/Google go through
    app stores which strip UTMs)."""
    result = {"found": False, "utm": {}, "raw_meta": None}
    if not YMOVE_API_KEY or not email:
        return result

    UTM_KEYS = ['utm_source', 'utm_medium', 'utm_campaign', 'utm_term', 'utm_content', 'ym_source',
                'utm_id', 'gclid', 'fbclid', 'ttclid', 'msclkid', 'twclid', 'ref', 'referrer']
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.get(
                f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup",
                headers={"X-Authorization": YMOVE_API_KEY},
                params={"email": email.strip().lower()}
            )
            if resp.status_code != 200:
                print(f"[ymove-utm] Lookup failed for {email}: HTTP {resp.status_code}")
                return result

            data = resp.json()
            if not data.get("found"):
                print(f"[ymove-utm] User not found: {email}")
                return result

            user = data.get("user", {})
            result["found"] = True

            # Extract meta parameters -- try common field names
            meta = user.get("meta") or user.get("metadata") or user.get("metaParameters") or user.get("params") or {}
            if isinstance(meta, str):
                try:
                    meta = json.loads(meta)
                except Exception:
                    meta = {}

            result["raw_meta"] = meta
            print(f"[ymove-utm] Raw meta for {email}: {json.dumps(meta, default=str)[:500]}")

            # Extract known UTM fields
            if isinstance(meta, dict):
                for key in UTM_KEYS:
                    val = meta.get(key, "")
                    if val:
                        result["utm"][key] = str(val)

            # Also check top-level user fields in case meta is nested differently
            for key in UTM_KEYS:
                if key not in result["utm"]:
                    val = user.get(key, "")
                    if val:
                        result["utm"][key] = str(val)

            if result["utm"]:
                print(f"[ymove-utm] Found UTMs for {email}: {result['utm']}")
            else:
                print(f"[ymove-utm] No UTM data found for {email} (meta keys: {list(meta.keys()) if isinstance(meta, dict) else 'not a dict'})")

    except Exception as e:
        print(f"[ymove-utm] Lookup error for {email}: {e}")

    return result


async def _store_utm_on_subscription(conn, stripe_sub_id: str, utm_result: dict):
    """S21: Store UTM attribution data on a subscription record."""
    utm = utm_result.get("utm", {})
    raw_meta = utm_result.get("raw_meta")

    if not utm and not raw_meta:
        return

    try:
        await conn.execute("""
            UPDATE subscriptions SET
                utm_source = COALESCE(NULLIF($1, ''), utm_source),
                utm_medium = COALESCE(NULLIF($2, ''), utm_medium),
                utm_campaign = COALESCE(NULLIF($3, ''), utm_campaign),
                utm_term = COALESCE(NULLIF($4, ''), utm_term),
                utm_content = COALESCE(NULLIF($5, ''), utm_content),
                ym_source = COALESCE(NULLIF($6, ''), ym_source),
                utm_meta_raw = $7,
                updated_at = NOW()
            WHERE stripe_subscription_id = $8
        """,
            utm.get("utm_source", ""),
            utm.get("utm_medium", ""),
            utm.get("utm_campaign", ""),
            utm.get("utm_term", ""),
            utm.get("utm_content", ""),
            utm.get("ym_source", ""),
            json.dumps(raw_meta, default=str) if raw_meta else None,
            stripe_sub_id
        )
        print(f"[ymove-utm] Stored UTM data on sub {stripe_sub_id}")
    except Exception as e:
        print(f"[ymove-utm] Store error for {stripe_sub_id}: {e}")


async def _ymove_handle_created(conn, event_data: dict, email: str, ymove_user_id: str, provider: str, first_name: str = "", last_name: str = ""):
    """Process subscriptionCreated from ymove."""
    now_ts = int(datetime.now(timezone.utc).timestamp())

    if provider == "stripe":
        # Stripe subs already arrive via direct Stripe webhook.
        # ymove adds email + name â attach email if missing.
        stripe_sub_id = event_data.get("stripeSubscriptionId", "")
        if not stripe_sub_id:
            return

        # Store event for audit
        eid = f"ymove_created_{stripe_sub_id}_{now_ts}"
        try:
            await conn.execute(
                """INSERT INTO subscription_events (stripe_event_id, event_type, stripe_subscription_id, source, data)
                   VALUES ($1, $2, $3, 'stripe', $4)
                   ON CONFLICT (stripe_event_id) DO NOTHING""",
                eid, "ymove.subscriptionCreated", stripe_sub_id,
                json.dumps({"ymove_user_id": ymove_user_id, "email": email, "provider": provider}, default=str)
            )
        except Exception as e:
            print(f"[ymove] Stripe event store error: {e}")

        # Attach email if subscription exists but has no email
        if email:
            try:
                existing = await conn.fetchrow(
                    "SELECT id, email FROM subscriptions WHERE stripe_subscription_id = $1",
                    stripe_sub_id
                )
                if existing and not existing["email"]:
                    await conn.execute(
                        "UPDATE subscriptions SET email = $1, updated_at = NOW() WHERE stripe_subscription_id = $2",
                        email, stripe_sub_id
                    )
                    print(f"[ymove] Attached email {email} to Stripe sub {stripe_sub_id}")
                # S17: Store names
                if existing and (first_name or last_name):
                    try:
                        await conn.execute(
                            "UPDATE subscriptions SET first_name = COALESCE(NULLIF($1, ''), first_name), last_name = COALESCE(NULLIF($2, ''), last_name) WHERE stripe_subscription_id = $3",
                            first_name, last_name, stripe_sub_id
                        )
                    except Exception:
                        pass
            except Exception as e:
                print(f"[ymove] Stripe email attach error: {e}")

        # S21: Look up UTM attribution from ymove meta parameters
        if email:
            try:
                utm_result = await _ymove_lookup_utm(email, ymove_user_id)
                if utm_result["found"]:
                    await _store_utm_on_subscription(conn, stripe_sub_id, utm_result)
            except Exception as e:
                print(f"[ymove-utm] Attribution lookup error: {e}")

    elif provider == "apple":
        # Apple subs: ymove is our PRIMARY source (no direct Apple webhook connected)
        transaction_id = event_data.get("transactionId", "")
        if not transaction_id:
            return

        product_id = event_data.get("productId", "")
        start_str = event_data.get("startDate", "")
        end_str = event_data.get("endDate", "")

        # Determine plan from productId
        plan_interval = "month"
        plan_amount = 1999
        if product_id and ("annual" in product_id.lower() or "year" in product_id.lower()):
            plan_interval = "year"
            plan_amount = 17999

        # Parse ISO dates

        period_start = _parse_iso(start_str)
        period_end = _parse_iso(end_str)

        # Store event
        eid = f"ymove_created_{transaction_id}_{now_ts}"
        try:
            await conn.execute(
                """INSERT INTO subscription_events (stripe_event_id, event_type, stripe_subscription_id, source, data)
                   VALUES ($1, $2, $3, 'apple', $4)
                   ON CONFLICT (stripe_event_id) DO NOTHING""",
                eid, "ymove.subscriptionCreated", transaction_id,
                json.dumps({"ymove_user_id": ymove_user_id, "email": email, "product_id": product_id}, default=str)
            )
        except Exception as e:
            print(f"[ymove] Apple event store error: {e}")

        # Upsert subscription â email + period dates are the key value from ymove
        try:
            await conn.execute("""
                INSERT INTO subscriptions (
                    stripe_customer_id, stripe_subscription_id, email, status,
                    plan_interval, plan_amount, currency, source,
                    current_period_start, current_period_end,
                    trial_start, trial_end,
                    created_at, updated_at
                ) VALUES ('', $1, $2, 'active', $3, $4, 'usd', 'apple', $5, $6, $5, $6, COALESCE($5, NOW()), NOW())
                ON CONFLICT (stripe_subscription_id) DO UPDATE SET
                    email = COALESCE(NULLIF(EXCLUDED.email, ''), subscriptions.email),
                    status = 'active',
                    plan_interval = EXCLUDED.plan_interval,
                    plan_amount = EXCLUDED.plan_amount,
                    current_period_start = COALESCE(EXCLUDED.current_period_start, subscriptions.current_period_start),
                    current_period_end = COALESCE(EXCLUDED.current_period_end, subscriptions.current_period_end),
                    updated_at = NOW()
            """,
                transaction_id, email, plan_interval, plan_amount,
                period_start, period_end
            )
            print(f"[ymove] Apple sub upserted: {transaction_id} | {email}")
        except Exception as e:
            print(f"[ymove] Apple sub upsert error: {e}")

        # S17: Store names from ymove
        if first_name or last_name:
            try:
                await conn.execute(
                    "UPDATE subscriptions SET first_name = COALESCE(NULLIF($1, ''), first_name), last_name = COALESCE(NULLIF($2, ''), last_name) WHERE stripe_subscription_id = $3",
                    first_name, last_name, transaction_id
                )
            except Exception as e:
                print(f"[ymove] Apple name store error: {e}")

        # S18: Dedup - deactivate synthetic Meg-import record if real transaction arrived
        if email:
            try:
                synthetic = await conn.fetch(
                    """SELECT id, stripe_subscription_id FROM subscriptions
                       WHERE lower(email) = lower($1)
                       AND stripe_subscription_id LIKE 'meg_apple_%'
                       AND stripe_subscription_id != $2
                       AND status IN ('active', 'trialing')""",
                    email, transaction_id
                )
                for syn in synthetic:
                    await conn.execute(
                        "UPDATE subscriptions SET status = 'canceled', canceled_at = NOW(), updated_at = NOW() WHERE id = $1",
                        syn["id"]
                    )
                    print(f"[ymove] Dedup: deactivated synthetic {syn['stripe_subscription_id']} for {email} (real: {transaction_id})")
            except Exception as e:
                print(f"[ymove] Apple dedup error: {e}")

        # S18: Detect trial-to-paid conversion (trial expired + still active + no cancel)
        try:
            conv_check = await conn.fetchrow(
                """SELECT id, trial_end FROM subscriptions
                   WHERE stripe_subscription_id = $1
                   AND converted_at IS NULL
                   AND status = 'active'
                   AND trial_end IS NOT NULL
                   AND trial_end < NOW()
                   AND canceled_at IS NULL""",
                transaction_id
            )
            if conv_check:
                await conn.execute(
                    "UPDATE subscriptions SET converted_at = $1 WHERE id = $2",
                    conv_check["trial_end"], conv_check["id"]
                )
                print(f"[ymove] Apple conversion detected for {transaction_id} (trial_end: {conv_check['trial_end']})")
        except Exception as e:
            print(f"[ymove] Apple conversion check error: {e}")

        # Assign readable_id if not yet set
        try:
            existing = await conn.fetchrow(
                "SELECT readable_id FROM subscriptions WHERE stripe_subscription_id = $1",
                transaction_id
            )
            if existing and not existing["readable_id"]:
                rid = await assign_readable_id(conn, "apple")
                await conn.execute(
                    "UPDATE subscriptions SET readable_id = $1 WHERE stripe_subscription_id = $2",
                    rid, transaction_id
                )
        except Exception as e:
            print(f"[ymove] Apple readable ID error: {e}")

    elif provider == "google":
        # Google subs â same pattern as Apple, using ymove uuid as ID
        ym_uuid = event_data.get("uuid", "")
        if not ym_uuid:
            return
        external_id = f"ym_google_{ym_uuid}"

        product_id = event_data.get("productId", "")
        start_str = event_data.get("startDate", "")
        end_str = event_data.get("endDate", "")

        plan_interval = "month"
        plan_amount = 1999
        if product_id and ("annual" in product_id.lower() or "year" in product_id.lower()):
            plan_interval = "year"
            plan_amount = 17999


        period_start = _parse_iso(start_str)
        period_end = _parse_iso(end_str)

        eid = f"ymove_created_{ym_uuid}_{now_ts}"
        try:
            await conn.execute(
                """INSERT INTO subscription_events (stripe_event_id, event_type, stripe_subscription_id, source, data)
                   VALUES ($1, $2, $3, 'google', $4)
                   ON CONFLICT (stripe_event_id) DO NOTHING""",
                eid, "ymove.subscriptionCreated", external_id,
                json.dumps({"ymove_user_id": ymove_user_id, "email": email, "product_id": product_id}, default=str)
            )
        except Exception as e:
            print(f"[ymove] Google event store error: {e}")

        try:
            await conn.execute("""
                INSERT INTO subscriptions (
                    stripe_customer_id, stripe_subscription_id, email, status,
                    plan_interval, plan_amount, currency, source,
                    current_period_start, current_period_end,
                    created_at, updated_at
                ) VALUES ('', $1, $2, 'active', $3, $4, 'usd', 'google', $5, $6, COALESCE($5, NOW()), NOW())
                ON CONFLICT (stripe_subscription_id) DO UPDATE SET
                    email = COALESCE(NULLIF(EXCLUDED.email, ''), subscriptions.email),
                    status = 'active',
                    updated_at = NOW()
            """,
                external_id, email, plan_interval, plan_amount,
                period_start, period_end
            )
            print(f"[ymove] Google sub upserted: {external_id} | {email}")
        except Exception as e:
            print(f"[ymove] Google sub upsert error: {e}")

        # S17: Store names from ymove
        if first_name or last_name:
            try:
                await conn.execute(
                    "UPDATE subscriptions SET first_name = COALESCE(NULLIF($1, ''), first_name), last_name = COALESCE(NULLIF($2, ''), last_name) WHERE stripe_subscription_id = $3",
                    first_name, last_name, external_id
                )
            except Exception as e:
                print(f"[ymove] Google name store error: {e}")

        # S18: Dedup - deactivate synthetic Meg-import record if real transaction arrived
        if email:
            try:
                synthetic = await conn.fetch(
                    """SELECT id, stripe_subscription_id FROM subscriptions
                       WHERE lower(email) = lower($1)
                       AND stripe_subscription_id LIKE 'meg_google_%'
                       AND stripe_subscription_id != $2
                       AND status IN ('active', 'trialing')""",
                    email, external_id
                )
                for syn in synthetic:
                    await conn.execute(
                        "UPDATE subscriptions SET status = 'canceled', canceled_at = NOW(), updated_at = NOW() WHERE id = $1",
                        syn["id"]
                    )
                    print(f"[ymove] Dedup: deactivated synthetic {syn['stripe_subscription_id']} for {email} (real: {external_id})")
            except Exception as e:
                print(f"[ymove] Google dedup error: {e}")

        # S18: Detect trial-to-paid conversion (trial expired + still active + no cancel)
        try:
            conv_check = await conn.fetchrow(
                """SELECT id, trial_end FROM subscriptions
                   WHERE stripe_subscription_id = $1
                   AND converted_at IS NULL
                   AND status = 'active'
                   AND trial_end IS NOT NULL
                   AND trial_end < NOW()
                   AND canceled_at IS NULL""",
                external_id
            )
            if conv_check:
                await conn.execute(
                    "UPDATE subscriptions SET converted_at = $1 WHERE id = $2",
                    conv_check["trial_end"], conv_check["id"]
                )
                print(f"[ymove] Google conversion detected for {external_id} (trial_end: {conv_check['trial_end']})")
        except Exception as e:
            print(f"[ymove] Google conversion check error: {e}")

        # Assign readable_id
        try:
            existing = await conn.fetchrow(
                "SELECT readable_id FROM subscriptions WHERE stripe_subscription_id = $1",
                external_id
            )
            if existing and not existing["readable_id"]:
                rid = await assign_readable_id(conn, "google")
                await conn.execute(
                    "UPDATE subscriptions SET readable_id = $1 WHERE stripe_subscription_id = $2",
                    rid, external_id
                )
        except Exception as e:
            print(f"[ymove] Google readable ID error: {e}")

    else:
        print(f"[ymove] Unknown provider: {provider}")


async def _ymove_handle_cancelled(conn, event_data: dict, email: str, ymove_user_id: str):
    """Process subscriptionCancelled from ymove.
    Cancel events have no provider field â we find the sub by email."""
    now_ts = int(datetime.now(timezone.utc).timestamp())

    # Store event for audit
    eid = f"ymove_cancelled_{ymove_user_id}_{now_ts}"
    try:
        await conn.execute(
            """INSERT INTO subscription_events (stripe_event_id, event_type, stripe_customer_id, stripe_subscription_id, source, data)
               VALUES ($1, $2, '', '', 'ymove', $3)
               ON CONFLICT (stripe_event_id) DO NOTHING""",
            eid, "ymove.subscriptionCancelled",
            json.dumps({"ymove_user_id": ymove_user_id, "email": email, "event": event_data}, default=str)
        )
    except Exception as e:
        print(f"[ymove] Cancel event store error: {e}")

    if not email:
        print("[ymove] Cancel event has no email â cannot match subscription")
        return

    # Find most recent active sub for this email and cancel it
    # For Stripe subs, the direct Stripe webhook will also fire â double-cancel is safe (no-op)
    # For Apple subs, this is the ONLY cancel signal we get
    try:
        active_sub = await conn.fetchrow(
            """SELECT id, stripe_subscription_id, source FROM subscriptions
               WHERE lower(email) = lower($1) AND status IN ('active', 'trialing')
               ORDER BY created_at DESC LIMIT 1""",
            email
        )
        if active_sub:
            # S26 Phase 3: For Stripe records, check period_end before cancelling.
            # Apple/Google use ymove's period-end semantics already — trust ymove for those.
            src = (active_sub.get("source") or "").lower()
            sub_id_check = active_sub.get("stripe_subscription_id") or ""
            handled_pending = False

            if src == "stripe" and sub_id_check.startswith("sub_") and STRIPE_SECRET_KEY:
                try:
                    import stripe as _s26_stripe
                    _s26_stripe.api_key = STRIPE_SECRET_KEY
                    _real = _s26_stripe.Subscription.retrieve(sub_id_check)
                    _pe = _real.get("current_period_end")
                    _now_ts = int(datetime.now(timezone.utc).timestamp())
                    if _pe and _pe > _now_ts:
                        await conn.execute(
                            """UPDATE subscriptions
                               SET cancel_state = 'pending',
                                   pending_cancel_at = $1,
                                   cancel_requested_at = COALESCE(cancel_requested_at, NOW()),
                                   updated_at = NOW()
                               WHERE id = $2""",
                            datetime.fromtimestamp(_pe, tz=timezone.utc), active_sub["id"]
                        )
                        handled_pending = True
                        print(f"[S26 ymove] Held Stripe cancel in grace for {email}")
                except Exception as _e:
                    print(f"[S26 ymove] Period-end check error for {sub_id_check}: {_e}")

            if not handled_pending:
                await conn.execute(
                    """UPDATE subscriptions SET status = 'canceled', canceled_at = NOW(),
                       cancel_state = 'expired', updated_at = NOW() WHERE id = $1""",
                    active_sub["id"]
                )
            print(f"[ymove] Cancelled {active_sub['source']} sub {active_sub['stripe_subscription_id']} for {email}")
        else:
            print(f"[ymove] No active sub found for {email} to cancel")
    except Exception as e:
        print(f"[ymove] Cancel processing error: {e}")


# --- Daily Digest System ---

async def gather_daily_stats() -> dict:
    """Query database for last 24 hours of activity."""
    if not db_pool:
        return {"error": "No database"}

    async with db_pool.acquire() as conn:
        # New subscriptions in last 24h
        new_subs = await conn.fetch(
            """SELECT source, status, plan_interval, plan_amount, email, created_at
               FROM subscriptions WHERE created_at > NOW() - INTERVAL '24 hours'
               AND created_at <= NOW()
               ORDER BY created_at DESC"""
        )
        new_subs_by_source = await conn.fetch(
            """SELECT source, COUNT(*) as count FROM subscriptions
               WHERE created_at > NOW() - INTERVAL '24 hours'
               AND created_at <= NOW()
               GROUP BY source ORDER BY count DESC"""
        )

        # Conversions (trial â paid) in last 24h
        conversions_today = await conn.fetch(
            """SELECT source, email, plan_interval, plan_amount, converted_at
               FROM subscriptions WHERE converted_at > NOW() - INTERVAL '24 hours'
               ORDER BY converted_at DESC"""
        )

        # Cancellations in last 24h (S22: include converted_at to distinguish paid vs trial)
        cancellations = await conn.fetch(
            """SELECT source, email, canceled_at, converted_at, trial_end FROM subscriptions
               WHERE canceled_at > NOW() - INTERVAL '24 hours'
               ORDER BY canceled_at DESC"""
        )

        # Current MRR
        mrr_monthly = await conn.fetchval(
            "SELECT COALESCE(SUM(plan_amount), 0) FROM subscriptions WHERE status = 'active' AND plan_interval = 'month'"
        )
        mrr_annual = await conn.fetchval(
            "SELECT COALESCE(SUM(plan_amount / 12), 0) FROM subscriptions WHERE status = 'active' AND plan_interval = 'year'"
        )
        current_mrr_cents = (mrr_monthly or 0) + (mrr_annual or 0)

        # MRR by source (for fee calc)
        mrr_by_source = await conn.fetch(
            """SELECT source,
                COALESCE(SUM(CASE WHEN plan_interval='month' THEN plan_amount ELSE 0 END), 0) +
                COALESCE(SUM(CASE WHEN plan_interval='year' THEN plan_amount/12 ELSE 0 END), 0) as mrr_cents
               FROM subscriptions WHERE status = 'active' GROUP BY source"""
        )

        # Active subs count
        active_count = await conn.fetchval(
            "SELECT COUNT(*) FROM subscriptions WHERE status IN ('active', 'trialing')"
        )
        trialing_count = await conn.fetchval(
            "SELECT COUNT(*) FROM subscriptions WHERE status = 'trialing'"
        )
        total_count = await conn.fetchval("SELECT COUNT(*) FROM subscriptions")

        # New leads in last 24h
        new_leads = await conn.fetch(
            """SELECT first_name, email, utm_source, utm_medium, utm_campaign, created_at
               FROM leads WHERE created_at > NOW() - INTERVAL '24 hours'
               ORDER BY created_at DESC"""
        )
        leads_by_source_24h = await conn.fetch(
            """SELECT COALESCE(NULLIF(utm_source,''), 'direct') as source, COUNT(*) as count
               FROM leads WHERE created_at > NOW() - INTERVAL '24 hours'
               GROUP BY source ORDER BY count DESC"""
        )

        # Page views last 24h
        pv_24h = await conn.fetchval(
            "SELECT COUNT(*) FROM page_views WHERE created_at > NOW() - INTERVAL '24 hours'"
        )
        pv_by_page = await conn.fetch(
            """SELECT page, COUNT(*) as count FROM page_views
               WHERE created_at > NOW() - INTERVAL '24 hours'
               GROUP BY page ORDER BY count DESC LIMIT 5"""
        )

        # Chat sessions last 24h
        chats_24h = await conn.fetchval(
            "SELECT COUNT(*) FROM chat_sessions WHERE created_at > NOW() - INTERVAL '24 hours'"
        )

        # Subscription events in last 24h
        recent_events = await conn.fetch(
            """SELECT event_type, source, created_at FROM subscription_events
               WHERE created_at > NOW() - INTERVAL '24 hours'
               ORDER BY created_at DESC LIMIT 30"""
        )

        # S23: Subscription UTM attribution (new subs in last 24h with UTMs)
        subs_by_utm_24h = await conn.fetch(
            """SELECT COALESCE(NULLIF(utm_source,''), 'direct') as channel, COUNT(*) as count
               FROM subscriptions WHERE created_at > NOW() - INTERVAL '24 hours'
               AND created_at <= NOW() AND utm_source != ''
               GROUP BY channel ORDER BY count DESC"""
        )

        # S23: Top traffic source (7d page views by UTM)
        top_traffic_7d = await conn.fetch(
            """SELECT COALESCE(NULLIF(utm_source,''), 'direct') as channel, COUNT(*) as views
               FROM page_views WHERE created_at > NOW() - INTERVAL '7 days'
               AND utm_source != ''
               GROUP BY channel ORDER BY views DESC LIMIT 5"""
        )

    # Calculate fees
    fee_breakdown = {}
    total_fees_cents = 0
    for r in mrr_by_source:
        src = r["source"]
        mrr = r["mrr_cents"] or 0
        if src == "apple":
            fee = round(mrr * 0.15)
        elif src == "google":
            fee = round(mrr * 0.15)
        else:
            fee = round(mrr * 0.029)
        fee_breakdown[src] = fee
        total_fees_cents += fee
    net_mrr_cents = current_mrr_cents - total_fees_cents

    return {
        "conversions_today": len(conversions_today),
        "conversion_details": [{"email": r["email"] or "n/a", "source": r["source"], "plan": f"${(r['plan_amount'] or 0)/100:.2f}/{r['plan_interval'] or '?'}"} for r in conversions_today],
        "new_subscriptions": len(new_subs),
        "new_subs_by_source": [{"source": r["source"], "count": r["count"]} for r in new_subs_by_source],
        "new_sub_details": [{"email": r["email"] or "n/a", "source": r["source"], "plan": f"${(r['plan_amount'] or 0)/100:.2f}/{r['plan_interval'] or '?'}"} for r in new_subs],
        "cancellations": len(cancellations),
        "cancellations_paid": len([c for c in cancellations if c["converted_at"] is not None]),
        "cancellations_trial": len([c for c in cancellations if c["converted_at"] is None]),
        "cancel_details": [{"email": r["email"] or "n/a", "source": r["source"], "type": "paid" if r["converted_at"] is not None else "trial"} for r in cancellations],
        "gross_mrr": f"${current_mrr_cents/100:,.2f}",
        "gross_mrr_cents": current_mrr_cents,
        "net_mrr": f"${net_mrr_cents/100:,.2f}",
        "net_mrr_cents": net_mrr_cents,
        "total_fees": f"${total_fees_cents/100:,.2f}",
        "fee_breakdown": fee_breakdown,
        "active_subscribers": active_count or 0,
        "trialing": trialing_count or 0,
        "total_subscribers": total_count or 0,
        "new_leads": len(new_leads),
        "leads_by_source": [{"source": r["source"], "count": r["count"]} for r in leads_by_source_24h],
        "page_views_24h": pv_24h or 0,
        "top_pages": [{"page": r["page"], "count": r["count"]} for r in pv_by_page],
        "chat_sessions_24h": chats_24h or 0,
        "events_24h": len(recent_events),
        "mrr_by_source": [{"source": r["source"], "mrr_cents": r["mrr_cents"] or 0} for r in mrr_by_source],
        "subs_by_utm_24h": [{"channel": r["channel"], "count": r["count"]} for r in subs_by_utm_24h],
        "top_traffic_7d": [{"channel": r["channel"], "views": r["views"]} for r in top_traffic_7d],
    }


async def generate_digest_insights(stats: dict) -> str:
    """Send daily stats to Claude for analysis and insights."""
    conversions = stats.get("conversions_today", 0)
    paid_cancels = stats.get("cancellations_paid", 0)
    trial_starts = stats.get("new_subscriptions", 0)
    trial_cancels = stats.get("cancellations_trial", 0)

    slim = {
        "net_paid_growth": conversions - paid_cancels,
        "net_trial_growth": trial_starts - trial_cancels,
        "conversions": conversions,
        "trial_starts": trial_starts,
        "paid_cancels": paid_cancels,
        "trial_cancels": trial_cancels,
        "gross_mrr": stats.get("gross_mrr", "$0"),
        "net_mrr": stats.get("net_mrr", "$0"),
        "active_subscribers": stats.get("active_subscribers", 0),
        "trialing": stats.get("trialing", 0),
        "new_leads": stats.get("new_leads", 0),
        "page_views_24h": stats.get("page_views_24h", 0),
    }

    # Only include UTM data if present
    utm_subs = stats.get("subs_by_utm_24h", [])
    utm_traffic = stats.get("top_traffic_7d", [])
    if utm_subs:
        slim["subscriber_channels_24h"] = utm_subs
    if utm_traffic:
        slim["top_traffic_source_7d"] = utm_traffic[0]

    stats_text = json.dumps(slim, default=str)
    prompt = f"Today's Movement & Miles metrics:\n{stats_text}"
    print(f"[Digest] Prompt length: {len(prompt)} chars")
    return await call_anthropic_raw(DIGEST_SYSTEM_PROMPT, [{"role": "user", "content": prompt}], max_tokens=400)


def build_digest_html(stats: dict, insights: str) -> str:
    """Build branded HTML email for the daily digest."""
    now_et = datetime.now(ZoneInfo("America/New_York"))
    date_str = now_et.strftime("%A, %B %d, %Y")

    # Build new subs detail rows
    sub_rows = ""
    for s in stats.get("new_sub_details", [])[:10]:
        sub_rows += f'<tr><td style="padding:6px 12px;font-size:14px;border-bottom:1px solid #e0e0e0">{s["email"]}</td><td style="padding:6px 12px;font-size:14px;border-bottom:1px solid #e0e0e0">{s["source"]}</td><td style="padding:6px 12px;font-size:14px;border-bottom:1px solid #e0e0e0">{s["plan"]}</td></tr>'

    cancel_rows = ""
    for c in stats.get("cancel_details", [])[:10]:
        ctype = c.get("type", "unknown")
        type_color = "#c0392b" if ctype == "paid" else "#b35a00"
        type_label = "PAID" if ctype == "paid" else "TRIAL"
        cancel_rows += f'<tr><td style="padding:6px 12px;font-size:14px;border-bottom:1px solid #e0e0e0">{c["email"]}</td><td style="padding:6px 12px;font-size:14px;border-bottom:1px solid #e0e0e0">{c["source"]}</td><td style="padding:6px 12px;font-size:14px;border-bottom:1px solid #e0e0e0;color:{type_color};font-weight:600">{type_label}</td></tr>'

    lead_source_rows = ""
    for ls in stats.get("leads_by_source", []):
        lead_source_rows += f'<tr><td style="padding:6px 12px;font-size:14px;border-bottom:1px solid #e0e0e0">{ls["source"]}</td><td style="padding:6px 12px;font-size:14px;border-bottom:1px solid #e0e0e0">{ls["count"]}</td></tr>'

    # Session 11: Convert markdown bold **text** to <strong>text</strong>
    insights_clean = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', insights)
    # Format insights â convert newlines and bullet points to HTML
    insights_html = insights_clean.replace("\n\n", "</p><p style='margin:0 0 12px'>").replace("\n- ", "<br>&#8226; ").replace("\n", "<br>")

    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#f7f7f7;font-family:Arial,Helvetica,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f7f7f7;padding:24px 0">
<tr><td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.08)">

<!-- Header -->
<tr><td style="background:#182241;padding:28px 32px">
  <h1 style="margin:0;color:#ffffff;font-size:22px;font-weight:600">Movement &amp; Miles</h1>
  <p style="margin:6px 0 0;color:rgba(255,255,255,0.7);font-size:13px">Daily Digest &mdash; {date_str}</p>
</td></tr>

<!-- AI Insights -->
<tr><td style="padding:28px 32px 20px">
  <h2 style="margin:0 0 12px;color:#182241;font-size:16px;font-weight:600;text-transform:uppercase;letter-spacing:0.05em">Today&rsquo;s Insights</h2>
  <div style="background:#f0f4f8;border-left:4px solid #182241;border-radius:0 8px 8px 0;padding:16px 20px;font-size:14px;line-height:1.6;color:#333">
    <p style="margin:0 0 12px">{insights_html}</p>
  </div>
</td></tr>

<!-- Key Metrics -->
<tr><td style="padding:0 32px 24px">
  <h2 style="margin:0 0 16px;color:#182241;font-size:16px;font-weight:600;text-transform:uppercase;letter-spacing:0.05em">Revenue Snapshot</h2>
  <table width="100%" cellpadding="0" cellspacing="0">
    <tr>
      <td style="padding:16px;background:#f7f7f7;border-radius:8px;text-align:center;width:25%">
        <div style="font-size:11px;color:#536c7c;font-weight:600;text-transform:uppercase;letter-spacing:0.05em">Gross MRR</div>
        <div style="font-size:24px;font-weight:700;color:#2d6a2d;margin-top:4px">{stats.get('gross_mrr','$0')}</div>
      </td>
      <td width="12"></td>
      <td style="padding:16px;background:#f7f7f7;border-radius:8px;text-align:center;width:25%">
        <div style="font-size:11px;color:#536c7c;font-weight:600;text-transform:uppercase;letter-spacing:0.05em">Net MRR</div>
        <div style="font-size:24px;font-weight:700;color:#2d6a2d;margin-top:4px">{stats.get('net_mrr','$0')}</div>
      </td>
      <td width="12"></td>
      <td style="padding:16px;background:#f7f7f7;border-radius:8px;text-align:center;width:25%">
        <div style="font-size:11px;color:#536c7c;font-weight:600;text-transform:uppercase;letter-spacing:0.05em">Active Subs</div>
        <div style="font-size:24px;font-weight:700;color:#182241;margin-top:4px">{stats.get('active_subscribers',0)}</div>
      </td>
      <td width="12"></td>
      <td style="padding:16px;background:#f7f7f7;border-radius:8px;text-align:center;width:25%">
        <div style="font-size:11px;color:#536c7c;font-weight:600;text-transform:uppercase;letter-spacing:0.05em">Fees / Mo</div>
        <div style="font-size:24px;font-weight:700;color:#b35a00;margin-top:4px">{stats.get('total_fees','$0')}</div>
      </td>
    </tr>
  </table>
</td></tr>

<!-- 24h Activity -->
<tr><td style="padding:0 32px 24px">
  <h2 style="margin:0 0 16px;color:#182241;font-size:16px;font-weight:600;text-transform:uppercase;letter-spacing:0.05em">Last 24 Hours</h2>
  <table width="100%" cellpadding="0" cellspacing="0">
    <tr>
      <td style="padding:12px 16px;background:#e8f4e8;border-radius:8px;text-align:center;width:20%">
        <div style="font-size:28px;font-weight:700;color:#2d6a2d">{stats.get('new_subscriptions',0)}</div>
        <div style="font-size:11px;color:#2d6a2d;font-weight:600;margin-top:2px">Trial Starts</div>
      </td>
      <td width="12"></td>
      <td style="padding:12px 16px;background:#e8f7e8;border-radius:8px;text-align:center;width:20%">
        <div style="font-size:28px;font-weight:700;color:#1b7a1b">{stats.get('conversions_today',0)}</div>
        <div style="font-size:11px;color:#1b7a1b;font-weight:600;margin-top:2px">Paid Conversions</div>
      </td>
      <td width="12"></td>
      <td style="padding:12px 16px;background:#fde8e8;border-radius:8px;text-align:center;width:20%">
        <div style="font-size:28px;font-weight:700;color:#c0392b">{stats.get('cancellations',0)}</div>
        <div style="font-size:11px;color:#c0392b;font-weight:600;margin-top:2px">Cancellations</div>
        <div style="font-size:10px;color:#c0392b;margin-top:2px">{stats.get('cancellations_paid',0)} paid &bull; {stats.get('cancellations_trial',0)} trial</div>
      </td>
      <td width="12"></td>
      <td style="padding:12px 16px;background:#e8eaf6;border-radius:8px;text-align:center;width:20%">
        <div style="font-size:28px;font-weight:700;color:#3949ab">{stats.get('new_leads',0)}</div>
        <div style="font-size:11px;color:#3949ab;font-weight:600;margin-top:2px">New Leads</div>
      </td>
      <td width="12"></td>
      <td style="padding:12px 16px;background:#f7f7f7;border-radius:8px;text-align:center;width:25%">
        <div style="font-size:28px;font-weight:700;color:#182241">{stats.get('page_views_24h',0)}</div>
        <div style="font-size:11px;color:#536c7c;font-weight:600;margin-top:2px">Page Views</div>
      </td>
    </tr>
  </table>
</td></tr>"""

    # New subscriptions detail table (only if there are any)
    if stats.get("new_subscriptions", 0) > 0:
        html += f"""
<!-- New Subscriptions Detail -->
<tr><td style="padding:0 32px 24px">
  <h3 style="margin:0 0 8px;color:#2d6a2d;font-size:14px;font-weight:600">New Subscriptions</h3>
  <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e0e0e0;border-radius:8px;overflow:hidden">
    <tr style="background:#f7f7f7"><th style="padding:8px 12px;font-size:11px;text-align:left;color:#536c7c;font-weight:600;text-transform:uppercase">Email</th><th style="padding:8px 12px;font-size:11px;text-align:left;color:#536c7c;font-weight:600;text-transform:uppercase">Source</th><th style="padding:8px 12px;font-size:11px;text-align:left;color:#536c7c;font-weight:600;text-transform:uppercase">Plan</th></tr>
    {sub_rows}
  </table>
</td></tr>"""

    # Cancellations detail (only if there are any)
    if stats.get("cancellations", 0) > 0:
        html += f"""
<!-- Cancellations Detail -->
<tr><td style="padding:0 32px 24px">
  <h3 style="margin:0 0 8px;color:#c0392b;font-size:14px;font-weight:600">Cancellations ({stats.get('cancellations_paid',0)} paid, {stats.get('cancellations_trial',0)} trial)</h3>
  <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e0e0e0;border-radius:8px;overflow:hidden">
    <tr style="background:#f7f7f7"><th style="padding:8px 12px;font-size:11px;text-align:left;color:#536c7c;font-weight:600;text-transform:uppercase">Email</th><th style="padding:8px 12px;font-size:11px;text-align:left;color:#536c7c;font-weight:600;text-transform:uppercase">Source</th><th style="padding:8px 12px;font-size:11px;text-align:left;color:#536c7c;font-weight:600;text-transform:uppercase">Type</th></tr>
    {cancel_rows}
  </table>
</td></tr>"""

    # Lead sources (only if there are leads)
    if stats.get("new_leads", 0) > 0 and lead_source_rows:
        html += f"""
<!-- Lead Sources -->
<tr><td style="padding:0 32px 24px">
  <h3 style="margin:0 0 8px;color:#3949ab;font-size:14px;font-weight:600">Lead Sources (24h)</h3>
  <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e0e0e0;border-radius:8px;overflow:hidden">
    <tr style="background:#f7f7f7"><th style="padding:8px 12px;font-size:11px;text-align:left;color:#536c7c;font-weight:600;text-transform:uppercase">Source</th><th style="padding:8px 12px;font-size:11px;text-align:left;color:#536c7c;font-weight:600;text-transform:uppercase">Count</th></tr>
    {lead_source_rows}
  </table>
</td></tr>"""

    # S23: Marketing attribution section
    subs_utm = stats.get("subs_by_utm_24h", [])
    top_traffic = stats.get("top_traffic_7d", [])
    if subs_utm or top_traffic:
        marketing_lines = ""
        if subs_utm:
            parts = ", ".join(f"{u['count']} from {u['channel']}" for u in subs_utm)
            marketing_lines += f'<p style="margin:0 0 6px;font-size:14px;color:#333"><strong>New subscribers (24h):</strong> {parts}</p>'
        if top_traffic:
            top = top_traffic[0]
            marketing_lines += f'<p style="margin:0;font-size:14px;color:#333"><strong>Top traffic source (7d):</strong> {top["channel"]} ({top["views"]:,} views)</p>'
        html += f"""
<!-- Marketing Attribution -->
<tr><td style="padding:0 32px 24px">
  <h2 style="margin:0 0 12px;color:#182241;font-size:16px;font-weight:600;text-transform:uppercase;letter-spacing:0.05em">Marketing</h2>
  <div style="background:#f0f4f8;border-radius:8px;padding:16px 20px">
    {marketing_lines}
  </div>
</td></tr>"""

    html += f"""
<!-- Footer -->
<tr><td style="padding:20px 32px 28px;border-top:1px solid #e0e0e0">
  <p style="margin:0;font-size:12px;color:#536c7c;text-align:center">
    Movement &amp; Miles Admin Dashboard &mdash;
    <a href="https://movementmiles2-production.up.railway.app/mm-admin" style="color:#182241;text-decoration:underline">Open Dashboard</a>
  </p>
</td></tr>

</table>
</td></tr>
</table>
</body>
</html>"""

    return html


async def send_digest_email(html: str, subject: str) -> dict:
    """Send email via Resend API."""
    if not RESEND_API_KEY:
        return {"error": "RESEND_API_KEY not configured"}
    if not DIGEST_RECIPIENTS:
        return {"error": "DIGEST_RECIPIENTS not configured"}

    recipients = [e.strip() for e in DIGEST_RECIPIENTS.split(",") if e.strip()]
    if not recipients:
        return {"error": "No valid recipients"}

    async with httpx.AsyncClient(timeout=15.0) as client:
        try:
            resp = await client.post(
                "https://api.resend.com/emails",
                headers={
                    "Authorization": f"Bearer {RESEND_API_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "from": DIGEST_FROM_EMAIL,
                    "to": recipients,
                    "subject": subject,
                    "html": html,
                },
            )
            resp.raise_for_status()
            data = resp.json()
            return {"status": "sent", "id": data.get("id", ""), "recipients": recipients}
        except Exception as e:
            return {"error": f"Resend API error: {str(e)}"}


async def run_daily_digest():
    """Orchestrator: auto-stamp conversions -> gather stats -> AI insights -> build email -> send."""
    print(f"[Digest] Starting daily digest at {datetime.now(ZoneInfo('America/New_York')).strftime('%Y-%m-%d %H:%M ET')}")
    try:
        # S18: Auto-stamp Apple/Google conversions (trial expired + active + no cancel)
        if db_pool:
            try:
                async with db_pool.acquire() as conn:
                    result = await conn.execute(
                        """UPDATE subscriptions SET converted_at = trial_end
                           WHERE converted_at IS NULL
                           AND source IN ('apple', 'google')
                           AND status = 'active'
                           AND trial_end IS NOT NULL
                           AND trial_end < NOW()
                           AND canceled_at IS NULL"""
                    )
                    stamped = int(result.split(" ")[-1]) if result else 0
                    if stamped > 0:
                        print(f"[Digest] Auto-stamped {stamped} Apple/Google conversions")
            except Exception as e:
                print(f"[Digest] Auto-stamp error: {e}")

        stats = await gather_daily_stats()
        if "error" in stats:
            print(f"[Digest] Error gathering stats: {stats['error']}")
            return

        insights = await generate_digest_insights(stats)

        now_et = datetime.now(ZoneInfo("America/New_York"))
        subject = f"M&M Daily Digest - {now_et.strftime('%b %d')} | {stats.get('conversions_today',0)} conversions, {stats.get('new_subscriptions',0)} trials, {stats.get('cancellations_paid',0)} paid cancels, {stats.get('cancellations_trial',0)} trial cancels"

        html = build_digest_html(stats, insights)
        result = await send_digest_email(html, subject)

        if "error" in result:
            print(f"[Digest] Send error: {result['error']}")
        else:
            print(f"[Digest] Sent successfully to {result['recipients']} (id: {result.get('id','')})")
    except Exception as e:
        print(f"[Digest] Unexpected error: {e}")


# --- Admin Endpoints ---

@app.post("/api/admin/login")
async def admin_login(req: LoginRequest):
    if req.password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Invalid password")
    return {"status": "ok"}


@app.post("/api/admin/fix-future-dates")
async def fix_future_dates(request: Request):
    """Fix subscriptions with created_at in the future (e.g. 2026-12-31 from bad import)."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    async with db_pool.acquire() as conn:
        # Find future-dated subs
        future_subs = await conn.fetch(
            "SELECT id, stripe_subscription_id, email, source, created_at FROM subscriptions WHERE created_at > NOW()"
        )
        if not future_subs:
            return {"status": "ok", "fixed": 0, "message": "No future-dated subscriptions found"}

        # Set created_at to their trial_start, or current_period_start, or updated_at as fallback
        fixed = 0
        details = []
        for s in future_subs:
            result = await conn.execute(
                """UPDATE subscriptions SET created_at = COALESCE(trial_start, current_period_start, updated_at, NOW())
                   WHERE id = $1 AND created_at > NOW()""",
                s["id"]
            )
            if result and result.endswith("1"):
                fixed += 1
                details.append({"email": s["email"] or "n/a", "source": s["source"], "old_date": str(s["created_at"])})

    return {"status": "ok", "fixed": fixed, "total_found": len(future_subs), "details": details}


@app.post("/api/admin/smart-backfill-conversions")
async def smart_backfill_conversions(request: Request):
    """S18: Stamp converted_at on Apple/Google subs that survived past their trial.
    Heuristic: trial_end has passed + still active + never cancelled = converted.
    Stamps converted_at = trial_end. Only where converted_at is currently NULL."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    preview = body.get("preview", True)

    async with db_pool.acquire() as conn:
        candidates = await conn.fetch(
            """SELECT id, stripe_subscription_id, email, source, trial_end, created_at
               FROM subscriptions
               WHERE converted_at IS NULL
               AND source IN ('apple', 'google')
               AND status = 'active'
               AND trial_end IS NOT NULL
               AND trial_end < NOW()
               AND canceled_at IS NULL
               ORDER BY source, trial_end"""
        )

        if preview:
            by_source = {}
            in_30d = 0
            for r in candidates:
                src = r["source"]
                by_source[src] = by_source.get(src, 0) + 1
                if r["trial_end"] and (datetime.now(timezone.utc) - r["trial_end"]).days <= 30:
                    in_30d += 1
            return {
                "status": "preview",
                "would_stamp": len(candidates),
                "by_source": by_source,
                "would_show_in_30d_count": in_30d,
                "sample": [
                    {"email": r["email"] or "n/a", "source": r["source"],
                     "trial_end": str(r["trial_end"]), "created_at": str(r["created_at"])}
                    for r in candidates[:20]
                ],
                "note": "These subs are active, past trial, never cancelled. converted_at will be set to trial_end. Check would_show_in_30d_count for dashboard impact."
            }

        stamped = 0
        for r in candidates:
            await conn.execute(
                "UPDATE subscriptions SET converted_at = $1 WHERE id = $2",
                r["trial_end"], r["id"]
            )
            stamped += 1

    return {"status": "ok", "stamped": stamped, "note": "converted_at set to trial_end for active subs that survived past trial."}


@app.post("/api/admin/cleanup-converted-at")
async def cleanup_converted_at(request: Request):
    """S18: NULL out fabricated converted_at on Apple/Google subs from S16 backfill.
    These dates were synthetic (trial_end = created_at + 30d) and pollute conversion metrics."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    preview = body.get("preview", True)

    async with db_pool.acquire() as conn:
        count = await conn.fetchval(
            "SELECT COUNT(*) FROM subscriptions WHERE source IN ('apple', 'google') AND converted_at IS NOT NULL"
        )
        if preview:
            by_source = await conn.fetch(
                """SELECT source, COUNT(*) as cnt FROM subscriptions
                   WHERE source IN ('apple', 'google') AND converted_at IS NOT NULL
                   GROUP BY source"""
            )
            return {
                "status": "preview",
                "would_null": count,
                "by_source": [{"source": r["source"], "count": r["cnt"]} for r in by_source],
                "note": "Send preview: false to execute. This NULLs converted_at on Apple/Google subs where dates were fabricated by S16 backfill."
            }
        result = await conn.execute(
            "UPDATE subscriptions SET converted_at = NULL WHERE source IN ('apple', 'google') AND converted_at IS NOT NULL"
        )
        cleaned = int(result.split(" ")[-1]) if result else 0

    return {"status": "ok", "cleaned": cleaned, "note": "converted_at NULLed on Apple/Google subs. Conversion metrics now reflect Stripe-only data."}


@app.post("/api/admin/backfill-names")
async def backfill_names(request: Request):
    """Backfill first_name/last_name on subscriptions from leads table."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    async with db_pool.acquire() as conn:
        result = await conn.execute("""
            UPDATE subscriptions s SET
                first_name = COALESCE(NULLIF(s.first_name, ''), l.first_name),
                last_name = COALESCE(NULLIF(s.last_name, ''), l.extra)
            FROM leads l
            WHERE lower(s.email) = lower(l.email)
            AND s.email != '' AND l.email != ''
            AND (s.first_name IS NULL OR s.first_name = '' OR s.last_name IS NULL OR s.last_name = '')
        """)
        count = int(result.split(" ")[-1]) if result else 0

    return {"status": "ok", "updated": count}


@app.get("/api/admin/ymove-log")
async def admin_ymove_log(request: Request):
    """View recent ymove webhook payloads for debugging."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")

    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT * FROM ymove_webhook_log ORDER BY created_at DESC LIMIT 50"
        )

    return {"events": [
        {"id": r["id"], "event_type": r["event_type"], "payload": r["payload"], "created_at": str(r["created_at"])}
        for r in rows
    ]}


# --- Session 17: ymove API Verification ---

@app.get("/api/admin/inspect-email")
async def inspect_email(request: Request):
    """S23 debug: dump all subscription records for an email + show why categorized as it was."""
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)
    email = request.query_params.get("email", "").strip().lower()
    if not email:
        return JSONResponse(status_code=400, content={"error": "email required"})
    if not db_pool:
        return JSONResponse(status_code=500, content={"error": "no db"})

    async with db_pool.acquire() as conn:
        records = await conn.fetch(
            """SELECT id, email, stripe_subscription_id, source, status, plan_amount,
                      created_at, canceled_at, updated_at, import_batch
               FROM subscriptions WHERE lower(email) = $1 ORDER BY created_at""",
            email
        )
        in_our_active = await conn.fetchval(
            """SELECT EXISTS(SELECT 1 FROM subscriptions
               WHERE lower(email) = $1 AND status IN ('active','trialing')
               AND source IN ('apple','google','undetermined')
               AND email != '' AND email IS NOT NULL)""", email
        )
        in_active_stripe = await conn.fetchval(
            """SELECT EXISTS(SELECT 1 FROM subscriptions
               WHERE lower(email) = $1 AND status IN ('active','trialing') AND source = 'stripe')""", email
        )
        in_all_known = await conn.fetchval(
            """SELECT EXISTS(SELECT 1 FROM subscriptions
               WHERE lower(email) = $1 AND email != '' AND email IS NOT NULL)""", email
        )
    return {
        "email": email,
        "record_count": len(records),
        "records": [dict(r) for r in records],
        "shadow_sync_categorization": {
            "would_be_in_our_active_set": in_our_active,
            "would_be_in_active_stripe_emails": in_active_stripe,
            "would_be_in_all_known_emails": in_all_known,
            "expected_branch": (
                "unchanged (in our_active_set)" if in_our_active
                else "active_stripe_in_ymove (in active_stripe_emails)" if in_active_stripe
                else "cross_platform_switcher (in all_known_emails)" if in_all_known
                else "truly_new"
            )
        }
    }


@app.get("/api/admin/inspect-ymove-user")
async def inspect_ymove_user(request: Request):
    """S23: Dump the full ymove response for a single email, including meta field.
    Usage: /api/admin/inspect-ymove-user?pw=mmadmin2026&email=user@example.com"""
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)
    email = request.query_params.get("email", "").strip().lower()
    if not email:
        return JSONResponse(status_code=400, content={"error": "email query param required"})
    if not YMOVE_API_KEY:
        return JSONResponse(status_code=500, content={"error": "YMOVE_API_KEY not set"})

    async with httpx.AsyncClient(timeout=15.0) as client:
        resp = await client.get(
            f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup",
            headers={"X-Authorization": YMOVE_API_KEY},
            params={"email": email}
        )
        return {
            "email": email,
            "http_status": resp.status_code,
            "raw_response": resp.json() if resp.status_code == 200 else resp.text
        }


@app.get("/api/admin/provider-test")
async def provider_test_get(request: Request):
    """S23: GET version of provider test - hit this in your browser.
    Compare individual vs bulk ymove API provider fields.
    Uses subscription ID patterns to pick KNOWN-GOOD users from each source."""
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)

    ymove_key = YMOVE_API_KEY
    if not ymove_key:
        return JSONResponse(status_code=500, content={"error": "YMOVE_API_KEY not set"})
    if not db_pool:
        return JSONResponse(status_code=500, content={"error": "No database connected"})

    # Pick KNOWN-GOOD users based on subscription ID pattern (not source label which may be wrong)
    test_emails = {}
    async with db_pool.acquire() as conn:
        # Stripe: sub_* prefix = came from Stripe webhook directly, 100% Stripe
        row = await conn.fetchrow(
            """SELECT email, stripe_subscription_id, source FROM subscriptions
               WHERE stripe_subscription_id LIKE 'sub_%' AND status = 'active'
               AND email != '' AND email IS NOT NULL
               ORDER BY created_at DESC LIMIT 1"""
        )
        if row:
            test_emails["stripe_verified"] = {"email": row["email"], "sub_id": row["stripe_subscription_id"], "source": row["source"], "why": "sub_* prefix = direct Stripe webhook"}

        # Apple: numeric ID = Apple transactionId from ymove webhook with provider=apple
        row = await conn.fetchrow(
            """SELECT email, stripe_subscription_id, source FROM subscriptions
               WHERE stripe_subscription_id ~ '^[0-9]+$' AND status = 'active'
               AND email != '' AND email IS NOT NULL
               ORDER BY created_at DESC LIMIT 1"""
        )
        if row:
            test_emails["apple_verified"] = {"email": row["email"], "sub_id": row["sub_id"] if "sub_id" in row else row["stripe_subscription_id"], "source": row["source"], "why": "numeric ID = Apple transactionId from ymove webhook"}

        # Google: ym_google_* prefix = came from ymove webhook with provider=google
        row = await conn.fetchrow(
            """SELECT email, stripe_subscription_id, source FROM subscriptions
               WHERE stripe_subscription_id LIKE 'ym_google_%' AND status = 'active'
               AND email != '' AND email IS NOT NULL
               ORDER BY created_at DESC LIMIT 1"""
        )
        if row:
            test_emails["google_verified"] = {"email": row["email"], "sub_id": row["stripe_subscription_id"], "source": row["source"], "why": "ym_google_* prefix = ymove webhook with provider=google"}

        # Shadow sync import: ymove_new_* = came from shadow sync (provider unknown)
        row = await conn.fetchrow(
            """SELECT email, stripe_subscription_id, source FROM subscriptions
               WHERE stripe_subscription_id LIKE 'ymove_new_%' AND status = 'active'
               AND email != '' AND email IS NOT NULL
               ORDER BY created_at DESC LIMIT 1"""
        )
        if row:
            test_emails["shadow_sync_import"] = {"email": row["email"], "sub_id": row["stripe_subscription_id"], "source": row["source"], "why": "ymove_new_* = shadow sync import, provider was guessed"}

    results = {}
    async with httpx.AsyncClient(timeout=15.0) as client:
        # Individual lookups for each test user
        for label, info in test_emails.items():
            try:
                resp = await client.get(
                    f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup",
                    headers={"X-Authorization": ymove_key},
                    params={"email": info["email"]}
                )
                raw = resp.json() if resp.status_code == 200 else {"http_status": resp.status_code}
                user_obj = raw.get("user", {})
                results[label] = {
                    "email": info["email"],
                    "our_sub_id": info["sub_id"],
                    "our_source_label": info["source"],
                    "id_pattern_means": info["why"],
                    "ymove_subscriptionProvider": user_obj.get("subscriptionProvider"),
                    "ymove_activeSubscription": user_obj.get("activeSubscription"),
                    "ymove_all_fields": {k: v for k, v in user_obj.items()},
                }
            except Exception as e:
                results[label] = {"email": info["email"], "error": str(e)}
            await asyncio.sleep(1.5)

        # Bulk: scan ALL pages and count how many have non-null provider
        bulk_stats = {"total_users": 0, "null_provider": 0, "non_null_provider": 0, "provider_values": {}, "sample_non_null": [], "pages_scanned": 0}
        try:
            page = 1
            while page <= 30:  # safety cap
                resp = await client.get(
                    f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup/all",
                    headers={"X-Authorization": ymove_key},
                    params={"status": "subscribed", "page": str(page)}
                )
                if resp.status_code == 429:
                    await asyncio.sleep(5)
                    continue
                if resp.status_code != 200:
                    bulk_stats["stopped_at"] = f"page {page}, HTTP {resp.status_code}"
                    break
                data = resp.json()
                users = data.get("users", [])
                if not users:
                    break
                for u in users:
                    bulk_stats["total_users"] += 1
                    prov = u.get("subscriptionProvider")
                    if prov is None or prov == "" or prov == "null":
                        bulk_stats["null_provider"] += 1
                    else:
                        bulk_stats["non_null_provider"] += 1
                        prov_str = str(prov).lower()
                        bulk_stats["provider_values"][prov_str] = bulk_stats["provider_values"].get(prov_str, 0) + 1
                        if len(bulk_stats["sample_non_null"]) < 5:
                            bulk_stats["sample_non_null"].append({"email": u.get("email"), "provider": prov})
                bulk_stats["pages_scanned"] = page
                total_pages = data.get("totalPages", 1)
                if page >= total_pages:
                    break
                page += 1
                await asyncio.sleep(1.0)
        except Exception as e:
            bulk_stats["error"] = str(e)

        results["_bulk_scan"] = bulk_stats

    return {
        "test": "provider_field_comparison_v2",
        "purpose": "Test KNOWN-GOOD users by ID pattern + scan ALL bulk pages for any non-null providers",
        "results": results,
        "conclusion": "If non_null_provider > 0, the API works for some users. If 0, Tosh needs to fix it."
    }


@app.post("/api/admin/ymove-verify")
async def ymove_verify(request: Request):
    """Verify emails against ymove Member Lookup API (S19 rewrite).
    Modes:
      {"smoke_test": true} - test API connectivity with Meg's email
      {"email": "single@ex.com"} - look up one email
      {"emails": [...]} - batch lookup with rate limiting
      {"pull_all_subscribed": true} - pull ALL active members, cross-ref our cancelled subs
      {"provider_test": true} - S23: Compare individual vs bulk API provider fields
    """
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)

    ymove_key = YMOVE_API_KEY
    if not ymove_key:
        return JSONResponse(status_code=500, content={
            "error": "YMOVE_API_KEY not set in environment variables."
        })

    body = await request.json()

    # S23: Provider field comparison test
    if body.get("provider_test"):
        if not db_pool:
            return JSONResponse(status_code=500, content={"error": "No database connected"})

        # Pick one known sub from each source
        test_emails = {}
        async with db_pool.acquire() as conn:
            for src in ("apple", "google", "stripe"):
                row = await conn.fetchrow(
                    """SELECT email, stripe_subscription_id, source FROM subscriptions
                       WHERE source = $1 AND status = 'active' AND email != '' AND email IS NOT NULL
                       ORDER BY created_at DESC LIMIT 1""", src
                )
                if row:
                    test_emails[src] = {"email": row["email"], "sub_id": row["stripe_subscription_id"]}

        results = {}
        async with httpx.AsyncClient(timeout=15.0) as client:
            # Individual lookups
            for src, info in test_emails.items():
                try:
                    resp = await client.get(
                        f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup",
                        headers={"X-Authorization": ymove_key},
                        params={"email": info["email"]}
                    )
                    raw = resp.json() if resp.status_code == 200 else {"http_status": resp.status_code}
                    user_obj = raw.get("user", {})
                    results[src] = {
                        "email": info["email"],
                        "our_sub_id": info["sub_id"],
                        "our_source": src,
                        "individual_lookup": {
                            "subscriptionProvider": user_obj.get("subscriptionProvider"),
                            "subscriptionPaymentProvider": user_obj.get("subscriptionPaymentProvider"),
                            "provider": user_obj.get("provider"),
                            "all_provider_fields": {k: v for k, v in user_obj.items() if "provider" in k.lower() or "payment" in k.lower() or "source" in k.lower()},
                            "full_user_keys": list(user_obj.keys()),
                        }
                    }
                except Exception as e:
                    results[src] = {"email": info["email"], "error": str(e)}

                await asyncio.sleep(1.5)

            # Bulk lookup - grab page 1 and check the same fields
            try:
                resp = await client.get(
                    f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup/all",
                    headers={"X-Authorization": ymove_key},
                    params={"status": "subscribed", "page": "1"}
                )
                if resp.status_code == 200:
                    bulk_data = resp.json()
                    bulk_users = bulk_data.get("users", [])
                    # Find our test emails in the bulk response
                    bulk_by_email = {(u.get("email") or "").strip().lower(): u for u in bulk_users}
                    for src, info in test_emails.items():
                        bulk_user = bulk_by_email.get(info["email"].lower())
                        if bulk_user:
                            results[src]["bulk_lookup"] = {
                                "subscriptionProvider": bulk_user.get("subscriptionProvider"),
                                "subscriptionPaymentProvider": bulk_user.get("subscriptionPaymentProvider"),
                                "provider": bulk_user.get("provider"),
                                "all_provider_fields": {k: v for k, v in bulk_user.items() if "provider" in k.lower() or "payment" in k.lower() or "source" in k.lower()},
                                "full_user_keys": list(bulk_user.keys()),
                            }
                        else:
                            results[src]["bulk_lookup"] = "not_found_on_page_1"
                    # Also include first 3 raw users from bulk for inspection
                    results["_bulk_sample"] = [{k: v for k, v in u.items() if "provider" in k.lower() or "payment" in k.lower() or "email" in k.lower() or "source" in k.lower()} for u in bulk_users[:3]]
                else:
                    results["_bulk_error"] = {"http_status": resp.status_code}
            except Exception as e:
                results["_bulk_error"] = str(e)

        return {
            "test": "provider_field_comparison",
            "purpose": "Compare individual member-lookup vs bulk member-lookup/all provider fields",
            "results": results,
            "next_step": "If individual returns provider but bulk does not, we can extract provider during the verify phase of shadow sync"
        }

    # Smoke test mode
    if body.get("smoke_test"):
        test_email = body.get("email", "takacsmeghan@gmail.com")
        return await _ymove_smoke_test(ymove_key, test_email)

    # Pull all subscribed members and cross-reference
    if body.get("pull_all_subscribed"):
        return await _ymove_pull_all_subscribed(ymove_key)

    # Single or batch email lookup
    emails = body.get("emails", [])
    if not emails and body.get("email"):
        emails = [body["email"]]
    if not emails:
        return JSONResponse(status_code=400, content={
            "error": "Provide email, emails, smoke_test, or pull_all_subscribed"
        })

    batch_size = min(body.get("batch_size", 10), 50)
    delay_seconds = max(body.get("delay_seconds", 1.5), 0.5)

    results = []
    rate_limit_hits = 0
    errors = 0
    total_delay = 0.0

    async with httpx.AsyncClient(timeout=15.0) as client:
        for i, email in enumerate(emails):
            if i > 0 and i % batch_size == 0:
                await asyncio.sleep(delay_seconds)
                total_delay += delay_seconds
            try:
                resp = await client.get(
                    f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup",
                    headers={"X-Authorization": ymove_key},
                    params={"email": email.strip().lower()}
                )
                if resp.status_code == 429:
                    rate_limit_hits += 1
                    retry_after = float(resp.headers.get("retry-after", "5"))
                    print(f"[ymove] Rate limited on {email}, backing off {retry_after}s")
                    await asyncio.sleep(retry_after)
                    total_delay += retry_after
                    resp = await client.get(
                        f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup",
                        headers={"X-Authorization": ymove_key},
                        params={"email": email.strip().lower()}
                    )
                if resp.status_code == 200:
                    data = resp.json()
                    status = _ymove_parse_status(data)
                    results.append({"email": email, "status": status, "raw": data})
                elif resp.status_code == 404:
                    results.append({"email": email, "status": "not_found", "raw": None})
                else:
                    results.append({
                        "email": email, "status": "error",
                        "http_status": resp.status_code, "raw": resp.text[:300]
                    })
                    errors += 1
            except Exception as e:
                results.append({"email": email, "status": "error", "error": str(e)})
                errors += 1

    active_count = sum(1 for r in results if r["status"] == "active")
    expired_count = sum(1 for r in results if r["status"] == "expired")
    not_found_count = sum(1 for r in results if r["status"] == "not_found")

    return {
        "total": len(emails), "verified": len(results),
        "active": active_count, "expired": expired_count,
        "not_found": not_found_count, "errors": errors,
        "rate_limit_hits": rate_limit_hits,
        "total_delay_seconds": round(total_delay, 1),
        "results": results
    }


async def _ymove_smoke_test(api_key: str, test_email: str) -> dict:
    """Quick connectivity test against ymove member-lookup with siteId 75."""
    async with httpx.AsyncClient(timeout=10.0) as client:
        try:
            url = f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup"
            resp = await client.get(
                url,
                headers={"X-Authorization": api_key},
                params={"email": test_email}
            )
            try:
                body = resp.json()
            except Exception:
                body = resp.text[:500]
            return {
                "status": "smoke_test_complete",
                "test_email": test_email,
                "http_status": resp.status_code,
                "api_url": url,
                "response": body,
                "success": resp.status_code == 200,
                "next_step": "If success, try {\"pull_all_subscribed\": true} to cross-ref candidates"
            }
        except Exception as e:
            return {"status": "error", "error": str(e)}


async def _ymove_pull_all_subscribed(api_key: str) -> dict:
    """Pull all subscribed members from ymove paginated API, cross-ref with our cancelled Apple/Google subs."""
    if not db_pool:
        return {"error": "No database connected"}

    # Step 1: Pull all subscribed members page by page
    all_members = []
    page = 1
    total_pages = None
    rate_limit_hits = 0
    last_data = {}

    async with httpx.AsyncClient(timeout=20.0) as client:
        while True:
            try:
                resp = await client.get(
                    f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup/all",
                    headers={"X-Authorization": api_key},
                    params={"status": "subscribed", "page": str(page)}
                )
                if resp.status_code == 429:
                    rate_limit_hits += 1
                    retry_after = float(resp.headers.get("retry-after", "5"))
                    print(f"[ymove] Rate limited on page {page}, backing off {retry_after}s")
                    await asyncio.sleep(retry_after)
                    continue
                if resp.status_code != 200:
                    return {
                        "error": f"ymove API returned {resp.status_code} on page {page}",
                        "body": resp.text[:500],
                        "members_so_far": len(all_members)
                    }
                last_data = resp.json()
                users = last_data.get("users", [])
                all_members.extend(users)
                total_pages = last_data.get("totalPages", 1)
                print(f"[ymove] Page {page}/{total_pages}: {len(users)} members (total: {len(all_members)})")
                if page >= total_pages:
                    break
                page += 1
                await asyncio.sleep(1.0)
            except Exception as e:
                return {"error": f"Failed on page {page}: {str(e)}", "members_so_far": len(all_members)}

    # Step 2: Build lookup of ymove active emails
    ymove_active = {}
    for m in all_members:
        email = (m.get("email") or "").strip().lower()
        if email:
            ymove_active[email] = {
                "provider": m.get("subscriptionProvider"),
                "plan_id": m.get("activeSubscriptionPlanID"),
                "first_name": m.get("firstName"),
                "last_name": m.get("lastName"),
                "signup_date": m.get("signupDate"),
            }

    # Step 3: Cross-ref with our cancelled Apple/Google subs
    async with db_pool.acquire() as conn:
        cancelled_rows = await conn.fetch("""
            SELECT DISTINCT ON (lower(email))
                id, lower(email) as em, source, status, canceled_at, created_at,
                stripe_subscription_id, first_name, last_name
            FROM subscriptions
            WHERE status = 'canceled' AND source IN ('apple', 'google') AND email != ''
            ORDER BY lower(email), created_at DESC
        """)
        active_emails = set(r["em"] for r in await conn.fetch(
            "SELECT DISTINCT lower(email) as em FROM subscriptions WHERE email != '' AND status IN ('active', 'trialing')"
        ))

    # Step 4: Categorize
    reactivation_candidates = []
    already_active = 0
    confirmed_cancelled = 0

    for row in cancelled_rows:
        email = row["em"]
        if email in active_emails:
            already_active += 1
            continue
        if email in ymove_active:
            ym = ymove_active[email]
            reactivation_candidates.append({
                "email": email,
                "db_source": row["source"],
                "db_sub_id": row["stripe_subscription_id"],
                "db_canceled_at": str(row["canceled_at"] or ""),
                "ymove_provider": ym["provider"],
                "ymove_plan_id": ym["plan_id"],
                "ymove_first_name": ym["first_name"],
                "ymove_last_name": ym["last_name"],
            })
        else:
            confirmed_cancelled += 1

    reactivation_candidates.sort(key=lambda x: x["email"])
    react_emails = [c["email"] for c in reactivation_candidates]

    return {
        "ymove_total_subscribed": last_data.get("total", len(all_members)),
        "ymove_pages_fetched": page,
        "ymove_unique_emails": len(ymove_active),
        "our_cancelled_apple_google": len(cancelled_rows),
        "already_active_in_db": already_active,
        "reactivation_candidates": len(reactivation_candidates),
        "confirmed_cancelled_in_ymove_too": confirmed_cancelled,
        "rate_limit_hits": rate_limit_hits,
        "candidates": reactivation_candidates,
        "candidate_emails": react_emails,
        "next_step": f"Review {len(reactivation_candidates)} candidates, then POST /api/admin/ymove-reactivate with emails list (preview first)"
    }


def _ymove_parse_status(data: dict) -> str:
    """Parse ymove member-lookup response. Format: {found: bool, user: {activeSubscription: bool, ...}}"""
    if not isinstance(data, dict):
        return "unknown"
    if not data.get("found"):
        return "not_found"
    user = data.get("user", {})
    if user.get("activeSubscription") is True:
        return "active"
    if user.get("previouslySubscribed") is True:
        return "expired"
    return "not_found"


@app.post("/api/admin/ymove-reactivate")
async def ymove_reactivate(request: Request):
    """Reactivate confirmed-active subs from ymove verification.
    Body: {"emails": [...], "preview": true/false}
    Only reactivates Apple/Google source subs (Stripe webhook is authoritative)."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    body = await request.json()
    emails = body.get("emails", [])
    if not emails:
        return JSONResponse(status_code=400, content={"error": "Provide emails list"})

    batch_id = body.get("batch_id", f"ymove_react_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}")
    preview = body.get("preview", True)

    async with db_pool.acquire() as conn:
        reactivated = 0
        skipped = 0
        details = []
        for email in emails:
            email_lower = email.strip().lower()
            row = await conn.fetchrow("""
                SELECT id, stripe_subscription_id, source, status, canceled_at
                FROM subscriptions
                WHERE lower(email) = $1 AND status = 'canceled' AND source IN ('apple', 'google')
                ORDER BY created_at DESC LIMIT 1
            """, email_lower)
            if not row:
                skipped += 1
                details.append({"email": email_lower, "action": "skipped", "reason": "no cancelled apple/google sub"})
                continue
            if preview:
                details.append({"email": email_lower, "action": "would_reactivate", "sub_id": row["stripe_subscription_id"], "source": row["source"], "canceled_at": str(row["canceled_at"] or "")})
                reactivated += 1
            else:
                await conn.execute("UPDATE subscriptions SET status = 'active', canceled_at = NULL, reactivated_at = NOW(), updated_at = NOW(), import_batch = $1 WHERE id = $2", batch_id, row["id"])
                reactivated += 1
                details.append({"email": email_lower, "action": "reactivated", "sub_id": row["stripe_subscription_id"], "source": row["source"]})

    return {
        "status": "preview" if preview else "ok",
        "batch_id": batch_id if not preview else None,
        "reactivated": reactivated, "skipped": skipped,
        "total_emails": len(emails), "details": details,
        "next_step": "Set preview: false to execute" if preview else f"Revert with POST /api/admin/revert-batch batch_id={batch_id}"
    }


@app.post("/api/admin/ymove-deactivate")
async def ymove_deactivate(request: Request):
    """Deactivate subs confirmed expired by ymove API verification (S19).
    Body: {"emails": [...], "preview": true/false}
    Only deactivates Apple/Google source subs marked active."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    body = await request.json()
    emails = body.get("emails", [])
    if not emails:
        return JSONResponse(status_code=400, content={"error": "Provide emails list"})

    batch_id = body.get("batch_id", f"ymove_deact_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}")
    preview = body.get("preview", True)

    async with db_pool.acquire() as conn:
        deactivated = 0
        skipped = 0
        details = []
        for email in emails:
            email_lower = email.strip().lower()
            row = await conn.fetchrow("""
                SELECT id, stripe_subscription_id, source, status, plan_amount
                FROM subscriptions
                WHERE lower(email) = $1 AND status IN ('active', 'trialing') AND source IN ('apple', 'google')
                ORDER BY created_at DESC LIMIT 1
            """, email_lower)
            if not row:
                skipped += 1
                details.append({"email": email_lower, "action": "skipped", "reason": "no active apple/google sub"})
                continue
            if preview:
                details.append({"email": email_lower, "action": "would_deactivate", "sub_id": row["stripe_subscription_id"], "source": row["source"], "plan_amount": row["plan_amount"]})
                deactivated += 1
            else:
                await conn.execute("""
                    UPDATE subscriptions
                    SET status = 'canceled', canceled_at = NOW(), updated_at = NOW(), import_batch = $1
                    WHERE id = $2
                """, batch_id, row["id"])
                deactivated += 1
                details.append({"email": email_lower, "action": "deactivated", "sub_id": row["stripe_subscription_id"], "source": row["source"]})

    return {
        "status": "preview" if preview else "ok",
        "batch_id": batch_id if not preview else None,
        "deactivated": deactivated, "skipped": skipped,
        "total_emails": len(emails), "details": details,
        "next_step": "Set preview: false to execute" if preview else f"Revert with POST /api/admin/revert-batch batch_id={batch_id}"
    }


# --- Session 20: ymove Shadow Sync ---

_active_sync_task = None  # Track running sync task


@app.post("/api/admin/ymove-shadow-sync")
async def ymove_shadow_sync(request: Request):
    """Shadow sync: pull ymove data, compute diff, optionally apply.
    Actions:
      run    - Start background sync task
      status - Check progress of running/latest sync
      diff   - Get categorized diff from latest completed sync
      apply  - Apply diff changes with batch_id (preview/confirm)
    """
    global _active_sync_task
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    if not YMOVE_API_KEY:
        raise HTTPException(status_code=500, detail="YMOVE_API_KEY not configured")

    body = await request.json()
    action = body.get("action", "status")

    if action == "run":
        # Check for already running sync
        async with db_pool.acquire() as conn:
            running = await conn.fetchrow(
                "SELECT id, started_at FROM ymove_sync_runs WHERE status = 'running' ORDER BY started_at DESC LIMIT 1"
            )
            if running:
                age_minutes = (datetime.now(timezone.utc) - running["started_at"]).total_seconds() / 60
                # Auto-expire stuck runs after 30 minutes
                if age_minutes > 30:
                    await conn.execute(
                        "UPDATE ymove_sync_runs SET status = 'failed', error = 'Timed out after 30 minutes', completed_at = NOW() WHERE id = $1",
                        running["id"]
                    )
                else:
                    return {"status": "already_running", "run_id": running["id"],
                            "running_for_minutes": round(age_minutes, 1),
                            "hint": "Use action: status to check progress"}

            # Create new run record
            run_id = await conn.fetchval(
                "INSERT INTO ymove_sync_runs (status, phase) VALUES ('running', 'init') RETURNING id"
            )

        # Launch background task
        _active_sync_task = asyncio.create_task(_run_shadow_sync(run_id))
        return {"status": "started", "run_id": run_id, "hint": "Poll with action: status (every 10-15s)"}

    elif action == "status":
        async with db_pool.acquire() as conn:
            run = await conn.fetchrow(
                "SELECT * FROM ymove_sync_runs ORDER BY started_at DESC LIMIT 1"
            )
        if not run:
            return {"status": "no_runs", "hint": "Start with action: run"}

        pct = ""
        if run["progress_total"] and run["progress_total"] > 0:
            pct = f" ({round(run['progress_current'] / run['progress_total'] * 100)}%)"

        result = {
            "run_id": run["id"],
            "status": run["status"],
            "phase": run["phase"],
            "progress": f"{run['progress_current']}/{run['progress_total']}{pct}" if run["progress_total"] else "initializing",
            "started_at": str(run["started_at"]),
            "completed_at": str(run["completed_at"]) if run["completed_at"] else None,
            "our_active_count": run["our_active_count"],
            "ymove_active_count": run["ymove_active_count"],
        }
        if run["error"]:
            result["error"] = run["error"]
        # Include summary if completed
        if run["status"] == "completed" and run["results"]:
            res_data = run["results"] if isinstance(run["results"], dict) else json.loads(run["results"])
            result["summary"] = {
                "to_deactivate": len(res_data.get("to_deactivate", [])),
                "to_reactivate": len(res_data.get("to_reactivate", [])),
                "cross_platform_switchers": len(res_data.get("cross_platform_switchers", [])),
                "active_stripe_in_ymove": res_data.get("active_stripe_in_ymove", 0),
                "truly_new": len(res_data.get("truly_new", [])),
                "unchanged": res_data.get("unchanged", 0),
                "not_found": len(res_data.get("not_found_in_ymove", [])),
                "pull_all_status": res_data.get("pull_all_status", "unknown"),
            }
        return result

    elif action == "diff":
        async with db_pool.acquire() as conn:
            run = await conn.fetchrow(
                "SELECT * FROM ymove_sync_runs WHERE status = 'completed' ORDER BY started_at DESC LIMIT 1"
            )
        if not run:
            return {"status": "no_completed_runs", "hint": "Start with action: run, wait for completion"}

        res_data = run["results"] if isinstance(run["results"], dict) else json.loads(run["results"])
        return {
            "run_id": run["id"],
            "completed_at": str(run["completed_at"]),
            "our_active_count": run["our_active_count"],
            "ymove_active_count": run["ymove_active_count"],
            "to_deactivate": res_data.get("to_deactivate", []),
            "to_reactivate": res_data.get("to_reactivate", []),
            "cross_platform_switchers": res_data.get("cross_platform_switchers", []),
            "active_stripe_in_ymove": res_data.get("active_stripe_in_ymove", 0),
            "truly_new": res_data.get("truly_new", []),
            "unchanged": res_data.get("unchanged", 0),
            "not_found_in_ymove": res_data.get("not_found_in_ymove", []),
            "errors": res_data.get("errors", 0),
            "pull_all_status": res_data.get("pull_all_status", "unknown"),
            "pull_all_emails_found": res_data.get("pull_all_emails_found", 0),
            "verified_count": res_data.get("verified_count", 0),
        }

    elif action == "apply":
        preview = body.get("preview", True)
        async with db_pool.acquire() as conn:
            run = await conn.fetchrow(
                "SELECT * FROM ymove_sync_runs WHERE status = 'completed' ORDER BY started_at DESC LIMIT 1"
            )
        if not run:
            return {"status": "no_completed_runs", "hint": "Run a sync first"}

        res_data = run["results"] if isinstance(run["results"], dict) else json.loads(run["results"])
        return await _apply_shadow_sync(res_data, preview, run["id"])

    else:
        return JSONResponse(status_code=400, content={
            "error": f"Unknown action: {action}. Use run, status, diff, or apply."
        })


async def _run_shadow_sync(run_id: int):
    """Background task: verify active Apple/Google emails + best-effort pull_all from ymove."""
    try:
        print(f"[Shadow Sync] Run {run_id} starting...")

        # S26 Phase 0: Expire any pending-cancel records whose period_end has passed.
        # Runs first so the rest of the sync sees the post-expiry state.
        try:
            async with db_pool.acquire() as conn:
                _sweep_result = await s26_expire_pending_cancels(conn)
            print(f"[Shadow Sync] S26 expiry sweep: {_sweep_result.get('expired_count', 0)} records expired")
        except Exception as _se:
            print(f"[Shadow Sync] S26 expiry sweep error (non-fatal): {_se}")

        # --- Gather our DB state ---
        async with db_pool.acquire() as conn:
            # All our active Apple/Google/undetermined subs with email
            our_active = await conn.fetch(
                """SELECT id, lower(email) as email, source, stripe_subscription_id, plan_amount, plan_interval
                   FROM subscriptions
                   WHERE status IN ('active', 'trialing') AND source IN ('apple', 'google', 'undetermined')
                   AND email != '' AND email IS NOT NULL
                   ORDER BY email"""
            )
            # All cancelled Apple/Google/undetermined subs (most recent per email, for reactivation)
            our_cancelled_ag = await conn.fetch(
                """SELECT DISTINCT ON (lower(s.email))
                   s.id, lower(s.email) as email, s.source, s.stripe_subscription_id
                   FROM subscriptions s
                   WHERE s.status = 'canceled' AND s.source IN ('apple', 'google', 'undetermined', 'manual')
                   AND s.email != '' AND s.email IS NOT NULL
                   AND (s.import_batch IS NULL OR s.import_batch NOT LIKE 's23_provider_cleanup%%')
                   AND (s.import_batch IS NULL OR s.import_batch NOT LIKE 's24_%%')
                   AND NOT EXISTS (
                       SELECT 1 FROM subscriptions s2
                       WHERE lower(s2.email) = lower(s.email)
                       AND s2.stripe_subscription_id LIKE 'sub_%%'
                       AND s2.status IN ('active', 'trialing')
                   )
                   ORDER BY lower(s.email), s.created_at DESC"""
            )
            # All known emails across all sources/statuses
            all_known_rows = await conn.fetch(
                "SELECT DISTINCT lower(email) as em FROM subscriptions WHERE email != '' AND email IS NOT NULL"
            )

        our_active_emails = [r["email"] for r in our_active]
        our_active_set = set(our_active_emails)
        our_active_lookup = {r["email"]: r for r in our_active}
        our_cancelled_ag_map = {r["email"]: r for r in our_cancelled_ag if r["email"] not in our_active_set}
        all_known_emails = set(r["em"] for r in all_known_rows)

        total_to_verify = len(our_active_emails)
        print(f"[Shadow Sync] {total_to_verify} active Apple/Google emails to verify, {len(our_cancelled_ag_map)} cancelled AG candidates")

        async with db_pool.acquire() as conn:
            await conn.execute(
                "UPDATE ymove_sync_runs SET phase = 'verify_existing', progress_total = $1, our_active_count = $2 WHERE id = $3",
                total_to_verify, total_to_verify, run_id
            )

        # --- Phase 1: Verify all active Apple/Google emails against ymove ---
        # S23: Also extract subscriptionProvider for self-healing
        verify_results = {}
        verify_providers = {}  # email -> provider from ymove API (for self-healing)
        batch_size = 10
        delay_seconds = 1.5
        processed = 0

        async with httpx.AsyncClient(timeout=15.0) as client:
            for i, email in enumerate(our_active_emails):
                if i > 0 and i % batch_size == 0:
                    await asyncio.sleep(delay_seconds)
                try:
                    resp = await client.get(
                        f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup",
                        headers={"X-Authorization": YMOVE_API_KEY},
                        params={"email": email}
                    )
                    if resp.status_code == 429:
                        retry_after = float(resp.headers.get("retry-after", "5"))
                        print(f"[Shadow Sync] Rate limited at {email}, backing off {retry_after}s")
                        await asyncio.sleep(retry_after)
                        resp = await client.get(
                            f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup",
                            headers={"X-Authorization": YMOVE_API_KEY},
                            params={"email": email}
                        )
                    if resp.status_code == 200:
                        data = resp.json()
                        verify_results[email] = _ymove_parse_status(data)
                        # S23: Extract provider for self-healing
                        user_obj = data.get("user", {})
                        prov = user_obj.get("subscriptionProvider")
                        if prov and isinstance(prov, str) and prov.strip():
                            verify_providers[email] = prov.strip().lower()
                    elif resp.status_code == 404:
                        verify_results[email] = "not_found"
                    else:
                        verify_results[email] = "error"
                except Exception as e:
                    verify_results[email] = "error"
                    print(f"[Shadow Sync] Verify error for {email}: {e}")

                processed += 1
                if processed % 50 == 0:
                    async with db_pool.acquire() as conn:
                        await conn.execute(
                            "UPDATE ymove_sync_runs SET progress_current = $1 WHERE id = $2",
                            processed, run_id
                        )

        # S25: Self-healing — update provider on any records where ymove returned a real value.
        # 'manual' EXCLUDED from allowed providers — ymove returns "manual" for users that have been
        # manually edited in their admin tool, but those users still pay via Apple/Google/Stripe.
        # Treating 'manual' as a provider value caused 9 records to be mislabeled this session.
        # See S25 Phase B notes.
        provider_healed = 0
        if verify_providers:
            async with db_pool.acquire() as conn:
                for email, prov in verify_providers.items():
                    r = our_active_lookup.get(email)
                    if r and r["source"] != prov and prov in ("apple", "google", "stripe"):
                        try:
                            await conn.execute(
                                "UPDATE subscriptions SET source = $1, updated_at = NOW() WHERE id = $2",
                                prov, r["id"]
                            )
                            provider_healed += 1
                            print(f"[Shadow Sync] Self-healed provider: {email} {r['source']} -> {prov}")
                        except Exception as e:
                            print(f"[Shadow Sync] Provider heal error for {email}: {e}")
            print(f"[Shadow Sync] Self-healed {provider_healed} provider labels")

        async with db_pool.acquire() as conn:
            await conn.execute(
                "UPDATE ymove_sync_runs SET progress_current = $1, phase = 'pull_all' WHERE id = $2",
                processed, run_id
            )

        print(f"[Shadow Sync] Phase 1 done: verified {processed} emails")

        # --- Phase 2: Best-effort pull all subscribed members ---
        ymove_all_emails = {}  # email -> provider (S22: dict for auto-import)
        pull_all_status = "starting"
        pull_all_pages = 0
        total_pages_est = 0

        try:
            page = 1
            async with httpx.AsyncClient(timeout=20.0) as client:
                while True:
                    try:
                        resp = await client.get(
                            f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup/all",
                            headers={"X-Authorization": YMOVE_API_KEY},
                            params={"status": "subscribed", "page": str(page)}
                        )
                    except httpx.TimeoutException:
                        print(f"[Shadow Sync] Timeout on pull_all page {page}")
                        pull_all_status = f"timeout_page_{page}"
                        break

                    if resp.status_code == 429:
                        retry_after = float(resp.headers.get("retry-after", "5"))
                        print(f"[Shadow Sync] Rate limited on pull_all page {page}, backing off {retry_after}s")
                        await asyncio.sleep(retry_after)
                        continue

                    if resp.status_code != 200:
                        pull_all_status = f"error_page_{page}_http_{resp.status_code}"
                        print(f"[Shadow Sync] pull_all error: HTTP {resp.status_code} on page {page}")
                        break

                    data = resp.json()
                    users = data.get("users", [])
                    for u in users:
                        em = (u.get("email") or "").strip().lower()
                        if em:
                            provider = (u.get("subscriptionProvider") or "undetermined").lower()
                            if provider not in ("apple", "google", "stripe", "manual", "undetermined"):
                                provider = "undetermined"
                            ymove_all_emails[em] = provider

                    total_pages_est = data.get("totalPages", 1)
                    pull_all_pages = page

                    if page >= total_pages_est:
                        pull_all_status = "success"
                        break

                    page += 1
                    await asyncio.sleep(1.0)

                    # Update progress every 50 pages
                    if page % 50 == 0:
                        async with db_pool.acquire() as conn:
                            await conn.execute(
                                "UPDATE ymove_sync_runs SET phase = $1 WHERE id = $2",
                                f"pull_all_page_{page}_of_{total_pages_est}", run_id
                            )
        except Exception as e:
            pull_all_status = f"error: {str(e)[:200]}"
            print(f"[Shadow Sync] pull_all exception: {e}")

        print(f"[Shadow Sync] Phase 2 done: pull_all_status={pull_all_status}, found {len(ymove_all_emails)} emails across {pull_all_pages} pages")

        # --- Phase 3: Compute diff ---
        async with db_pool.acquire() as conn:
            await conn.execute(
                "UPDATE ymove_sync_runs SET phase = 'computing_diff', ymove_active_count = $1 WHERE id = $2",
                len(ymove_all_emails), run_id
            )
            # Query active Stripe emails to separate from cross-platform switchers
            active_stripe_rows = await conn.fetch(
                "SELECT DISTINCT lower(email) as em FROM subscriptions WHERE email != '' AND status IN ('active', 'trialing') AND source = 'stripe'"
            )
        active_stripe_emails = set(r["em"] for r in active_stripe_rows)

        to_deactivate = []
        to_reactivate = []
        cross_platform_switchers = []  # Cancelled Stripe, now active on Apple/Google
        active_stripe_in_ymove = 0     # Active Stripe users ymove also tracks (no action)
        truly_new = []
        unchanged = 0
        not_found = []
        verify_errors = 0
        # S26 Bug B fix: records where individual lookup said 'expired' but bulk pull
        # says 'subscribed'. Possible transient ymove glitch (the Kelsey case).
        # Do NOT deactivate these — flag for review instead.
        conflicting_expired_vs_bulk = []

        # Categorize our active subs based on ymove verification
        for email, ymove_status in verify_results.items():
            if ymove_status == "active":
                unchanged += 1
            elif ymove_status == "expired":
                r = our_active_lookup.get(email)
                if r:
                    # S26 Bug B: cross-check the bulk pull. Only deactivate if bulk also agrees.
                    # If bulk pull failed or returned 0 results, fall through to old behavior
                    # (otherwise a single bulk failure would freeze all deactivations).
                    bulk_pull_usable = (pull_all_status == "success" and len(ymove_all_emails) > 0)
                    if bulk_pull_usable and email in ymove_all_emails:
                        # Conflict: individual says expired, bulk says subscribed. Defer.
                        conflicting_expired_vs_bulk.append({
                            "email": email,
                            "sub_id": r["stripe_subscription_id"],
                            "source": r["source"],
                            "db_id": r["id"],
                            "bulk_provider": ymove_all_emails.get(email),
                        })
                    else:
                        to_deactivate.append({
                            "email": email,
                            "sub_id": r["stripe_subscription_id"],
                            "source": r["source"],
                            "plan_amount": r["plan_amount"],
                            "db_id": r["id"],
                        })
            elif ymove_status == "not_found":
                r = our_active_lookup.get(email)
                if r:
                    not_found.append({
                        "email": email,
                        "sub_id": r["stripe_subscription_id"],
                        "source": r["source"],
                        "db_id": r["id"],
                    })
            else:
                verify_errors += 1

        # Categorize ymove subscribers we don't have as active
        if pull_all_status == "success":
            for email in ymove_all_emails:
                if email in our_active_set:
                    continue
                if email in our_cancelled_ag_map:
                    row = our_cancelled_ag_map[email]
                    to_reactivate.append({
                        "email": email,
                        "sub_id": row["stripe_subscription_id"],
                        "source": row["source"],
                        "db_id": row["id"],
                    })
                elif email in active_stripe_emails:
                    # Active Stripe sub, ymove just tracks them too. No action needed.
                    active_stripe_in_ymove += 1
                elif email in all_known_emails:
                    # Known email but no active sub of any kind. Real cross-platform switcher.
                    cross_platform_switchers.append({"email": email, "provider": ymove_all_emails.get(email, "undetermined")})
                else:
                    truly_new.append({"email": email, "provider": ymove_all_emails.get(email, "undetermined")})

        # Build final results
        results = {
            "to_deactivate": to_deactivate,
            "to_reactivate": to_reactivate,
            "cross_platform_switchers": cross_platform_switchers,
            "active_stripe_in_ymove": active_stripe_in_ymove,
            "truly_new": truly_new,
            "unchanged": unchanged,
            "not_found_in_ymove": not_found,
            "errors": verify_errors,
            "pull_all_status": pull_all_status,
            "pull_all_pages": pull_all_pages,
            "pull_all_total_pages": total_pages_est,
            "pull_all_emails_found": len(ymove_all_emails),
            "verified_count": len(verify_results),
            "provider_healed": provider_healed,
            "s26_conflicting_expired_vs_bulk": conflicting_expired_vs_bulk,
        }

        async with db_pool.acquire() as conn:
            await conn.execute(
                """UPDATE ymove_sync_runs SET
                   status = 'completed', phase = 'done', completed_at = NOW(),
                   progress_current = $1, results = $2,
                   ymove_active_count = $3
                   WHERE id = $4""",
                processed, json.dumps(results, default=str),
                len(ymove_all_emails), run_id
            )

        print(f"[Shadow Sync] Run {run_id} COMPLETED. "
              f"Deactivate: {len(to_deactivate)}, Reactivate: {len(to_reactivate)}, "
              f"Switchers: {len(cross_platform_switchers)}, Stripe-in-ymove: {active_stripe_in_ymove}, "
              f"New: {len(truly_new)}, Unchanged: {unchanged}, Not found: {len(not_found)}")

    except Exception as e:
        print(f"[Shadow Sync] Run {run_id} FAILED: {e}")
        import traceback
        traceback.print_exc()
        try:
            async with db_pool.acquire() as conn:
                await conn.execute(
                    "UPDATE ymove_sync_runs SET status = 'failed', error = $1, completed_at = NOW() WHERE id = $2",
                    str(e)[:500], run_id
                )
        except Exception:
            pass


async def _apply_shadow_sync(results: dict, preview: bool, run_id: int) -> dict:
    """Apply shadow sync diff: deactivate expired, reactivate confirmed active."""
    batch_id = f"shadow_{run_id}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

    to_deactivate = results.get("to_deactivate", [])
    to_reactivate = results.get("to_reactivate", [])
    cross_platform_switchers = results.get("cross_platform_switchers", [])
    active_stripe_in_ymove = results.get("active_stripe_in_ymove", 0)
    truly_new = results.get("truly_new", [])

    if preview:
        deact_mrr = sum(r.get("plan_amount", 1999) for r in to_deactivate)
        react_mrr = len(to_reactivate) * 1999
        return {
            "status": "preview",
            "batch_id": batch_id,
            "would_deactivate": len(to_deactivate),
            "would_reactivate": len(to_reactivate),
            "cross_platform_switchers": len(cross_platform_switchers),
            "active_stripe_in_ymove": active_stripe_in_ymove,
            "truly_new_found": len(truly_new),
            "mrr_impact_deactivate_cents": -deact_mrr,
            "mrr_impact_reactivate_cents": react_mrr,
            "net_mrr_impact_cents": react_mrr - deact_mrr,
            "deactivate_sample": to_deactivate[:25],
            "reactivate_sample": to_reactivate[:25],
            "switcher_sample": cross_platform_switchers[:25],
            "truly_new_sample": truly_new[:25],
            "note": "Send preview: false to execute deactivations and reactivations. Switchers and new subscribers are NOT auto-applied (require manual review/import)."
        }

    deactivated = 0
    reactivated = 0
    errors = 0

    async with db_pool.acquire() as conn:
        for item in to_deactivate:
            try:
                # S26 Phase 3: For Stripe-source records, check Stripe API for period_end
                # before cancelling. If still in grace period, set pending state instead.
                # For Apple/Google, trust ymove (it already uses period-end semantics).
                src = (item.get("source") or "").lower()
                sub_id_check = item.get("stripe_subscription_id") or ""
                handled_as_pending = False

                if src == "stripe" and sub_id_check.startswith("sub_") and STRIPE_SECRET_KEY:
                    try:
                        import stripe as _s26_stripe
                        _s26_stripe.api_key = STRIPE_SECRET_KEY
                        _real = _s26_stripe.Subscription.retrieve(sub_id_check)
                        _pe = _real.get("current_period_end")
                        _now_ts = int(datetime.now(timezone.utc).timestamp())
                        if _pe and _pe > _now_ts:
                            await conn.execute(
                                """UPDATE subscriptions
                                   SET cancel_state = 'pending',
                                       pending_cancel_at = $1,
                                       cancel_requested_at = COALESCE(cancel_requested_at, NOW()),
                                       import_batch = $2,
                                       updated_at = NOW()
                                   WHERE id = $3 AND status IN ('active', 'trialing')""",
                                datetime.fromtimestamp(_pe, tz=timezone.utc), batch_id, item["db_id"]
                            )
                            handled_as_pending = True
                            print(f"[S26 ShadowSync] Held Stripe cancel in grace: {item.get('email')}")
                    except Exception as _e:
                        print(f"[S26 ShadowSync] Period-end check error for {sub_id_check}: {_e}")
                        # Fall through to normal cancel on error

                if not handled_as_pending:
                    result = await conn.execute(
                        """UPDATE subscriptions SET status = 'canceled', canceled_at = NOW(),
                           cancel_state = 'expired',
                           updated_at = NOW(), import_batch = $1
                           WHERE id = $2 AND status IN ('active', 'trialing')""",
                        batch_id, item["db_id"]
                    )
                    if result and result.endswith("1"):
                        deactivated += 1
                else:
                    deactivated += 1
            except Exception as e:
                print(f"[Shadow Sync Apply] Deactivate error {item.get('email')}: {e}")
                errors += 1

        for item in to_reactivate:
            try:
                result = await conn.execute(
                    """UPDATE subscriptions SET status = 'active', canceled_at = NULL,
                       reactivated_at = NOW(), updated_at = NOW(), import_batch = $1
                       WHERE id = $2 AND status = 'canceled'""",
                    batch_id, item["db_id"]
                )
                if result and result.endswith("1"):
                    reactivated += 1
            except Exception as e:
                print(f"[Shadow Sync Apply] Reactivate error {item.get('email')}: {e}")
                errors += 1

    return {
        "status": "applied",
        "batch_id": batch_id,
        "deactivated": deactivated,
        "reactivated": reactivated,
        "cross_platform_switchers": len(cross_platform_switchers),
        "active_stripe_in_ymove": active_stripe_in_ymove,
        "truly_new_found": len(truly_new),
        "errors": errors,
        "revert": f"POST /api/admin/revert-batch with batch_id={batch_id}",
        "note": f"Applied {deactivated} deactivations, {reactivated} reactivations. "
                f"{len(cross_platform_switchers)} switchers and {len(truly_new)} new subscribers NOT auto-applied (need manual review)."
    }


@app.post("/api/admin/ymove-import-new")
async def ymove_import_new(request: Request):
    """Import new Apple/Google subscribers discovered by shadow sync.
    Takes emails, verifies each against ymove to get provider, creates records.
    Body: {"emails": [...], "preview": true/false, "skip_test_accounts": true}
    """
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    if not YMOVE_API_KEY:
        raise HTTPException(status_code=500, detail="YMOVE_API_KEY not configured")

    body = await request.json()
    emails = body.get("emails", [])
    if not emails:
        return JSONResponse(status_code=400, content={"error": "Provide emails list"})

    preview = body.get("preview", True)
    skip_test = body.get("skip_test_accounts", True)

    # Filter test accounts
    test_domains = ["ymove.app"]
    test_patterns = ["test", "dsfg", "asdf", "qwer"]
    filtered_emails = []
    skipped_test = []
    for email in emails:
        em = email.strip().lower()
        if not em or "@" not in em:
            continue
        domain = em.split("@")[-1]
        local = em.split("@")[0]
        is_test = False
        if skip_test:
            if domain in test_domains:
                is_test = True
            for pat in test_patterns:
                if pat in local:
                    is_test = True
                    break
        if is_test:
            skipped_test.append(em)
        else:
            filtered_emails.append(em)

    # Check which emails already have active Apple/Google subs (skip those)
    async with db_pool.acquire() as conn:
        active_ag_rows = await conn.fetch(
            "SELECT DISTINCT lower(email) as em FROM subscriptions WHERE email != '' AND status IN ('active', 'trialing') AND source IN ('apple', 'google')"
        )
    active_ag_emails = set(r["em"] for r in active_ag_rows)

    new_emails = [e for e in filtered_emails if e not in active_ag_emails]
    already_known = [e for e in filtered_emails if e in active_ag_emails]

    # Verify against ymove to get provider and confirm active
    verified = []
    not_active = []
    verify_errors = []
    batch_size = 10
    delay_seconds = 1.5

    async with httpx.AsyncClient(timeout=15.0) as client:
        for i, email in enumerate(new_emails):
            if i > 0 and i % batch_size == 0:
                await asyncio.sleep(delay_seconds)
            try:
                resp = await client.get(
                    f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup",
                    headers={"X-Authorization": YMOVE_API_KEY},
                    params={"email": email}
                )
                if resp.status_code == 429:
                    retry_after = float(resp.headers.get("retry-after", "5"))
                    await asyncio.sleep(retry_after)
                    resp = await client.get(
                        f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup",
                        headers={"X-Authorization": YMOVE_API_KEY},
                        params={"email": email}
                    )
                if resp.status_code == 200:
                    data = resp.json()
                    status = _ymove_parse_status(data)
                    if status == "active":
                        # Try to determine provider from response
                        user_data = data.get("user", {})
                        provider = (user_data.get("subscriptionProvider") or
                                    user_data.get("provider") or "undetermined").lower()
                        if provider not in ("apple", "google", "stripe", "manual", "undetermined"):
                            provider = "undetermined"
                        verified.append({"email": email, "provider": provider, "raw": user_data})
                    else:
                        not_active.append({"email": email, "status": status})
                else:
                    verify_errors.append({"email": email, "http_status": resp.status_code})
            except Exception as e:
                verify_errors.append({"email": email, "error": str(e)})

    if preview:
        by_provider = {}
        for v in verified:
            p = v["provider"]
            by_provider[p] = by_provider.get(p, 0) + 1
        return {
            "status": "preview",
            "total_input": len(emails),
            "skipped_test_accounts": len(skipped_test),
            "already_active_apple_google": len(already_known),
            "new_to_verify": len(new_emails),
            "verified_active": len(verified),
            "not_active": len(not_active),
            "verify_errors": len(verify_errors),
            "by_provider": by_provider,
            "would_import": len(verified),
            "est_mrr_impact_cents": len(verified) * 1999,
            "verified_sample": [{"email": v["email"], "provider": v["provider"]} for v in verified[:30]],
            "skipped_test_sample": skipped_test[:10],
            "not_active_sample": not_active[:10],
            "already_known_sample": already_known[:10],
            "note": "Send preview: false to create Apple/Google records for verified-active emails."
        }

    # Create records
    batch_id = f"ymove_import_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"
    imported = 0
    errors = 0

    async with db_pool.acquire() as conn:
        for v in verified:
            email = v["email"]
            provider = v["provider"]
            email_hash = hashlib.md5(email.encode()).hexdigest()[:16]
            syn_id = f"ymove_new_{provider}_{email_hash}"
            # S23: Use ymove signupDate as created_at if available
            raw = v.get("raw", {})
            signup_str = raw.get("signupDate") or raw.get("createdAt") or ""
            signup_dt = _parse_iso(signup_str) if signup_str else None
            try:
                await conn.execute("""
                    INSERT INTO subscriptions (
                        stripe_customer_id, stripe_subscription_id, email, status,
                        plan_interval, plan_amount, currency, source,
                        created_at, updated_at, import_batch
                    ) VALUES ('', $1, $2, 'active', 'month', 1999, 'usd', $3, COALESCE($5, NOW()), NOW(), $4)
                    ON CONFLICT (stripe_subscription_id) DO NOTHING
                """, syn_id, email, provider, batch_id, signup_dt)
                imported += 1
            except Exception as e:
                print(f"[ymove import] Error for {email}: {e}")
                errors += 1

    return {
        "status": "ok",
        "batch_id": batch_id,
        "imported": imported,
        "errors": errors,
        "skipped_test_accounts": len(skipped_test),
        "already_active_apple_google": len(already_known),
        "not_active_in_ymove": len(not_active),
        "est_mrr_added_cents": imported * 1999,
        "revert": f"POST /api/admin/revert-batch with batch_id={batch_id}"
    }


@app.post("/api/admin/backfill-ymove-dates")
async def backfill_ymove_dates(request: Request):
    """Backfill created_at on ymove-imported subs using their real signupDate from ymove API.
    Finds subs with import_batch LIKE 'ymove_import_%' or 'ymove_react_%', looks up each in ymove,
    and updates created_at to signupDate if available and earlier than current created_at.
    Body: {"preview": true/false}"""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    if not YMOVE_API_KEY:
        raise HTTPException(status_code=500, detail="YMOVE_API_KEY not configured")

    body = await request.json()
    preview = body.get("preview", True)

    async with db_pool.acquire() as conn:
        candidates = await conn.fetch(
            """SELECT id, email, source, created_at, import_batch, stripe_subscription_id
               FROM subscriptions
               WHERE import_batch LIKE 'ymove_import_%' OR import_batch LIKE 'ymove_react_%'
               ORDER BY created_at DESC"""
        )

    if not candidates:
        return {"status": "ok", "message": "No ymove-imported subs found", "updated": 0}

    updated = 0
    skipped = 0
    errors = 0
    details = []

    async with httpx.AsyncClient(timeout=15.0) as client:
        for i, row in enumerate(candidates):
            email = (row["email"] or "").strip().lower()
            if not email:
                skipped += 1
                continue

            if i > 0 and i % 10 == 0:
                await asyncio.sleep(1.5)

            try:
                resp = await client.get(
                    f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup",
                    headers={"X-Authorization": YMOVE_API_KEY},
                    params={"email": email}
                )
                if resp.status_code == 429:
                    retry_after = float(resp.headers.get("retry-after", "5"))
                    await asyncio.sleep(retry_after)
                    resp = await client.get(
                        f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup",
                        headers={"X-Authorization": YMOVE_API_KEY},
                        params={"email": email}
                    )
                if resp.status_code != 200:
                    errors += 1
                    continue

                data = resp.json()
                user = data.get("user", {})
                signup_str = user.get("signupDate") or user.get("createdAt") or ""
                if not signup_str:
                    skipped += 1
                    details.append({"email": email, "action": "skipped", "reason": "no signupDate in ymove"})
                    continue

                signup_dt = _parse_iso(signup_str)
                if not signup_dt:
                    skipped += 1
                    details.append({"email": email, "action": "skipped", "reason": f"unparseable date: {signup_str}"})
                    continue

                current_created = row["created_at"]
                if signup_dt >= current_created:
                    skipped += 1
                    details.append({"email": email, "action": "skipped", "reason": "ymove date not earlier"})
                    continue

                detail = {
                    "email": email,
                    "old_created_at": str(current_created),
                    "new_created_at": str(signup_dt),
                    "batch": row["import_batch"],
                }

                if preview:
                    detail["action"] = "would_update"
                else:
                    async with db_pool.acquire() as conn2:
                        await conn2.execute(
                            "UPDATE subscriptions SET created_at = $1, updated_at = NOW() WHERE id = $2",
                            signup_dt, row["id"]
                        )
                    detail["action"] = "updated"

                details.append(detail)
                updated += 1

            except Exception as e:
                errors += 1
                details.append({"email": email, "action": "error", "reason": str(e)})

    return {
        "status": "preview" if preview else "ok",
        "total_candidates": len(candidates),
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "details": details[:50],
        "note": "Set preview: false to apply" if preview else "Done. MRR trend chart should now reflect accurate dates."
    }


async def run_daily_shadow_sync():
    """Automated daily shadow sync: verify all Apple/Google against ymove, auto-apply safe changes."""
    print(f"[Daily Sync] Starting automated shadow sync at {datetime.now(ZoneInfo('America/New_York')).strftime('%Y-%m-%d %H:%M ET')}")
    try:
        if not db_pool or not YMOVE_API_KEY:
            print("[Daily Sync] Skipped: missing db_pool or YMOVE_API_KEY")
            return

        # Check for already running sync
        async with db_pool.acquire() as conn:
            running = await conn.fetchrow(
                "SELECT id, started_at FROM ymove_sync_runs WHERE status = 'running' ORDER BY started_at DESC LIMIT 1"
            )
            if running:
                age_minutes = (datetime.now(timezone.utc) - running["started_at"]).total_seconds() / 60
                if age_minutes > 30:
                    await conn.execute(
                        "UPDATE ymove_sync_runs SET status = 'failed', error = 'Timed out after 30 minutes', completed_at = NOW() WHERE id = $1",
                        running["id"]
                    )
                else:
                    print(f"[Daily Sync] Skipped: sync already running (id={running['id']}, {round(age_minutes,1)}m)")
                    return

            # Create run record
            run_id = await conn.fetchval(
                "INSERT INTO ymove_sync_runs (status, phase) VALUES ('running', 'init') RETURNING id"
            )

        # Run the sync (same function as manual trigger)
        await _run_shadow_sync(run_id)

        # Wait briefly for completion to be written
        await asyncio.sleep(2)

        # Auto-apply safe changes (deactivations + reactivations only)
        async with db_pool.acquire() as conn:
            run = await conn.fetchrow(
                "SELECT * FROM ymove_sync_runs WHERE id = $1", run_id
            )

        if run and run["status"] == "completed" and run["results"]:
            res_data = run["results"] if isinstance(run["results"], dict) else json.loads(run["results"])
            to_deactivate = res_data.get("to_deactivate", [])
            to_reactivate = res_data.get("to_reactivate", [])
            switchers = len(res_data.get("cross_platform_switchers", []))
            truly_new = len(res_data.get("truly_new", []))

            if to_deactivate or to_reactivate:
                apply_result = await _apply_shadow_sync(res_data, False, run_id)
                print(f"[Daily Sync] Auto-applied: {apply_result.get('deactivated', 0)} deactivated, "
                      f"{apply_result.get('reactivated', 0)} reactivated, "
                      f"batch_id={apply_result.get('batch_id', 'n/a')}")
            else:
                print("[Daily Sync] No changes to apply. Database is in sync.")

            # S23: Alert mode — instead of silently importing, cross-ref with Meg imports, Stripe, and report
            switcher_list = res_data.get("cross_platform_switchers", [])
            truly_new_list = res_data.get("truly_new", [])
            all_unknown = switcher_list + [t for t in truly_new_list if not _is_test_email(t.get("email", ""))]

            if all_unknown:
                # Three-step waterfall: Meg import → Stripe API → undetermined
                meg_identified = []
                stripe_identified = []
                non_stripe = []
                stripe_errors = 0

                async with db_pool.acquire() as conn:
                    for item in all_unknown:
                        email = item.get("email", "")
                        if not email:
                            continue

                        # S24 Step 0: Skip if email already has ANY active record.
                        # Bug fix: previous waterfall only checked Meg/Stripe ID patterns,
                        # not "is there already an active record for this email at all?"
                        # Result: emails with existing ymove_new_apple_* records would get
                        # a NEW ymove_new_undetermined_* record inserted alongside, because
                        # the synthetic IDs differ in the provider segment so ON CONFLICT misses.
                        # 9 manual duplicates created on 2026-04-11 traced to this gap.
                        existing_active = await conn.fetchval(
                            """SELECT id FROM subscriptions
                               WHERE lower(email) = lower($1)
                               AND status IN ('active', 'trialing')
                               LIMIT 1""", email
                        )
                        if existing_active:
                            print(f"[Daily Sync] Step 0 skip: {email} already has active record (id={existing_active})")
                            continue

                        # Step 1: Check Meg import records for known provider
                        meg_record = await conn.fetchrow(
                            """SELECT source FROM subscriptions
                               WHERE lower(email) = lower($1)
                               AND (stripe_subscription_id LIKE 'import_apple_%%'
                                    OR stripe_subscription_id LIKE 'import_google_%%'
                                    OR stripe_subscription_id LIKE 'meg_apple_%%'
                                    OR stripe_subscription_id LIKE 'meg_google_%%')
                               ORDER BY created_at DESC LIMIT 1""", email
                        )
                        if meg_record and meg_record["source"] in ("apple", "google"):
                            meg_identified.append({"email": email, "source": meg_record["source"]})
                            continue

                        # Step 2: Check Stripe
                        try:
                            customers = stripe.Customer.list(email=email, limit=1)
                            if customers.data:
                                cust = customers.data[0]
                                subs = stripe.Subscription.list(customer=cust.id, status="active", limit=1)
                                if subs.data:
                                    stripe_identified.append({
                                        "email": email,
                                        "stripe_customer_id": cust.id,
                                        "stripe_sub_id": subs.data[0].id,
                                    })
                                    continue
                            non_stripe.append({"email": email})
                        except Exception as e:
                            stripe_errors += 1
                            non_stripe.append({"email": email})

                # Auto-import Meg-identified users with correct source
                if meg_identified:
                    batch_id_meg = f"autosync_meg_{run_id}_{datetime.now(timezone.utc).strftime('%Y%m%d')}"
                    imported_meg = 0
                    async with db_pool.acquire() as conn:
                        for item in meg_identified:
                            email_hash = hashlib.md5(item["email"].encode()).hexdigest()[:16]
                            syn_id = f"ymove_new_{item['source']}_{email_hash}"
                            try:
                                await conn.execute("""
                                    INSERT INTO subscriptions (
                                        stripe_customer_id, stripe_subscription_id, email, status,
                                        plan_interval, plan_amount, currency, source,
                                        created_at, updated_at, import_batch
                                    ) VALUES ('', $1, $2, 'active', 'month', 1999, 'usd', $3, NOW(), NOW(), $4)
                                    ON CONFLICT (stripe_subscription_id) DO NOTHING
                                """, syn_id, item["email"], item["source"], batch_id_meg)
                                imported_meg += 1
                            except Exception as e:
                                print(f"[Daily Sync] Meg-identified import error for {item['email']}: {e}")
                    print(f"[Daily Sync] Auto-imported {imported_meg} Meg-identified users (batch={batch_id_meg})")

                # Auto-import Stripe-identified users with correct source
                if stripe_identified:
                    batch_id_stripe = f"autosync_stripe_{run_id}_{datetime.now(timezone.utc).strftime('%Y%m%d')}"
                    imported_stripe = 0
                    async with db_pool.acquire() as conn:
                        for item in stripe_identified:
                            try:
                                await conn.execute("""
                                    INSERT INTO subscriptions (
                                        stripe_customer_id, stripe_subscription_id, email, status,
                                        plan_interval, plan_amount, currency, source,
                                        created_at, updated_at, import_batch
                                    ) VALUES ($1, $2, $3, 'active', 'month', 1999, 'usd', 'stripe', NOW(), NOW(), $4)
                                    ON CONFLICT (stripe_subscription_id) DO NOTHING
                                """, item["stripe_customer_id"], item["stripe_sub_id"], item["email"], batch_id_stripe)
                                imported_stripe += 1
                            except Exception as e:
                                print(f"[Daily Sync] Stripe cross-ref import error for {item['email']}: {e}")
                    print(f"[Daily Sync] Auto-imported {imported_stripe} Stripe-identified users (batch={batch_id_stripe})")

                # Import remaining as undetermined
                if non_stripe:
                    batch_id_undet = f"autosync_undetermined_{run_id}_{datetime.now(timezone.utc).strftime('%Y%m%d')}"
                    imported_undet = 0
                    async with db_pool.acquire() as conn:
                        for item in non_stripe:
                            email = item["email"]
                            email_hash = hashlib.md5(email.encode()).hexdigest()[:16]
                            syn_id = f"ymove_new_undetermined_{email_hash}"
                            try:
                                await conn.execute("""
                                    INSERT INTO subscriptions (
                                        stripe_customer_id, stripe_subscription_id, email, status,
                                        plan_interval, plan_amount, currency, source,
                                        created_at, updated_at, import_batch
                                    ) VALUES ('', $1, $2, 'active', 'month', 1999, 'usd', 'undetermined', NOW(), NOW(), $3)
                                    ON CONFLICT (stripe_subscription_id) DO NOTHING
                                """, syn_id, email, batch_id_undet)
                                imported_undet += 1
                            except Exception as e:
                                print(f"[Daily Sync] Undetermined import error for {email}: {e}")
                    print(f"[Daily Sync] Imported {imported_undet} undetermined users (batch={batch_id_undet})")

                # Log summary
                print(f"[Daily Sync] Gap report: {len(all_unknown)} users in ymove not in our DB. "
                      f"Meg-identified: {len(meg_identified)}, Stripe: {len(stripe_identified)}, "
                      f"Undetermined: {len(non_stripe)}, Stripe errors: {stripe_errors}")
            else:
                print("[Daily Sync] No unknown users found. Webhook pipeline is healthy.")
        else:
            status = run["status"] if run else "unknown"
            error = run.get("error", "") if run else ""
            print(f"[Daily Sync] Sync did not complete successfully. Status: {status}, Error: {error}")

    except Exception as e:
        print(f"[Daily Sync] Error: {e}")
        import traceback
        traceback.print_exc()


# --- S23: Retroactive Provider Cleanup ---

@app.post("/api/admin/cancel-all-duplicates")
async def cancel_all_duplicates(request: Request):
    """S23: Auto-find and safely cancel all duplicate active subscriptions.
    For each email with multiple active records, keeps the one with a real ID
    pattern (sub_*, numeric Apple, ym_google_*) and cancels the synthetic one(s)."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    body = await request.json()
    if not body.get("confirm"):
        return JSONResponse(status_code=400, content={"error": "Provide {confirm: true}"})

    batch_id = f"s23_provider_cleanup_dedup_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

    async with db_pool.acquire() as conn:
        # Get all active records for emails that have duplicates
        rows = await conn.fetch("""
            SELECT s.id, lower(s.email) as email, s.stripe_subscription_id, s.source, s.created_at
            FROM subscriptions s
            INNER JOIN (
                SELECT lower(email) as em FROM subscriptions
                WHERE status IN ('active', 'trialing') AND email != '' AND email IS NOT NULL
                GROUP BY lower(email) HAVING COUNT(*) > 1
            ) dups ON lower(s.email) = dups.em
            WHERE s.status IN ('active', 'trialing')
            ORDER BY s.email, s.created_at
        """)

        # Group by email
        groups = {}
        for r in rows:
            groups.setdefault(r["email"], []).append(dict(r))

        cancelled = []
        kept = []
        skipped = []

        def is_real(sub_id):
            if not sub_id:
                return False
            if sub_id.startswith("sub_"):
                return True
            if sub_id.startswith("ym_google_"):
                return True
            if sub_id.isdigit():
                return True
            return False

        for email, records in groups.items():
            real_records = [r for r in records if is_real(r["stripe_subscription_id"])]
            synthetic_records = [r for r in records if not is_real(r["stripe_subscription_id"])]

            if not real_records:
                # No real sibling, leave alone
                skipped.append({"email": email, "reason": "no real sibling found", "records": records})
                continue

            # Keep first real one
            keep_record = real_records[0]
            kept.append({"email": email, "kept_id": keep_record["id"], "kept_sub": keep_record["stripe_subscription_id"]})

            # Cancel everything else (other real ones AND synthetic ones)
            to_cancel = [r for r in records if r["id"] != keep_record["id"]]
            for r in to_cancel:
                try:
                    await conn.execute(
                        """UPDATE subscriptions SET status = 'canceled', canceled_at = NOW(),
                           import_batch = $1, updated_at = NOW() WHERE id = $2""",
                        batch_id, r["id"]
                    )
                    cancelled.append({"email": email, "cancelled_id": r["id"], "sub": r["stripe_subscription_id"], "source": r["source"]})
                except Exception as e:
                    skipped.append({"email": email, "id": r["id"], "error": str(e)})

    return {
        "status": "done",
        "batch_id": batch_id,
        "duplicate_emails_processed": len(groups),
        "records_cancelled": len(cancelled),
        "records_kept": len(kept),
        "skipped": len(skipped),
        "cancelled": cancelled,
        "kept": kept,
        "skipped_details": skipped,
        "revert": f"POST /api/admin/revert-batch with batch_id={batch_id}"
    }


@app.post("/api/admin/cancel-duplicate")
async def cancel_duplicate(request: Request):
    """S23: Safely cancel a duplicate record by ID (preserves history).
    Body: {"id": 123, "confirm": true}
    Only cancels if a real sibling exists for the same email.
    Tags with import_batch='s23_dedup_safe_cancel' so shadow sync won't reactivate."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    body = await request.json()
    record_id = body.get("id")
    confirm = body.get("confirm", False)
    if not record_id or not confirm:
        return JSONResponse(status_code=400, content={"error": "Provide {id: int, confirm: true}"})

    batch_id = f"s23_provider_cleanup_dedup_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

    async with db_pool.acquire() as conn:
        target = await conn.fetchrow(
            "SELECT id, email, stripe_subscription_id, source, status FROM subscriptions WHERE id = $1",
            record_id
        )
        if not target:
            return {"status": "not_found", "id": record_id}

        sibling = await conn.fetchrow(
            """SELECT id, stripe_subscription_id, source FROM subscriptions
               WHERE lower(email) = lower($1) AND id != $2
               AND status IN ('active', 'trialing')
               AND (stripe_subscription_id LIKE 'sub_%'
                    OR stripe_subscription_id ~ '^[0-9]+$'
                    OR stripe_subscription_id LIKE 'ym_google_%')
               LIMIT 1""", target["email"], record_id
        )
        if not sibling:
            return {"status": "blocked", "reason": "No real sibling record found for this email. Refusing to cancel.",
                    "target": dict(target)}

        await conn.execute(
            """UPDATE subscriptions
               SET status = 'canceled', canceled_at = NOW(),
                   import_batch = $1, updated_at = NOW()
               WHERE id = $2""",
            batch_id, record_id
        )
        return {
            "status": "cancelled",
            "batch_id": batch_id,
            "cancelled_record": dict(target),
            "preserved_sibling": dict(sibling),
            "revert": f"POST /api/admin/revert-batch with batch_id={batch_id}"
        }


@app.post("/api/admin/delete-duplicate")
async def delete_duplicate(request: Request):
    """S23: Hard delete a specific duplicate record by ID. Use with caution.
    Body: {"id": 123, "confirm": true}
    Use Reconciliation Audit duplicate list to find IDs to delete.
    Only deletes if a real sub_* / numeric / ym_google_* sibling exists for the same email."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    body = await request.json()
    record_id = body.get("id")
    confirm = body.get("confirm", False)
    if not record_id or not confirm:
        return JSONResponse(status_code=400, content={"error": "Provide {id: int, confirm: true}"})

    async with db_pool.acquire() as conn:
        target = await conn.fetchrow(
            "SELECT id, email, stripe_subscription_id, source, status FROM subscriptions WHERE id = $1",
            record_id
        )
        if not target:
            return {"status": "not_found", "id": record_id}

        # Safety: only allow delete if sibling exists for same email with real ID pattern
        sibling = await conn.fetchrow(
            """SELECT id, stripe_subscription_id, source FROM subscriptions
               WHERE lower(email) = lower($1) AND id != $2
               AND status IN ('active', 'trialing')
               AND (stripe_subscription_id LIKE 'sub_%'
                    OR stripe_subscription_id ~ '^[0-9]+$'
                    OR stripe_subscription_id LIKE 'ym_google_%')
               LIMIT 1""", target["email"], record_id
        )
        if not sibling:
            return {"status": "blocked", "reason": "No real sibling record found for this email. Refusing to delete to prevent data loss.",
                    "target": dict(target)}

        await conn.execute("DELETE FROM subscriptions WHERE id = $1", record_id)
        return {
            "status": "deleted",
            "deleted_record": dict(target),
            "preserved_sibling": dict(sibling)
        }


@app.post("/api/admin/cleanup-manual-duplicates")
async def cleanup_manual_duplicates(request: Request):
    """S24: Clean up manual/manual duplicates where neither record is a 'real sibling'.

    The existing cancel-all-duplicates endpoint refuses to operate when both
    records are synthetic (ymove_new_*, import_*, meg_*) because there's no
    real Stripe sub_*, Apple numeric, or ym_google_* to anchor on. That's
    intentional safety, but it leaves a gap for the specific case where both
    duplicates are correctly classified manual records that just happen to
    have different synthetic IDs (e.g. ymove_new_apple_X and ymove_new_undetermined_X
    for the same email hash, created by separate sync runs before the Step 0 fix).

    This endpoint handles that exact case with strict guards:
      - Both records must have status IN ('active', 'trialing')
      - Both records must have source = 'manual'
      - Both records must have a synthetic ID (not sub_*, not numeric, not ym_google_*)
      - Email must be the same

    Action: keep the OLDER record (preserves history), cancel the newer one
    with batch tag s24_manual_dedup_<timestamp>. Future shadow syncs will
    skip these via the s24_% reactivation guard added in this same session.

    Body: {preview: true|false}
      preview=true (default): show planned actions, no writes
      preview=false: apply
    """
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    try:
        body = await request.json()
    except Exception:
        body = {}
    preview = body.get("preview", True)
    batch_id = f"s24_manual_dedup_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

    def is_synthetic(sub_id: str) -> bool:
        if not sub_id:
            return True
        if sub_id.startswith("sub_"):
            return False
        if sub_id.startswith("ym_google_"):
            return False
        if sub_id.isdigit():
            return False
        return True

    async with db_pool.acquire() as conn:
        # Find emails with multiple active manual records
        groups = await conn.fetch("""
            SELECT lower(email) as email, COUNT(*) as cnt
            FROM subscriptions
            WHERE status IN ('active', 'trialing')
              AND source = 'manual'
              AND email IS NOT NULL AND email != ''
            GROUP BY lower(email)
            HAVING COUNT(*) > 1
        """)

        plan = []
        for g in groups:
            email = g["email"]
            records = await conn.fetch("""
                SELECT id, email, stripe_subscription_id, source, status, created_at
                FROM subscriptions
                WHERE lower(email) = $1
                  AND status IN ('active', 'trialing')
                  AND source = 'manual'
                ORDER BY created_at ASC
            """, email)
            records = [dict(r) for r in records]

            # Strict guard: every record in the group must be synthetic
            if not all(is_synthetic(r["stripe_subscription_id"]) for r in records):
                plan.append({
                    "email": email,
                    "action": "skip",
                    "reason": "at least one record is a real sibling, refusing to touch",
                    "records": [
                        {**r, "created_at": str(r["created_at"])} for r in records
                    ],
                })
                continue

            keeper = records[0]   # oldest by created_at ASC
            cancellees = records[1:]
            plan.append({
                "email": email,
                "action": "cancel_newer",
                "keep": {
                    "id": keeper["id"],
                    "stripe_subscription_id": keeper["stripe_subscription_id"],
                    "created_at": str(keeper["created_at"]),
                },
                "cancel": [
                    {
                        "id": c["id"],
                        "stripe_subscription_id": c["stripe_subscription_id"],
                        "created_at": str(c["created_at"]),
                    }
                    for c in cancellees
                ],
            })

        if preview:
            return {
                "status": "preview",
                "batch_id": batch_id,
                "groups_found": len(groups),
                "actionable": sum(1 for p in plan if p["action"] == "cancel_newer"),
                "skipped": sum(1 for p in plan if p["action"] == "skip"),
                "plan": plan,
                "note": "Set preview: false to apply.",
            }

        # APPLY mode
        cancelled_ids = []
        errors = []
        for p in plan:
            if p["action"] != "cancel_newer":
                continue
            for c in p["cancel"]:
                try:
                    await conn.execute(
                        """UPDATE subscriptions
                           SET status = 'canceled',
                               canceled_at = NOW(),
                               import_batch = $1,
                               updated_at = NOW()
                           WHERE id = $2""",
                        batch_id, c["id"]
                    )
                    cancelled_ids.append(c["id"])
                except Exception as e:
                    errors.append({"id": c["id"], "error": str(e)})

        return {
            "status": "applied",
            "batch_id": batch_id,
            "cancelled_count": len(cancelled_ids),
            "cancelled_ids": cancelled_ids,
            "errors": errors,
            "revert": f"POST /api/admin/revert-batch with batch_id={batch_id}",
        }


@app.post("/api/admin/backfill-utms")
async def backfill_utms(request: Request):
    """S24: Backfill UTM attribution on existing Stripe subscriptions.

    Tosh shipped the meta-field passthrough on 2026-04-11, which means
    subscriptionProvider is populated AND member-lookup now returns UTM params
    in the meta field. New Stripe webhooks already capture UTMs via the existing
    pipeline in _ymove_handle_created. This endpoint backfills the ~1000 Stripe
    subs that were created before the fix.

    Scope: Stripe-only. Apple/Google subs go through app stores which strip UTMs,
    so they are not relevant for ad attribution.

    Body: {preview: true|false, limit: int (optional), status_filter: "active"|"all"}
      preview=true (default): count + sample, no writes
      preview=false: actually call ymove and write UTMs
      limit: cap rows processed (useful for staged runs)
      status_filter:
        "active" (default): only active+trialing Stripe subs (~1k records, ~4 min)
        "all": every Stripe sub including cancelled (~6k records, ~25 min)
               useful for future historical analysis of churn by channel

    Idempotent: WHERE filter excludes rows that already have utm_source set,
    and _store_utm_on_subscription uses COALESCE so it cannot blank existing data.
    """
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    if not YMOVE_API_KEY:
        raise HTTPException(status_code=500, detail="YMOVE_API_KEY not configured")

    try:
        body = await request.json()
    except Exception:
        body = {}
    preview = body.get("preview", True)
    limit = body.get("limit")
    status_filter = body.get("status_filter", "active")
    if status_filter not in ("active", "all"):
        return JSONResponse(status_code=400, content={
            "error": "status_filter must be 'active' or 'all'"
        })
    batch_id = f"s24_utm_backfill_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

    async with db_pool.acquire() as conn:
        # Stripe subs only, no UTMs yet, must have email to look up
        status_clause = "AND status IN ('active', 'trialing')" if status_filter == "active" else ""
        query = f"""
            SELECT id, stripe_subscription_id, email, status, created_at
            FROM subscriptions
            WHERE stripe_subscription_id LIKE 'sub_%'
              AND email IS NOT NULL
              AND email != ''
              AND (utm_source IS NULL OR utm_source = '')
              {status_clause}
            ORDER BY created_at DESC
        """
        if limit and isinstance(limit, int) and limit > 0:
            query += f" LIMIT {int(limit)}"

        rows = await conn.fetch(query)
        candidate_count = len(rows)

        if preview:
            return {
                "status": "preview",
                "batch_id": batch_id,
                "candidates": candidate_count,
                "limit_applied": limit,
                "status_filter": status_filter,
                "sample": [
                    {
                        "id": r["id"],
                        "stripe_subscription_id": r["stripe_subscription_id"],
                        "email": r["email"],
                        "status": r["status"],
                        "created_at": str(r["created_at"]),
                    }
                    for r in rows[:5]
                ],
                "note": "Set preview: false to apply. Estimated runtime: ~"
                        + str(round(candidate_count * 0.25)) + "s at 4 lookups/sec.",
            }

        # APPLY mode
        processed = 0
        found_with_utm = 0
        found_no_utm = 0
        not_found = 0
        errors = 0
        attached_samples = []

        for r in rows:
            sub_id = r["stripe_subscription_id"]
            email = r["email"]
            try:
                utm_result = await _ymove_lookup_utm(email)
                processed += 1
                if not utm_result.get("found"):
                    not_found += 1
                    continue
                if utm_result.get("utm"):
                    await _store_utm_on_subscription(conn, sub_id, utm_result)
                    found_with_utm += 1
                    if len(attached_samples) < 10:
                        attached_samples.append({
                            "stripe_subscription_id": sub_id,
                            "email": email,
                            "utm": utm_result["utm"],
                        })
                else:
                    found_no_utm += 1
            except Exception as e:
                errors += 1
                print(f"[s24-backfill-utms] Error on {sub_id} ({email}): {e}")

            # Be polite to ymove API: ~4 req/sec
            await asyncio.sleep(0.05)

        # Tag the batch on rows we successfully attached UTMs to
        if found_with_utm > 0:
            try:
                await conn.execute(
                    """UPDATE subscriptions
                       SET import_batch = COALESCE(import_batch, '') || $1,
                           updated_at = NOW()
                       WHERE stripe_subscription_id LIKE 'sub_%'
                         AND utm_source IS NOT NULL
                         AND utm_source != ''
                         AND updated_at > NOW() - INTERVAL '5 minutes'""",
                    f" {batch_id}"
                )
            except Exception as e:
                print(f"[s24-backfill-utms] Batch tag error: {e}")

        return {
            "status": "applied",
            "batch_id": batch_id,
            "candidates": candidate_count,
            "processed": processed,
            "found_with_utm": found_with_utm,
            "found_no_utm": found_no_utm,
            "not_found": not_found,
            "errors": errors,
            "attached_samples": attached_samples,
        }


@app.post("/api/admin/provider-cleanup")
async def provider_cleanup(request: Request):
    """S23: Reclassify ymove_new_*/ymove_switch_* records that were auto-imported with guessed providers.
    Step 1: Check if we already have a real sub_* Stripe record (cancel duplicate)
    Step 2: Check Meg import records for known Apple/Google source
    Step 3: Cross-reference with Stripe API
    Step 4: Report remaining undetermined for future resolution (when Tosh fixes API)
    Body: {"preview": true/false}
    """
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    body = await request.json()
    preview = body.get("preview", True)
    batch_id = f"s23_provider_cleanup_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

    async with db_pool.acquire() as conn:
        # Find all active ymove_new_* and ymove_switch_* records
        candidates = await conn.fetch("""
            SELECT id, email, stripe_subscription_id, source, status, import_batch
            FROM subscriptions
            WHERE status IN ('active', 'trialing')
            AND (stripe_subscription_id LIKE 'ymove_new_%' OR stripe_subscription_id LIKE 'ymove_switch_%')
            ORDER BY created_at
        """)

    if not candidates:
        return {"status": "ok", "message": "No ymove_new_*/ymove_switch_* active records found. Nothing to clean up."}

    stripe_duplicates = []  # Have real sub_* record already — cancel the ymove_new_* copy
    apple_google_duplicates = []  # Have real Apple transactionId or ym_google_* record — cancel the ymove_new_* copy
    meg_identified = []     # Meg import tells us the provider
    stripe_orphans = []     # Stripe confirms active but we don't have a sub_* record
    non_stripe = []
    stripe_errors = 0

    async with db_pool.acquire() as conn:
        for r in candidates:
            email = (r["email"] or "").strip().lower()
            if not email:
                non_stripe.append({"id": r["id"], "email": "", "sub_id": r["stripe_subscription_id"],
                                   "old_source": r["source"], "new_source": "undetermined", "note": "No email to cross-ref"})
                continue

            # Step 1: do we already have a sub_* record for this email?
            existing_stripe = await conn.fetchrow(
                """SELECT id, stripe_subscription_id, status FROM subscriptions
                   WHERE lower(email) = $1 AND stripe_subscription_id LIKE 'sub_%'
                   AND status IN ('active', 'trialing')""", email
            )
            if existing_stripe:
                stripe_duplicates.append({
                    "id": r["id"], "email": email, "sub_id": r["stripe_subscription_id"],
                    "old_source": r["source"], "action": "cancel_duplicate",
                    "real_sub_id": existing_stripe["stripe_subscription_id"],
                    "note": f"Duplicate of Stripe sub: {existing_stripe['stripe_subscription_id']}"
                })
                continue

            # Step 1b: do we already have a real Apple (numeric ID) or Google (ym_google_*) record?
            existing_ag = await conn.fetchrow(
                """SELECT id, stripe_subscription_id, source, status FROM subscriptions
                   WHERE lower(email) = $1
                   AND id != $2
                   AND status IN ('active', 'trialing')
                   AND (
                       (stripe_subscription_id ~ '^[0-9]+$')
                       OR (stripe_subscription_id LIKE 'ym_google_%%')
                   )
                   ORDER BY created_at DESC LIMIT 1""", email, r["id"]
            )
            if existing_ag:
                apple_google_duplicates.append({
                    "id": r["id"], "email": email, "sub_id": r["stripe_subscription_id"],
                    "old_source": r["source"], "action": "cancel_duplicate",
                    "real_sub_id": existing_ag["stripe_subscription_id"],
                    "real_source": existing_ag["source"],
                    "note": f"Duplicate of {existing_ag['source']} sub: {existing_ag['stripe_subscription_id']}"
                })
                continue

            # Step 2: do we have a Meg import record for this email? (any status)
            meg_record = await conn.fetchrow(
                """SELECT source, stripe_subscription_id FROM subscriptions
                   WHERE lower(email) = $1
                   AND (stripe_subscription_id LIKE 'import_apple_%'
                        OR stripe_subscription_id LIKE 'import_google_%'
                        OR stripe_subscription_id LIKE 'meg_apple_%'
                        OR stripe_subscription_id LIKE 'meg_google_%')
                   ORDER BY created_at DESC LIMIT 1""", email
            )
            if meg_record:
                meg_source = meg_record["source"]
                if meg_source in ("apple", "google"):
                    meg_identified.append({
                        "id": r["id"], "email": email, "sub_id": r["stripe_subscription_id"],
                        "old_source": r["source"], "new_source": meg_source,
                        "meg_sub_id": meg_record["stripe_subscription_id"],
                        "note": f"Meg import confirms source={meg_source} ({meg_record['stripe_subscription_id']})"
                    })
                    continue

            # Step 3: does Stripe know this email?
            try:
                customers = stripe.Customer.list(email=email, limit=1)
                if customers.data:
                    cust = customers.data[0]
                    subs = stripe.Subscription.list(customer=cust.id, status="active", limit=1)
                    if subs.data:
                        stripe_orphans.append({
                            "id": r["id"], "email": email, "sub_id": r["stripe_subscription_id"],
                            "old_source": r["source"], "action": "reclassify_stripe",
                            "stripe_customer_id": cust.id, "stripe_sub_id": subs.data[0].id,
                            "note": "Active Stripe sub confirmed but no sub_* record in our DB"
                        })
                        continue
                non_stripe.append({"id": r["id"], "email": email, "sub_id": r["stripe_subscription_id"],
                                   "old_source": r["source"], "new_source": "undetermined",
                                   "note": "No Stripe sub, no Meg import — Apple/Google/Manual"})
            except Exception as e:
                stripe_errors += 1
                non_stripe.append({"id": r["id"], "email": email, "sub_id": r["stripe_subscription_id"],
                                   "old_source": r["source"], "new_source": "undetermined",
                                   "note": f"Stripe lookup error: {str(e)[:100]}"})

    if preview:
        return {
            "status": "preview",
            "batch_id": batch_id,
            "total_candidates": len(candidates),
            "stripe_duplicates_to_cancel": len(stripe_duplicates),
            "apple_google_duplicates_to_cancel": len(apple_google_duplicates),
            "meg_identified": len(meg_identified),
            "stripe_orphans_to_reclassify": len(stripe_orphans),
            "non_stripe_to_undetermined": len(non_stripe),
            "stripe_errors": stripe_errors,
            "duplicate_details": stripe_duplicates,
            "ag_duplicate_details": apple_google_duplicates,
            "meg_details": meg_identified,
            "orphan_details": stripe_orphans,
            "undetermined_details": non_stripe,
            "note": "Send preview: false to apply. All duplicates get cancelled. Meg-identified get correct source. Stripe orphans get source='stripe'. Rest get source='undetermined'."
        }

    # Apply changes
    cancelled_dupes = 0
    updated_meg = 0
    updated_stripe = 0
    updated_undet = 0
    errors = 0

    async with db_pool.acquire() as conn:
        # Cancel Stripe duplicates
        for item in stripe_duplicates:
            try:
                await conn.execute(
                    "UPDATE subscriptions SET status = 'canceled', canceled_at = NOW(), import_batch = $1, updated_at = NOW() WHERE id = $2",
                    batch_id, item["id"]
                )
                cancelled_dupes += 1
            except Exception as e:
                errors += 1

        # Cancel Apple/Google duplicates
        for item in apple_google_duplicates:
            try:
                await conn.execute(
                    "UPDATE subscriptions SET status = 'canceled', canceled_at = NOW(), import_batch = $1, updated_at = NOW() WHERE id = $2",
                    batch_id, item["id"]
                )
                cancelled_dupes += 1
            except Exception as e:
                errors += 1

        # Apply Meg-identified source
        for item in meg_identified:
            try:
                await conn.execute(
                    "UPDATE subscriptions SET source = $1, import_batch = $2, updated_at = NOW() WHERE id = $3",
                    item["new_source"], batch_id, item["id"]
                )
                updated_meg += 1
            except Exception as e:
                errors += 1

        # Reclassify Stripe orphans
        for item in stripe_orphans:
            try:
                await conn.execute(
                    "UPDATE subscriptions SET source = 'stripe', import_batch = $1, updated_at = NOW() WHERE id = $2",
                    batch_id, item["id"]
                )
                updated_stripe += 1
            except Exception as e:
                errors += 1

        # Reclassify non-Stripe to undetermined
        for item in non_stripe:
            try:
                await conn.execute(
                    "UPDATE subscriptions SET source = 'undetermined', import_batch = $1, updated_at = NOW() WHERE id = $2",
                    batch_id, item["id"]
                )
                updated_undet += 1
            except Exception as e:
                errors += 1

    return {
        "status": "applied",
        "batch_id": batch_id,
        "total_processed": len(candidates),
        "cancelled_duplicates": cancelled_dupes,
        "updated_from_meg": updated_meg,
        "updated_to_stripe": updated_stripe,
        "updated_to_undetermined": updated_undet,
        "errors": errors,
        "revert": f"POST /api/admin/revert-batch with batch_id={batch_id}"
    }


@app.post("/api/admin/send-test-digest")
async def send_test_digest(request: Request):
    """Manually trigger a daily digest email (for testing)."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)

    if not RESEND_API_KEY:
        raise HTTPException(status_code=500, detail="RESEND_API_KEY not configured")
    if not DIGEST_RECIPIENTS:
        raise HTTPException(status_code=500, detail="DIGEST_RECIPIENTS not configured")

    stats = await gather_daily_stats()
    if "error" in stats:
        raise HTTPException(status_code=500, detail=stats["error"])

    insights = await generate_digest_insights(stats)

    now_et = datetime.now(ZoneInfo("America/New_York"))
    subject = f"[TEST] M&M Daily Digest - {now_et.strftime('%b %d')} | {stats.get('conversions_today',0)} conversions, {stats.get('new_subscriptions',0)} trials, {stats.get('cancellations_paid',0)} paid cancels, {stats.get('cancellations_trial',0)} trial cancels"

    html = build_digest_html(stats, insights)
    result = await send_digest_email(html, subject)

    if "error" in result:
        raise HTTPException(status_code=500, detail=result["error"])

    return {"status": "sent", "recipients": result["recipients"], "stats_summary": {
        "new_subs": stats.get("new_subscriptions", 0),
        "cancellations": stats.get("cancellations", 0),
        "new_leads": stats.get("new_leads", 0),
        "gross_mrr": stats.get("gross_mrr", "$0"),
        "net_mrr": stats.get("net_mrr", "$0"),
    }}


# --- Admin Data Tools ---


@app.post("/api/admin/backfill-stripe")
async def backfill_stripe(request: Request):
    """Pull all Stripe subscriptions and upsert into DB."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    if not STRIPE_SECRET_KEY:
        raise HTTPException(status_code=500, detail="Stripe not configured")

    count = 0
    errors = 0
    try:
        subs_iter = stripe.Subscription.list(
            limit=100, status="all", expand=["data.customer"]
        )
        for sub in subs_iter.auto_paging_iter():
            try:
                sub_id = sub.id
                cust_obj = sub.get("customer") if isinstance(sub, dict) else sub.customer
                customer_id = cust_obj.get("id", str(cust_obj)) if isinstance(cust_obj, dict) else (cust_obj.id if hasattr(cust_obj, "id") else str(cust_obj))
                status = sub.status

                plan_amount = 0
                plan_interval = ""
                try:
                    items_data = sub["items"]["data"]
                    if items_data:
                        price = items_data[0]["price"]
                        plan_amount = price.get("unit_amount", 0) or 0
                        recurring = price.get("recurring")
                        if recurring:
                            plan_interval = recurring.get("interval", "") or ""
                except Exception:
                    pass

                email = ""
                try:
                    cust = sub.get("customer")
                    if isinstance(cust, dict):
                        email = cust.get("email", "") or ""
                    elif hasattr(cust, "email"):
                        email = cust.email or ""
                except Exception:
                    pass


                real_created = _ts(sub.created) if hasattr(sub, 'created') else None

                async with db_pool.acquire() as conn:
                    await conn.execute("""
                        INSERT INTO subscriptions (
                            stripe_customer_id, stripe_subscription_id, email, status,
                            plan_interval, plan_amount, currency, source,
                            trial_start, trial_end,
                            current_period_start, current_period_end,
                            canceled_at, created_at, updated_at
                        ) VALUES ($1,$2,$3,$4,$5,$6,$7,'stripe',$8,$9,$10,$11,$12,COALESCE($13,NOW()),NOW())
                        ON CONFLICT (stripe_subscription_id) DO UPDATE SET
                            status = EXCLUDED.status,
                            plan_interval = EXCLUDED.plan_interval,
                            plan_amount = EXCLUDED.plan_amount,
                            email = EXCLUDED.email,
                            trial_start = EXCLUDED.trial_start,
                            trial_end = EXCLUDED.trial_end,
                            current_period_start = EXCLUDED.current_period_start,
                            current_period_end = EXCLUDED.current_period_end,
                            canceled_at = EXCLUDED.canceled_at,
                            created_at = EXCLUDED.created_at,
                            updated_at = NOW()
                    """,
                        customer_id, sub_id, email, status,
                        plan_interval, plan_amount, sub.currency or "usd",
                        _ts(sub.trial_start), _ts(sub.trial_end),
                        _ts(sub.current_period_start), _ts(sub.current_period_end),
                        _ts(sub.canceled_at), real_created
                    )
                count += 1
            except Exception as e:
                print(f"Backfill sub error: {e}")
                errors += 1
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Stripe API error: {str(e)}")

    return {"status": "ok", "imported": count, "errors": errors}


@app.post("/api/admin/sync-trialing")
async def sync_trialing(request: Request):
    """Sync all Stripe subs stuck as 'trialing' with Stripe's real current status.
    Fixes zombie records from before the webhook was connected."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    if not STRIPE_SECRET_KEY:
        raise HTTPException(status_code=500, detail="Stripe not configured")

    async with db_pool.acquire() as conn:
        zombies = await conn.fetch(
            "SELECT id, stripe_subscription_id FROM subscriptions WHERE status = 'trialing' AND source = 'stripe'"
        )

    results = {"total_checked": len(zombies), "now_active": 0, "now_canceled": 0, "now_past_due": 0, "still_trialing": 0, "not_found": 0, "errors": 0}

    for z in zombies:
        sub_id = z["stripe_subscription_id"]
        if not sub_id or not sub_id.startswith("sub_"):
            results["not_found"] += 1
            continue
        try:
            real_sub = stripe.Subscription.retrieve(sub_id)
            real_status = real_sub.status
            if real_status == "trialing":
                results["still_trialing"] += 1
                continue


            async with db_pool.acquire() as conn:
                await conn.execute(
                    """UPDATE subscriptions SET
                       status = $1,
                       canceled_at = $2,
                       current_period_start = $3,
                       current_period_end = $4,
                       updated_at = NOW()
                       WHERE id = $5""",
                    real_status,
                    _ts(real_sub.canceled_at),
                    _ts(real_sub.current_period_start),
                    _ts(real_sub.current_period_end),
                    z["id"]
                )

                # If now active, they converted from trial and we missed it
                if real_status == "active":
                    await conn.execute(
                        "UPDATE subscriptions SET converted_at = COALESCE(trial_end, NOW()) WHERE id = $1 AND converted_at IS NULL",
                        z["id"]
                    )
                    results["now_active"] += 1
                elif real_status == "canceled":
                    results["now_canceled"] += 1
                elif real_status == "past_due":
                    results["now_past_due"] += 1
                else:
                    # incomplete, incomplete_expired, unpaid, paused
                    results["now_canceled"] += 1

        except stripe.error.InvalidRequestError:
            # Sub doesn't exist in Stripe anymore
            async with db_pool.acquire() as conn:
                await conn.execute(
                    "UPDATE subscriptions SET status = 'canceled', canceled_at = NOW(), updated_at = NOW() WHERE id = $1",
                    z["id"]
                )
            results["not_found"] += 1
        except Exception as e:
            print(f"Sync trialing error for {sub_id}: {e}")
            results["errors"] += 1

    return results


@app.post("/api/admin/fix-stripe-dates")
async def fix_stripe_dates(request: Request):
    """One-time fix: update created_at on all Stripe subs to use Stripe's real created timestamp.
    This corrects the backfill issue where all subs got created_at = NOW() instead of their real date."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    if not STRIPE_SECRET_KEY:
        raise HTTPException(status_code=500, detail="Stripe not configured")

    fixed = 0
    errors = 0
    skipped = 0
    try:
        subs_iter = stripe.Subscription.list(limit=100, status="all")
        for sub in subs_iter.auto_paging_iter():
            try:
                sub_id = sub.id
                stripe_created = sub.created
                if not stripe_created:
                    skipped += 1
                    continue
                real_dt = datetime.fromtimestamp(stripe_created, tz=timezone.utc)
                async with db_pool.acquire() as conn:
                    result = await conn.execute(
                        "UPDATE subscriptions SET created_at = $1 WHERE stripe_subscription_id = $2 AND source = 'stripe'",
                        real_dt, sub_id
                    )
                    if result and result.endswith('1'):
                        fixed += 1
                    else:
                        skipped += 1
            except Exception as e:
                print(f"Fix date error for {sub.id}: {e}")
                errors += 1
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Stripe API error: {str(e)}")

    return {"status": "ok", "fixed": fixed, "skipped": skipped, "errors": errors}


@app.post("/api/admin/reset-data")
async def reset_data(request: Request):
    """Wipe all tables. Requires confirm=RESET in body."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    body = await request.json()
    if body.get("confirm") != "RESET":
        raise HTTPException(status_code=400, detail="Must send confirm=RESET")
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    async with db_pool.acquire() as conn:
        await conn.execute("TRUNCATE leads, page_views, chat_sessions, subscriptions, subscription_events, ad_spend RESTART IDENTITY")

    return {"status": "ok", "message": "All data cleared"}


@app.post("/api/admin/import-leads-csv")
async def import_leads_csv(request: Request, file: UploadFile = File(...)):
    """Import leads from CSV file upload."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")


    contents = await file.read()
    text = contents.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))


    async with db_pool.acquire() as conn:
        existing = await conn.fetch("SELECT DISTINCT lower(email) as em FROM leads WHERE email != ''")
        existing_emails = set(r["em"] for r in existing)

        imported = 0
        skipped = 0
        for row in reader:
            email = _find_col(row, ["email", "e-mail"])
            if not email or email.lower() in existing_emails:
                skipped += 1
                continue
            first_name = _find_col(row, ["first_name", "first", "firstname", "first name"])
            last_name = _find_col(row, ["last_name", "last", "lastname", "last name"])
            referral = _find_col(row, ["source", "referral_source", "referral", "how_heard"])

            await conn.execute(
                """INSERT INTO leads (first_name, email, extra, referral_source)
                   VALUES ($1, $2, $3, $4)""",
                first_name, email, last_name, referral
            )
            imported += 1
            existing_emails.add(email.lower())

    return {"status": "ok", "imported": imported, "skipped": skipped}


@app.post("/api/admin/import-subscribers-csv")
async def import_subscribers_csv(request: Request, file: UploadFile = File(...)):
    """Import Apple/Google subscribers from CSV into subscriptions table."""

    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    contents = await file.read()
    text = contents.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))


    async with db_pool.acquire() as conn:
        existing = await conn.fetch("SELECT stripe_subscription_id FROM subscriptions")
        existing_ids = set(r["stripe_subscription_id"] for r in existing)

        imported = 0
        skipped = 0
        for row in reader:
            email = _find_col(row, ["email", "e-mail"])
            if not email:
                skipped += 1
                continue

            source_raw = _find_col(row, ["source"]).lower()
            if source_raw == "google":
                source = "google"
            else:
                source = "apple"

            email_hash = hashlib.md5(email.lower().encode()).hexdigest()[:16]
            syn_id = f"import_{source}_{email_hash}"

            if syn_id in existing_ids:
                skipped += 1
                continue

            date_str = _find_col(row, ["date", "sign up date", "signup_date"])
            period_start = None
            if date_str:
                try:
                    period_start = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                except Exception:
                    pass

            plan_amount = 1999
            plan_interval = "month"

            try:
                await conn.execute("""
                    INSERT INTO subscriptions (
                        stripe_customer_id, stripe_subscription_id, email, status,
                        plan_interval, plan_amount, currency, source,
                        current_period_start, created_at, updated_at
                    ) VALUES ('', $1, $2, 'active', $3, $4, 'usd', $5, $6, COALESCE($7, NOW()), NOW())
                    ON CONFLICT (stripe_subscription_id) DO UPDATE SET
                        email = EXCLUDED.email,
                        status = EXCLUDED.status,
                        updated_at = NOW()
                """,
                    syn_id, email, plan_interval, plan_amount, source, period_start, period_start
                )
                imported += 1
                existing_ids.add(syn_id)
            except Exception as e:
                print(f"Subscriber import error: {e}")
                skipped += 1

    return {"status": "ok", "imported": imported, "skipped": skipped}


@app.post("/api/admin/reconcile-cancellations")
async def reconcile_cancellations(request: Request, file: UploadFile = File(...)):
    """Upload CSV of cancelled emails. Matches against subscriptions and flips to canceled."""

    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    contents = await file.read()
    text = contents.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    cancelled_emails = set()
    for row in reader:
        for key in row:
            if key.strip().lower() in ("email", "e-mail"):
                val = (row[key] or "").strip().lower()
                if val and "@" in val:
                    cancelled_emails.add(val)

    if not cancelled_emails:
        return {"status": "ok", "matched": 0, "total_emails": 0}

    async with db_pool.acquire() as conn:
        active_subs = await conn.fetch(
            "SELECT id, email FROM subscriptions WHERE status IN ('active', 'trialing')"
        )
        matched = 0
        for sub in active_subs:
            if sub["email"] and sub["email"].strip().lower() in cancelled_emails:
                await conn.execute(
                    "UPDATE subscriptions SET status = 'canceled', canceled_at = NOW(), updated_at = NOW() WHERE id = $1",
                    sub["id"]
                )
                matched += 1

    return {"status": "ok", "matched": matched, "total_emails": len(cancelled_emails)}


@app.post("/api/admin/backfill-trial-dates")
async def backfill_trial_dates(request: Request):
    """Sets trial_start/end on imported subs missing trial data."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    async with db_pool.acquire() as conn:
        updated = await conn.execute(
            """UPDATE subscriptions
               SET trial_start = created_at,
                   trial_end = created_at + INTERVAL '30 days'
               WHERE trial_start IS NULL
               AND source IN ('apple', 'google')
               AND created_at IS NOT NULL"""
        )

    count = int(updated.split(" ")[-1]) if updated else 0
    return {"status": "ok", "updated": count}


# --- Session 11: Backfill converted_at for historical data ---

@app.post("/api/admin/backfill-conversions")
async def backfill_conversions(request: Request):
    """One-time backfill: stamp converted_at on subs that clearly converted from trial.
    Uses heuristic: had a trial (trial_end set) AND current_period_start moved past trial_end."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    async with db_pool.acquire() as conn:
        # Stripe subs: trial_end exists and current_period moved past it
        stripe_result = await conn.execute(
            """UPDATE subscriptions
               SET converted_at = trial_end
               WHERE converted_at IS NULL
               AND trial_end IS NOT NULL
               AND current_period_start IS NOT NULL
               AND current_period_start > trial_end
               AND source = 'stripe'""")
        stripe_count = int(stripe_result.split(" ")[-1]) if stripe_result else 0

        # Apple/Google imported subs: if active and trial_end is set, they converted
        appgoogle_result = await conn.execute(
            """UPDATE subscriptions
               SET converted_at = trial_end
               WHERE converted_at IS NULL
               AND trial_end IS NOT NULL
               AND status = 'active'
               AND source IN ('apple', 'google')""")
        appgoogle_count = int(appgoogle_result.split(" ")[-1]) if appgoogle_result else 0

    return {
        "status": "ok",
        "stripe_backfilled": stripe_count,
        "apple_google_backfilled": appgoogle_count,
        "total": stripe_count + appgoogle_count,
    }


@app.post("/api/admin/backfill-readable-ids")
async def backfill_readable_ids(request: Request):
    """Assign persistent readable_id to all existing subscriptions that lack one."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    async with db_pool.acquire() as conn:
        # Get all subs without readable_id, ordered by created_at so IDs are chronological
        rows = await conn.fetch(
            """SELECT id, stripe_subscription_id, source FROM subscriptions
               WHERE readable_id IS NULL
               ORDER BY source, created_at"""
        )

        # Get current max ID per source
        counters = {}
        for src in ("stripe", "apple", "google"):
            row = await conn.fetchrow(
                "SELECT readable_id FROM subscriptions WHERE source = $1 AND readable_id IS NOT NULL ORDER BY readable_id DESC LIMIT 1",
                src
            )
            if row and row["readable_id"]:
                try:
                    counters[src] = int(row["readable_id"].split("-")[-1])
                except (ValueError, IndexError):
                    counters[src] = 0
            else:
                counters[src] = 0

        assigned = 0
        for r in rows:
            src = r["source"] or "stripe"
            counters[src] = counters.get(src, 0) + 1
            rid = f"{src.upper()}-{counters[src]:04d}"
            await conn.execute(
                "UPDATE subscriptions SET readable_id = $1 WHERE id = $2",
                rid, r["id"]
            )
            assigned += 1

    return {
        "status": "ok",
        "assigned": assigned,
        "counters": {k: v for k, v in counters.items()},
    }


# --- Session 11: Subscriptions CSV Export ---

@app.get("/api/admin/subscriptions-csv")
async def admin_subscriptions_csv(request: Request):
    """Export all subscribers (Stripe + Apple + Google) as CSV with lead attribution."""

    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT s.*,
                   l.first_name as lead_name,
                   l.utm_source as lead_utm_source,
                   l.utm_medium as lead_utm_medium,
                   l.utm_campaign as lead_utm_campaign,
                   l.created_at as lead_date
            FROM subscriptions s
            LEFT JOIN LATERAL (
                SELECT first_name, utm_source, utm_medium, utm_campaign, created_at
                FROM leads
                WHERE lower(leads.email) = lower(s.email) AND s.email != ''
                ORDER BY created_at DESC LIMIT 1
            ) l ON true
            ORDER BY s.created_at DESC
        """)

    # readable_id now comes from DB (assigned persistently)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "readable_id", "email", "name", "source", "status",
        "plan", "signup_date",
        "trial_start", "trial_end", "converted_at", "cancel_date",
        "months_active", "est_lifetime_revenue",
        "utm_source", "utm_medium", "utm_campaign",
        "email_known",
        "renewal_count",
        "last_renewed_at"
    ])

    for r in rows:
        source = r.get("source", "stripe") or "stripe"
        sub_id = r.get("stripe_subscription_id", "") or ""

        readable_id = r.get("readable_id", "") or sub_id or f"{source.upper()}-{r['id']}"

        email = r.get("email", "") or ""
        email_known = "TRUE" if email and "@" in email else "FALSE"

        plan_amount = r.get("plan_amount", 0) or 0
        plan_interval = r.get("plan_interval", "") or ""
        created = r.get("created_at")
        canceled = r.get("canceled_at")

        # Plan display
        if plan_interval == "year":
            plan_display = f"${plan_amount / 100:.2f}/yr"
        elif plan_interval == "month":
            plan_display = f"${plan_amount / 100:.2f}/mo"
        else:
            plan_display = f"${plan_amount / 100:.2f}" if plan_amount else "$0.00"

        # Months active
        months_active = 0
        est_revenue = 0
        if created and plan_amount:
            end_date = canceled if canceled else datetime.now(timezone.utc)
            months_active = round(max(1, (end_date - created).days / 30), 1)
            monthly_equiv = plan_amount / 12 if plan_interval == "year" else plan_amount
            est_revenue = round(monthly_equiv * months_active)

        writer.writerow([
            readable_id,
            email,
            r.get("lead_name", "") or "",
            source,
            r.get("status", ""),
            plan_display,
            str(created.date()) if created else "",
            str(r.get("trial_start", "").date() if r.get("trial_start") else ""),
            str(r.get("trial_end", "").date() if r.get("trial_end") else ""),
            str(r.get("converted_at", "").date() if r.get("converted_at") else ""),
            str(canceled.date()) if canceled else "",
            months_active,
            f"${est_revenue / 100:.2f}",
            r.get("lead_utm_source", "") or "",
            r.get("lead_utm_medium", "") or "",
            r.get("lead_utm_campaign", "") or "",
            email_known,
            r.get("renewal_count", 0) or 0,
            str(r.get("last_renewed_at", "").date() if r.get("last_renewed_at") else ""),
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=mm-subscriptions.csv"}
    )



# --- Session 11: Multi-Tab XLSX Export (Meg format) ---

@app.get("/api/admin/subscriptions-xlsx")
async def admin_subscriptions_xlsx(request: Request):
    """Export subscribers as multi-tab XLSX matching Meg spreadsheet layout."""

    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")

    async with db_pool.acquire() as conn:
        all_subs = await conn.fetch("""
            SELECT s.*,
                   l.first_name as lead_name,
                   l.extra as lead_last_name,
                   l.utm_source as lead_utm_source,
                   l.utm_medium as lead_utm_medium,
                   l.utm_campaign as lead_utm_campaign
            FROM subscriptions s
            LEFT JOIN LATERAL (
                SELECT first_name, extra, utm_source, utm_medium, utm_campaign
                FROM leads
                WHERE lower(leads.email) = lower(s.email) AND s.email != ''
                ORDER BY created_at DESC LIMIT 1
            ) l ON true
            ORDER BY s.created_at
        """)

        offered_leads = await conn.fetch("""
            SELECT l.first_name, l.extra as last_name, l.email,
                   l.experience_level, l.goals, l.recommended_plan,
                   l.utm_source, l.utm_medium, l.utm_campaign, l.created_at
            FROM leads l
            WHERE l.email != ''
            AND NOT EXISTS (
                SELECT 1 FROM subscriptions s
                WHERE lower(s.email) = lower(l.email)
                AND s.status IN ('active', 'trialing')
            )
            AND lower(l.email) NOT IN (
                SELECT DISTINCT lower(email) FROM subscriptions
                WHERE email != '' AND status = 'canceled'
            )
            ORDER BY l.created_at DESC
        """)

    apple_google_active = []
    stripe_active = []
    trialing = []
    cancelled = []

    for r in all_subs:
        status = r.get("status", "")
        source = r.get("source", "stripe") or "stripe"
        if status == "trialing":
            trialing.append(r)
        elif status == "canceled":
            cancelled.append(r)
        elif status == "active":
            if source in ("apple", "google"):
                apple_google_active.append(r)
            else:
                stripe_active.append(r)

    wb = Workbook()

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill_navy = PatternFill("solid", fgColor="182241")
    hdr_fill_green = PatternFill("solid", fgColor="2d6a2d")
    hdr_fill_blue = PatternFill("solid", fgColor="3949ab")
    hdr_fill_orange = PatternFill("solid", fgColor="b35a00")
    hdr_fill_red = PatternFill("solid", fgColor="c0392b")
    data_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center")
    thin_border = Border(bottom=Side(style="thin", color="E0E0E0"))

    SUB_HEADERS = [
        "Readable ID", "First Name", "Last Name", "Email", "Sign-up Date",
        "Source", "Status", "Plan", "Trial Start", "Trial End",
        "Converted", "Renewals", "Last Renewed", "Cancel Date",
        "Months Active", "Est Revenue", "UTM Source", "UTM Medium", "UTM Campaign"
    ]

    def sub_row(r):
        source = r.get("source", "stripe") or "stripe"
        rid = r.get("readable_id", "") or r.get("stripe_subscription_id", "") or ""
        plan_amount = r.get("plan_amount", 0) or 0
        plan_interval = r.get("plan_interval", "") or ""
        created = r.get("created_at")
        canceled = r.get("canceled_at")
        if plan_interval == "year":
            plan_disp = f"${plan_amount/100:.2f}/yr"
        elif plan_interval == "month":
            plan_disp = f"${plan_amount/100:.2f}/mo"
        else:
            plan_disp = f"${plan_amount/100:.2f}" if plan_amount else ""
        months = 0
        revenue = 0
        if created and plan_amount:
            end = canceled if canceled else datetime.now(timezone.utc)
            months = round(max(1, (end - created).days / 30), 1)
            meq = plan_amount / 12 if plan_interval == "year" else plan_amount
            revenue = round(meq * months)
        return [
            rid,
            r.get("lead_name", "") or "",
            r.get("lead_last_name", "") or "",
            r.get("email", "") or "",
            created.strftime("%Y-%m-%d") if created else "",
            source.upper(),
            r.get("status", ""),
            plan_disp,
            r["trial_start"].strftime("%Y-%m-%d") if r.get("trial_start") else "",
            r["trial_end"].strftime("%Y-%m-%d") if r.get("trial_end") else "",
            r["converted_at"].strftime("%Y-%m-%d") if r.get("converted_at") else "",
            r.get("renewal_count", 0) or 0,
            r["last_renewed_at"].strftime("%Y-%m-%d") if r.get("last_renewed_at") else "",
            canceled.strftime("%Y-%m-%d") if canceled else "",
            months,
            f"${revenue/100:.2f}" if revenue else "",
            r.get("lead_utm_source", "") or "",
            r.get("lead_utm_medium", "") or "",
            r.get("lead_utm_campaign", "") or "",
        ]

    def write_sub_sheet(ws, rows, hdr_fill):
        for ci, h in enumerate(SUB_HEADERS, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center
        for ri, r in enumerate(rows, 2):
            vals = sub_row(r)
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.font = data_font
                c.border = thin_border
        for ci in range(1, len(SUB_HEADERS) + 1):
            max_len = len(SUB_HEADERS[ci-1])
            for ri in range(2, min(len(rows)+2, 100)):
                val = ws.cell(row=ri, column=ci).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 30)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions if len(rows) > 0 else "A1:S1"

    ws1 = wb.active
    ws1.title = "Apple & Google Members"
    write_sub_sheet(ws1, apple_google_active, hdr_fill_navy)

    ws2 = wb.create_sheet("Stripe Members")
    write_sub_sheet(ws2, stripe_active, hdr_fill_green)

    ws3 = wb.create_sheet("Free Month Trial")
    write_sub_sheet(ws3, trialing, hdr_fill_blue)

    ws4 = wb.create_sheet("Downloaded & Offered Trial")
    lead_headers = [
        "First Name", "Last Name", "Email", "Date",
        "Experience", "Goals", "Recommended Plan",
        "UTM Source", "UTM Medium", "UTM Campaign"
    ]
    for ci, h in enumerate(lead_headers, 1):
        c = ws4.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill_orange
        c.alignment = center
    for ri, r in enumerate(offered_leads, 2):
        vals = [
            r.get("first_name", "") or "",
            r.get("last_name", "") or "",
            r.get("email", "") or "",
            r["created_at"].strftime("%Y-%m-%d") if r.get("created_at") else "",
            r.get("experience_level", "") or "",
            r.get("goals", "") or "",
            r.get("recommended_plan", "") or "",
            r.get("utm_source", "") or "",
            r.get("utm_medium", "") or "",
            r.get("utm_campaign", "") or "",
        ]
        for ci, v in enumerate(vals, 1):
            c = ws4.cell(row=ri, column=ci, value=v)
            c.font = data_font
            c.border = thin_border
    for ci in range(1, len(lead_headers) + 1):
        max_len = len(lead_headers[ci-1])
        for ri in range(2, min(len(offered_leads)+2, 100)):
            val = ws4.cell(row=ri, column=ci).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws4.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 30)
    ws4.freeze_panes = "A2"
    if len(offered_leads) > 0:
        ws4.auto_filter.ref = ws4.dimensions

    ws5 = wb.create_sheet("Cancelled Subscription")
    write_sub_sheet(ws5, cancelled, hdr_fill_red)
    for ri in range(2, len(cancelled) + 2):
        for ci in range(1, len(SUB_HEADERS) + 1):
            ws5.cell(row=ri, column=ci).fill = PatternFill("solid", fgColor="FFE0E0")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mm-subscribers.xlsx"}
    )


# --- Session 17: Meg Format XLSX Export ---

@app.get("/api/admin/subscriptions-meg-xlsx")
async def admin_subscriptions_meg_xlsx(request: Request):
    """Export subscribers in Meg's exact spreadsheet format.
    Tab 1: Apple & Google Members (no headers) - first, last, email, date, source
    Tab 2: Stripe Members (no headers) - first, last, email, date
    Tab 3: Free Month Trial (headers) - First Name, Last Name, Email, Sign Up Date, Payment Type
    Tab 4: Downloaded & Offered Trial (headers) - First Name, Last Name, Email, Sign Up Date
    Tab 5: Cancelled Subscription (no headers) - first, last, email"""

    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")

    async with db_pool.acquire() as conn:
        all_subs = await conn.fetch("""
            SELECT s.*, s.first_name as sub_first, s.last_name as sub_last,
                   l.first_name as lead_name, l.extra as lead_last_name
            FROM subscriptions s
            LEFT JOIN LATERAL (
                SELECT first_name, extra FROM leads
                WHERE lower(leads.email) = lower(s.email) AND s.email != ''
                ORDER BY created_at DESC LIMIT 1
            ) l ON true
            ORDER BY s.created_at
        """)
        offered_leads = await conn.fetch("""
            SELECT l.first_name, l.extra as last_name, l.email, l.created_at
            FROM leads l
            WHERE l.email != ''
            AND NOT EXISTS (
                SELECT 1 FROM subscriptions s
                WHERE lower(s.email) = lower(l.email)
                AND s.status IN ('active', 'trialing')
            )
            AND lower(l.email) NOT IN (
                SELECT DISTINCT lower(email) FROM subscriptions
                WHERE email != '' AND status = 'canceled'
            )
            ORDER BY l.created_at DESC
        """)

    apple_google_active = []
    stripe_active = []
    trialing = []
    cancelled = []
    for r in all_subs:
        status = r.get("status", "")
        source = r.get("source", "stripe") or "stripe"
        if status == "trialing":
            trialing.append(r)
        elif status == "canceled":
            cancelled.append(r)
        elif status == "active":
            if source in ("apple", "google"):
                apple_google_active.append(r)
            else:
                stripe_active.append(r)

    wb = Workbook()
    hdr_font = Font(name="Arial", bold=True, size=11)
    data_font = Font(name="Arial", size=10)

    def get_name(r):
        first = r.get("sub_first", "") or r.get("lead_name", "") or ""
        last = r.get("sub_last", "") or r.get("lead_last_name", "") or ""
        return first, last

    def auto_width(ws, num_cols, num_rows):
        for ci in range(1, num_cols + 1):
            max_len = 10
            for ri in range(1, min(num_rows + 2, 100)):
                val = ws.cell(row=ri, column=ci).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 30)

    # Tab 1: Apple & Google Members (NO headers, matching Meg format)
    ws1 = wb.active
    ws1.title = "Apple & Google Members"
    for ri, r in enumerate(apple_google_active, 1):
        first, last = get_name(r)
        created = r.get("created_at")
        source = (r.get("source", "") or "").upper()
        ws1.cell(row=ri, column=1, value=first).font = data_font
        ws1.cell(row=ri, column=2, value=last).font = data_font
        ws1.cell(row=ri, column=3, value=r.get("email", "") or "").font = data_font
        ws1.cell(row=ri, column=4, value=created.strftime("%Y-%m-%d") if created else "").font = data_font
        ws1.cell(row=ri, column=5, value=source).font = data_font
    auto_width(ws1, 5, len(apple_google_active))

    # Tab 2: Stripe Members (NO headers)
    ws2 = wb.create_sheet("Stripe Members")
    for ri, r in enumerate(stripe_active, 1):
        first, last = get_name(r)
        created = r.get("created_at")
        ws2.cell(row=ri, column=1, value=first).font = data_font
        ws2.cell(row=ri, column=2, value=last).font = data_font
        ws2.cell(row=ri, column=3, value=r.get("email", "") or "").font = data_font
        ws2.cell(row=ri, column=4, value=created.strftime("%Y-%m-%d") if created else "").font = data_font
    auto_width(ws2, 4, len(stripe_active))

    # Tab 3: Free Month Trial (WITH headers)
    ws3 = wb.create_sheet("Free month trial")
    for ci, h in enumerate(["First Name", "Last Name", "Email", "Sign Up Date", "Payment Type"], 1):
        ws3.cell(row=1, column=ci, value=h).font = hdr_font
    for ri, r in enumerate(trialing, 2):
        first, last = get_name(r)
        created = r.get("created_at")
        source = (r.get("source", "stripe") or "stripe").capitalize()
        ws3.cell(row=ri, column=1, value=first).font = data_font
        ws3.cell(row=ri, column=2, value=last).font = data_font
        ws3.cell(row=ri, column=3, value=r.get("email", "") or "").font = data_font
        ws3.cell(row=ri, column=4, value=created.strftime("%Y-%m-%d") if created else "").font = data_font
        ws3.cell(row=ri, column=5, value=source).font = data_font
    auto_width(ws3, 5, len(trialing) + 1)
    ws3.freeze_panes = "A2"

    # Tab 4: Downloaded & Offered Trial (WITH headers)
    ws4 = wb.create_sheet("downloaded, offered free trial")
    for ci, h in enumerate(["First Name", "Last Name", "Email", "Sign Up Date"], 1):
        ws4.cell(row=1, column=ci, value=h).font = hdr_font
    for ri, r in enumerate(offered_leads, 2):
        ws4.cell(row=ri, column=1, value=r.get("first_name", "") or "").font = data_font
        ws4.cell(row=ri, column=2, value=r.get("last_name", "") or "").font = data_font
        ws4.cell(row=ri, column=3, value=r.get("email", "") or "").font = data_font
        created = r.get("created_at")
        ws4.cell(row=ri, column=4, value=created.strftime("%Y-%m-%d") if created else "").font = data_font
    auto_width(ws4, 4, len(offered_leads) + 1)
    ws4.freeze_panes = "A2"

    # Tab 5: Cancelled Subscription (NO headers)
    # S18: Filter out rows with no email (cosmetic fix)
    cancelled_with_email = [r for r in cancelled if r.get("email")]
    ws5 = wb.create_sheet("cancelled subscription")
    for ri, r in enumerate(cancelled_with_email, 1):
        first, last = get_name(r)
        ws5.cell(row=ri, column=1, value=first).font = data_font
        ws5.cell(row=ri, column=2, value=last).font = data_font
        ws5.cell(row=ri, column=3, value=r.get("email", "") or "").font = data_font
    auto_width(ws5, 3, len(cancelled_with_email))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mm-meg-format.xlsx"}
    )


# --- Session 11: User Journey CSV (lead -> subscriber timeline) ---

@app.get("/api/admin/user-journey-csv")
async def admin_user_journey_csv(request: Request):
    """Export lead-to-subscriber journey: joins leads + subscriptions by email."""

    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT
                l.email,
                l.first_name,
                l.created_at as lead_date,
                l.utm_source, l.utm_medium, l.utm_campaign,
                l.experience_level, l.goals, l.recommended_plan,
                s.source as sub_source,
                s.status as sub_status,
                s.plan_interval, s.plan_amount,
                s.trial_start, s.trial_end, s.converted_at,
                s.created_at as sub_date,
                s.canceled_at,
                s.current_period_start, s.current_period_end
            FROM leads l
            LEFT JOIN LATERAL (
                SELECT * FROM subscriptions
                WHERE lower(subscriptions.email) = lower(l.email) AND l.email != ''
                ORDER BY created_at DESC LIMIT 1
            ) s ON true
            WHERE l.email != ''
            ORDER BY l.created_at DESC
        """)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "email", "first_name",
        "lead_date", "utm_source", "utm_medium", "utm_campaign",
        "experience_level", "goals", "recommended_plan",
        "subscribed", "sub_source", "sub_status", "sub_date",
        "plan_interval", "plan_amount",
        "trial_start", "trial_end", "converted_at", "canceled_at",
        "days_lead_to_sub", "est_lifetime_revenue"
    ])

    for r in rows:
        subscribed = "yes" if r.get("sub_source") else "no"

        days_to_sub = ""
        if r.get("lead_date") and r.get("sub_date"):
            delta = r["sub_date"] - r["lead_date"]
            days_to_sub = max(0, delta.days)

        plan_amount = r.get("plan_amount", 0) or 0
        plan_interval = r.get("plan_interval", "") or ""
        sub_date = r.get("sub_date")
        canceled = r.get("canceled_at")
        est_revenue = 0
        if sub_date and plan_amount:
            end_date = canceled if canceled else datetime.now(timezone.utc)
            months_active = max(1, (end_date - sub_date).days / 30)
            monthly_equiv = plan_amount / 12 if plan_interval == "year" else plan_amount
            est_revenue = round(monthly_equiv * months_active)

        writer.writerow([
            r.get("email", ""),
            r.get("first_name", ""),
            str(r.get("lead_date", "") or ""),
            r.get("utm_source", "") or "",
            r.get("utm_medium", "") or "",
            r.get("utm_campaign", "") or "",
            r.get("experience_level", "") or "",
            r.get("goals", "") or "",
            r.get("recommended_plan", "") or "",
            subscribed,
            r.get("sub_source", "") or "",
            r.get("sub_status", "") or "",
            str(r.get("sub_date", "") or ""),
            plan_interval,
            f"${plan_amount / 100:.2f}" if plan_amount else "$0.00",
            str(r.get("trial_start", "") or ""),
            str(r.get("trial_end", "") or ""),
            str(r.get("converted_at", "") or ""),
            str(canceled or ""),
            str(days_to_sub),
            f"${est_revenue / 100:.2f}",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=mm-user-journey.csv"}
    )


# --- Session 11: Ad Spend CRUD ---

@app.post("/api/admin/ad-spend")
async def save_ad_spend(req: AdSpendRequest, request: Request):
    """Save or update ad spend for a month/channel combination."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")

    if not req.month or not req.channel:
        raise HTTPException(status_code=400, detail="month and channel required")

    amount_cents = round(req.amount_dollars * 100)
    async with db_pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO ad_spend (month, channel, amount_cents, updated_at)
            VALUES ($1, $2, $3, NOW())
            ON CONFLICT (month, channel) DO UPDATE SET
                amount_cents = EXCLUDED.amount_cents,
                updated_at = NOW()
        """, req.month, req.channel.lower().strip(), amount_cents)

    return {"status": "ok", "month": req.month, "channel": req.channel, "amount_cents": amount_cents}


@app.get("/api/admin/ad-spend")
async def get_ad_spend(request: Request):
    """Retrieve all ad spend records."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM ad_spend ORDER BY month DESC, channel")

    return {"spend": [
        {"month": r["month"], "channel": r["channel"], "amount_cents": r["amount_cents"]}
        for r in rows
    ]}


@app.delete("/api/admin/ad-spend")
async def delete_ad_spend(request: Request):
    """Delete a specific ad spend entry."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")

    body = await request.json()
    month = body.get("month", "")
    channel = body.get("channel", "")
    if not month or not channel:
        raise HTTPException(status_code=400, detail="month and channel required")

    async with db_pool.acquire() as conn:
        await conn.execute("DELETE FROM ad_spend WHERE month = $1 AND channel = $2", month, channel)

    return {"status": "ok"}


# --- S23: UTM Links CRUD ---

@app.post("/api/admin/utm-links")
async def save_utm_link(request: Request):
    """Save a generated UTM link."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")
    body = await request.json()
    base_url = (body.get("base_url") or "").strip()
    utm_source = (body.get("utm_source") or "").strip()
    if not base_url or not utm_source:
        raise HTTPException(status_code=400, detail="base_url and utm_source required")
    label = (body.get("label") or "").strip()
    utm_medium = (body.get("utm_medium") or "").strip()
    utm_campaign = (body.get("utm_campaign") or "").strip()
    utm_term = (body.get("utm_term") or "").strip()
    utm_content = (body.get("utm_content") or "").strip()
    ym_source = (body.get("ym_source") or "").strip()
    full_url = (body.get("full_url") or "").strip()
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            """INSERT INTO utm_links (label, base_url, utm_source, utm_medium, utm_campaign, utm_term, utm_content, ym_source, full_url)
               VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9) RETURNING id, created_at""",
            label, base_url, utm_source, utm_medium, utm_campaign, utm_term, utm_content, ym_source, full_url
        )
    return {"status": "ok", "id": row["id"], "created_at": str(row["created_at"])}


@app.get("/api/admin/utm-links")
async def get_utm_links(request: Request):
    """List all saved UTM links."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM utm_links ORDER BY created_at DESC")
    return {"links": [dict(r) for r in rows]}


@app.delete("/api/admin/utm-links")
async def delete_utm_link(request: Request):
    """Delete a UTM link by ID."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")
    link_id = request.query_params.get("id")
    if not link_id:
        raise HTTPException(status_code=400, detail="id required")
    async with db_pool.acquire() as conn:
        await conn.execute("DELETE FROM utm_links WHERE id = $1", int(link_id))
    return {"status": "ok", "deleted": int(link_id)}


# --- Session 11: Attach Email + Search (consolidated) ---

@app.post("/api/admin/attach-email")
async def attach_email(request: Request):
    """Permanently attach an email to an Apple/Google subscription.
    Once attached, leads-table matching gives full name + UTM attribution."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")

    body = await request.json()
    sub_id = (body.get("subscription_id") or "").strip()
    email = (body.get("email") or "").strip().lower()

    if not sub_id or not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Valid subscription_id and email required")

    async with db_pool.acquire() as conn:
        # Find the subscription
        row = await conn.fetchrow(
            "SELECT id, source, email as current_email FROM subscriptions WHERE stripe_subscription_id = $1",
            sub_id
        )
        if not row:
            raise HTTPException(status_code=404, detail=f"Subscription '{sub_id}' not found")

        await conn.execute(
            "UPDATE subscriptions SET email = $1, updated_at = NOW() WHERE stripe_subscription_id = $2",
            email, sub_id
        )

        # Check if we can match to a lead now
        lead = await conn.fetchrow(
            "SELECT first_name, utm_source, utm_medium, utm_campaign FROM leads WHERE lower(email) = $1 ORDER BY created_at DESC LIMIT 1",
            email
        )

    result = {
        "status": "ok",
        "subscription_id": sub_id,
        "email": email,
        "source": row["source"],
        "previous_email": row["current_email"] or "(none)",
    }
    if lead:
        result["matched_lead"] = {
            "name": lead["first_name"] or "",
            "utm_source": lead["utm_source"] or "",
            "utm_medium": lead["utm_medium"] or "",
            "utm_campaign": lead["utm_campaign"] or "",
        }
    else:
        result["matched_lead"] = None

    return result


@app.get("/api/admin/search-subscriptions")
async def search_subscriptions(request: Request):
    """Search subscriptions by ID prefix or email for the attach-email UI."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")

    q = request.query_params.get("q", "").strip()
    if len(q) < 2:
        return {"results": []}

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT stripe_subscription_id, email, source, status, plan_amount, plan_interval
            FROM subscriptions
            WHERE stripe_subscription_id ILIKE $1 OR email ILIKE $1
            ORDER BY created_at DESC LIMIT 10
        """, f"%{q}%")

    return {"results": [
        {
            "id": r["stripe_subscription_id"],
            "email": r["email"] or "",
            "source": r["source"],
            "status": r["status"],
            "plan": f"${(r['plan_amount'] or 0)/100:.2f}/{r['plan_interval'] or '?'}",
        }
        for r in rows
    ]}


@app.get("/api/admin/stats")
async def admin_stats(request: Request):
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)

    if not db_pool:
        return {"error": "No database connected"}

    # Date range filtering (optional query params)
    date_from_str = request.query_params.get("from", "")
    date_to_str = request.query_params.get("to", "")
    date_from_dt = None
    date_to_dt = None
    try:
        if date_from_str:
            date_from_dt = datetime.strptime(date_from_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        if date_to_str:
            date_to_dt = datetime.strptime(date_to_str, "%Y-%m-%d").replace(tzinfo=timezone.utc) + timedelta(days=1)
    except Exception as e:
        print(f"Date parse error: {e}")

    async with db_pool.acquire() as conn:
        # Leads (with optional date filter)
        if date_from_dt or date_to_dt:
            lead_count_q = "SELECT COUNT(*) FROM leads WHERE 1=1"
            lead_list_q = "SELECT * FROM leads WHERE 1=1"
            idx = 1
            args = []
            if date_from_dt:
                lead_count_q += " AND created_at >= $" + str(idx)
                lead_list_q += " AND created_at >= $" + str(idx)
                args.append(date_from_dt)
                idx += 1
            if date_to_dt:
                lead_count_q += " AND created_at < $" + str(idx)
                lead_list_q += " AND created_at < $" + str(idx)
                args.append(date_to_dt)
                idx += 1
            lead_list_q += " ORDER BY created_at DESC LIMIT 50"
            total_leads = await conn.fetchval(lead_count_q, *args)
            recent_leads = await conn.fetch(lead_list_q, *args)
        else:
            total_leads = await conn.fetchval("SELECT COUNT(*) FROM leads")
            recent_leads = await conn.fetch(
                "SELECT * FROM leads ORDER BY created_at DESC LIMIT 50"
            )

        # Page views (with optional date filter)
        if date_from_dt or date_to_dt:
            pv_q = "SELECT COUNT(*) FROM page_views WHERE 1=1"
            pv_args = []
            pv_idx = 1
            if date_from_dt:
                pv_q += " AND created_at >= $" + str(pv_idx)
                pv_args.append(date_from_dt)
                pv_idx += 1
            if date_to_dt:
                pv_q += " AND created_at < $" + str(pv_idx)
                pv_args.append(date_to_dt)
                pv_idx += 1
            total_views = await conn.fetchval(pv_q, *pv_args)
            today_views = await conn.fetchval(
                "SELECT COUNT(*) FROM page_views WHERE created_at::date = CURRENT_DATE"
            )
        else:
            total_views = await conn.fetchval("SELECT COUNT(*) FROM page_views")
            today_views = await conn.fetchval(
                "SELECT COUNT(*) FROM page_views WHERE created_at::date = CURRENT_DATE"
            )
        views_by_page = await conn.fetch(
            """SELECT page, COUNT(*) as views FROM page_views
               WHERE created_at > NOW() - INTERVAL '7 days'
               GROUP BY page ORDER BY views DESC LIMIT 10"""
        )
        views_by_day = await conn.fetch(
            """SELECT created_at::date as day, COUNT(*) as views FROM page_views
               WHERE created_at > NOW() - INTERVAL '7 days'
               GROUP BY day ORDER BY day"""
        )
        # S22: Page views by UTM source (7 days)
        pv_by_utm_source = await conn.fetch(
            """SELECT COALESCE(NULLIF(utm_source, ''), 'direct') as channel, COUNT(*) as views
               FROM page_views
               WHERE created_at > NOW() - INTERVAL '7 days'
               GROUP BY channel ORDER BY views DESC"""
        )

        # Chat sessions
        total_chats = await conn.fetchval("SELECT COUNT(*) FROM chat_sessions")
        chats_by_type = await conn.fetch(
            "SELECT session_type, COUNT(*) as count FROM chat_sessions GROUP BY session_type"
        )

        # Phase 5: UTM attribution
        leads_by_source = await conn.fetch(
            """SELECT COALESCE(NULLIF(utm_source,''), 'direct') as channel, COUNT(*) as count
               FROM leads GROUP BY channel ORDER BY count DESC"""
        )
        leads_by_medium = await conn.fetch(
            """SELECT COALESCE(NULLIF(utm_medium,''), 'none') as medium, COUNT(*) as count
               FROM leads GROUP BY medium ORDER BY count DESC"""
        )
        leads_by_campaign = await conn.fetch(
            """SELECT COALESCE(NULLIF(utm_campaign,''), 'none') as campaign, COUNT(*) as count
               FROM leads GROUP BY campaign ORDER BY count DESC LIMIT 10"""
        )

        # Subscription stats (Phase 3)
        active_subs = await conn.fetchval(
            "SELECT COUNT(*) FROM subscriptions WHERE status IN ('active', 'trialing')"
        )
        trialing_subs = await conn.fetchval(
            "SELECT COUNT(*) FROM subscriptions WHERE status = 'trialing'"
        )
        canceled_subs = await conn.fetchval(
            "SELECT COUNT(*) FROM subscriptions WHERE status = 'canceled'"
        )
        total_subs = await conn.fetchval("SELECT COUNT(*) FROM subscriptions")

        # MRR calculation: sum of active monthly amounts
        # Monthly subs count as-is, annual subs divided by 12
        mrr_monthly = await conn.fetchval(
            """SELECT COALESCE(SUM(plan_amount), 0) FROM subscriptions
               WHERE status = 'active' AND plan_interval = 'month'"""
        )
        mrr_annual = await conn.fetchval(
            """SELECT COALESCE(SUM(plan_amount / 12), 0) FROM subscriptions
               WHERE status = 'active' AND plan_interval = 'year'"""
        )
        mrr_cents = (mrr_monthly or 0) + (mrr_annual or 0)

        # Session 11: Fixed trial -> paid conversion
        # "Ever converted" = has converted_at OR (had trial and current_period moved past trial_end)
        total_ever_trialed = await conn.fetchval(
            "SELECT COUNT(*) FROM subscriptions WHERE trial_start IS NOT NULL"
        )
        converted_from_trial = await conn.fetchval(
            """SELECT COUNT(*) FROM subscriptions
               WHERE trial_start IS NOT NULL
               AND (converted_at IS NOT NULL
                    OR (trial_end IS NOT NULL AND current_period_start IS NOT NULL AND current_period_start > trial_end))"""
        )

        # Rolling cohort data for conversion windows
        cohort_subs = await conn.fetch(
            """SELECT created_at, converted_at, canceled_at, status, trial_start, trial_end, current_period_start
               FROM subscriptions WHERE trial_start IS NOT NULL"""
        )

        # Recent conversion activity
        conversions_7d = await conn.fetchval(
            "SELECT COUNT(*) FROM subscriptions WHERE converted_at > NOW() - INTERVAL '7 days'"
        ) or 0
        conversions_30d = await conn.fetchval(
            "SELECT COUNT(*) FROM subscriptions WHERE converted_at > NOW() - INTERVAL '30 days'"
        ) or 0

        # Churn: canceled in last 30 days vs active at start of period
        churned_30d = await conn.fetchval(
            """SELECT COUNT(*) FROM subscriptions
               WHERE status = 'canceled' AND canceled_at > NOW() - INTERVAL '30 days'"""
        )
        # S22: Split churn into paid vs trial
        churned_30d_paid = await conn.fetchval(
            """SELECT COUNT(*) FROM subscriptions
               WHERE status = 'canceled' AND canceled_at > NOW() - INTERVAL '30 days'
               AND converted_at IS NOT NULL"""
        )
        churned_30d_trial = await conn.fetchval(
            """SELECT COUNT(*) FROM subscriptions
               WHERE status = 'canceled' AND canceled_at > NOW() - INTERVAL '30 days'
               AND converted_at IS NULL"""
        )

        # Recent subscription events
        recent_events = await conn.fetch(
            """SELECT event_type, stripe_customer_id, source, created_at
               FROM subscription_events ORDER BY created_at DESC LIMIT 20"""
        )

        # Subscriptions by status
        subs_by_status = await conn.fetch(
            "SELECT status, COUNT(*) as count FROM subscriptions GROUP BY status ORDER BY count DESC"
        )

        # Phase 4: Subscriptions by source
        subs_by_source = await conn.fetch(
            "SELECT source, COUNT(*) as count, SUM(CASE WHEN status IN ('active','trialing') THEN 1 ELSE 0 END) as active_count FROM subscriptions GROUP BY source ORDER BY count DESC"
        )

        # MRR by source
        mrr_by_source = await conn.fetch(
            """SELECT source,
                COALESCE(SUM(CASE WHEN plan_interval='month' THEN plan_amount ELSE 0 END), 0) as mrr_monthly,
                COALESCE(SUM(CASE WHEN plan_interval='year' THEN plan_amount/12 ELSE 0 END), 0) as mrr_annual
               FROM subscriptions WHERE status = 'active'
               GROUP BY source"""
        )

        # S22: Subscription UTM attribution (from ymove meta parameters)
        subs_by_utm_source = await conn.fetch(
            """SELECT COALESCE(NULLIF(utm_source, ''), 'none') as channel, COUNT(*) as count
               FROM subscriptions WHERE utm_source IS NOT NULL AND utm_source != ''
               GROUP BY channel ORDER BY count DESC"""
        )
        subs_by_utm_medium = await conn.fetch(
            """SELECT COALESCE(NULLIF(utm_medium, ''), 'none') as medium, COUNT(*) as count
               FROM subscriptions WHERE utm_source IS NOT NULL AND utm_source != ''
               GROUP BY medium ORDER BY count DESC"""
        )
        subs_by_utm_campaign = await conn.fetch(
            """SELECT COALESCE(NULLIF(utm_campaign, ''), 'none') as campaign, COUNT(*) as count
               FROM subscriptions WHERE utm_source IS NOT NULL AND utm_source != ''
               GROUP BY campaign ORDER BY count DESC LIMIT 10"""
        )
        subs_with_utm = await conn.fetchval(
            "SELECT COUNT(*) FROM subscriptions WHERE utm_source IS NOT NULL AND utm_source != ''"
        )

        # Phase 5b: MRR trend - weekly (last 12 weeks) + monthly (last 12 months)
        # Note: reactivated subs count from their created_at because they were paying
        # on Apple/Google the whole time - our DB just had bad data before S19 reconciliation.
        mrr_trend_weekly = await conn.fetch(
            """SELECT
                to_char(w, 'YYYY-MM-DD') as period,
                COALESCE(SUM(CASE WHEN s.plan_interval='month' THEN s.plan_amount ELSE 0 END), 0) as monthly_total,
                COALESCE(SUM(CASE WHEN s.plan_interval='year' THEN s.plan_amount/12 ELSE 0 END), 0) as annual_total
               FROM generate_series(
                   date_trunc('week', NOW() - INTERVAL '12 weeks'),
                   date_trunc('week', NOW()),
                   '1 week'::interval
               ) AS w
               LEFT JOIN subscriptions s ON s.created_at <= w
                   AND (s.canceled_at IS NULL OR s.canceled_at > w)
                   AND s.status != 'incomplete_expired'
               GROUP BY w ORDER BY w"""
        )
        mrr_trend_monthly = await conn.fetch(
            """SELECT
                to_char(w, 'YYYY-MM') as period,
                COALESCE(SUM(CASE WHEN s.plan_interval='month' THEN s.plan_amount ELSE 0 END), 0) as monthly_total,
                COALESCE(SUM(CASE WHEN s.plan_interval='year' THEN s.plan_amount/12 ELSE 0 END), 0) as annual_total
               FROM generate_series(
                   date_trunc('month', NOW() - INTERVAL '12 months'),
                   date_trunc('month', NOW()),
                   '1 month'::interval
               ) AS w
               LEFT JOIN subscriptions s ON s.created_at <= w
                   AND (s.canceled_at IS NULL OR s.canceled_at > w)
                   AND s.status != 'incomplete_expired'
               GROUP BY w ORDER BY w"""
        )

        # Avg subscriber lifetime (all who ever converted â includes churned for honest LTV)
        avg_sub_age = await conn.fetchval(
            """SELECT COALESCE(AVG(EXTRACT(EPOCH FROM (COALESCE(canceled_at, NOW()) - created_at)) / 86400), 0)
               FROM subscriptions
               WHERE converted_at IS NOT NULL
               OR (trial_end IS NOT NULL AND current_period_start IS NOT NULL AND current_period_start > trial_end)"""
        )

        # Avg monthly revenue per converted sub (includes churned for honest LTV)
        avg_monthly_per_sub = await conn.fetchval(
            """SELECT COALESCE(AVG(
                CASE WHEN plan_interval='month' THEN plan_amount
                     WHEN plan_interval='year' THEN plan_amount/12
                     ELSE 0 END
               ), 0) FROM subscriptions
               WHERE converted_at IS NOT NULL
               OR (trial_end IS NOT NULL AND current_period_start IS NOT NULL AND current_period_start > trial_end)"""
        )

        # Total subscribers ever (for lifetime calc)
        total_subs_ever = await conn.fetchval("SELECT COUNT(*) FROM subscriptions")

        # ARR
        arr_cents = mrr_cents * 12

        # Session 11: Ad spend data
        ad_spend_rows = await conn.fetch("SELECT * FROM ad_spend ORDER BY month DESC, channel")

        # Session 15: Conversion funnel by UTM source
        funnel_rows = await conn.fetch("""
            SELECT
              COALESCE(NULLIF(l.utm_source, ''), 'unknown') as channel,
              COUNT(*)::int as trials,
              SUM(CASE WHEN s.converted_at IS NOT NULL
                   OR (s.trial_end IS NOT NULL AND s.current_period_start IS NOT NULL
                       AND s.current_period_start > s.trial_end)
                   THEN 1 ELSE 0 END)::int as converted,
              SUM(CASE WHEN NOT (s.converted_at IS NOT NULL
                   OR (s.trial_end IS NOT NULL AND s.current_period_start IS NOT NULL
                       AND s.current_period_start > s.trial_end))
                   AND s.status = 'trialing' AND s.trial_end IS NOT NULL AND s.trial_end > NOW()
                   THEN 1 ELSE 0 END)::int as still_trialing
            FROM subscriptions s
            LEFT JOIN LATERAL (
              SELECT utm_source FROM leads
              WHERE lower(leads.email) = lower(s.email) AND s.email != ''
              ORDER BY created_at DESC LIMIT 1
            ) l ON true
            WHERE s.trial_start IS NOT NULL
            GROUP BY channel
            ORDER BY trials DESC
        """)

        # Ad spend totals by channel (all months summed)
        ad_spend_totals = await conn.fetch(
            "SELECT channel, SUM(amount_cents)::int as total_cents FROM ad_spend GROUP BY channel"
        )

    # Build response
    trial_conversion_rate = 0
    if total_ever_trialed and total_ever_trialed > 0:
        trial_conversion_rate = round((converted_from_trial / total_ever_trialed) * 100, 1)

    # Rolling cohort conversion windows (7d, 30d, 60d, 90d, all)
    now_utc = datetime.now(timezone.utc)
    cohort_windows = [
        {"label": "7d", "days": 7},
        {"label": "30d", "days": 30},
        {"label": "60d", "days": 60},
        {"label": "90d", "days": 90},
        {"label": "all", "days": None},
    ]
    conversion_cohorts = []
    for w in cohort_windows:
        cutoff = now_utc - timedelta(days=w["days"]) if w["days"] else None
        trials = converted = canceled = still_trialing = mature = 0
        for s in cohort_subs:
            created = s["created_at"]
            if not created:
                continue
            if cutoff and created < cutoff:
                continue
            trials += 1
            if (now_utc - created).days >= 30:
                mature += 1
            has_converted = (s["converted_at"] is not None) or (s["trial_end"] and s["current_period_start"] and s["current_period_start"] > s["trial_end"])
            if has_converted:
                converted += 1
            elif s["status"] == "canceled":
                canceled += 1
            elif s["status"] == "trialing" and s["trial_end"] and s["trial_end"] > now_utc:
                still_trialing += 1
            else:
                canceled += 1
        decided = converted + canceled
        conversion_cohorts.append({
            "window": w["label"],
            "trials": trials,
            "converted": converted,
            "canceled": canceled,
            "still_trialing": still_trialing,
            "conversion_rate": round((converted / decided) * 100, 1) if decided > 0 else None,
            "maturity_pct": round((mature / trials) * 100, 1) if trials > 0 else 0,
        })

    churn_rate = 0
    if active_subs and active_subs > 0:
        churn_rate = round((churned_30d / (active_subs + churned_30d)) * 100, 1)

    # Session 15: Build conversion funnel data
    spend_by_channel = {r["channel"]: r["total_cents"] for r in ad_spend_totals}
    conversion_funnel = []
    all_trials = all_converted = all_canceled = all_trialing = 0
    all_spend = sum(spend_by_channel.values())

    for fr in funnel_rows:
        ch = fr["channel"]
        trials = fr["trials"]
        converted = fr["converted"]
        st = fr["still_trialing"]
        canceled = trials - converted - st
        decided = converted + canceled
        rate = round((converted / decided) * 100, 1) if decided > 0 else None
        spend = spend_by_channel.get(ch, 0)
        cpa = round(spend / converted) if converted > 0 and spend > 0 else None

        all_trials += trials
        all_converted += converted
        all_canceled += canceled
        all_trialing += st

        conversion_funnel.append({
            "channel": ch,
            "trials": trials,
            "converted": converted,
            "canceled": canceled,
            "still_trialing": st,
            "conv_rate": rate,
            "ad_spend_cents": spend,
            "cpa_cents": cpa,
        })

    all_decided = all_converted + all_canceled
    all_rate = round((all_converted / all_decided) * 100, 1) if all_decided > 0 else None
    all_cpa = round(all_spend / all_converted) if all_converted > 0 and all_spend > 0 else None
    conversion_funnel.insert(0, {
        "channel": "__all__",
        "trials": all_trials,
        "converted": all_converted,
        "canceled": all_canceled,
        "still_trialing": all_trialing,
        "conv_rate": all_rate,
        "ad_spend_cents": all_spend,
        "cpa_cents": all_cpa,
    })

    # Last sync status for dashboard indicator
    last_sync = None
    async with db_pool.acquire() as conn:
        sync_row = await conn.fetchrow(
            "SELECT status, started_at, completed_at, our_active_count, ymove_active_count, results, error FROM ymove_sync_runs ORDER BY started_at DESC LIMIT 1"
        )
        if sync_row:
            last_sync = {
                "status": sync_row["status"],
                "started_at": str(sync_row["started_at"]) if sync_row["started_at"] else None,
                "completed_at": str(sync_row["completed_at"]) if sync_row["completed_at"] else None,
                "our_active_count": sync_row["our_active_count"],
                "ymove_active_count": sync_row["ymove_active_count"],
                "results": sync_row["results"],
                "error": sync_row["error"],
            }

    return {
        "leads": {
            "total": total_leads,
            "recent": [dict(r) for r in recent_leads],
            "by_source": [{"channel": r["channel"], "count": r["count"]} for r in leads_by_source],
            "by_medium": [{"medium": r["medium"], "count": r["count"]} for r in leads_by_medium],
            "by_campaign": [{"campaign": r["campaign"], "count": r["count"]} for r in leads_by_campaign],
        },
        "page_views": {
            "total": total_views,
            "today": today_views,
            "by_page": [{"page": r["page"], "views": r["views"]} for r in views_by_page],
            "by_day": [{"day": str(r["day"]), "views": r["views"]} for r in views_by_day],
            "by_utm_source": [{"channel": r["channel"], "views": r["views"]} for r in pv_by_utm_source],
        },
        "chats": {
            "total": total_chats,
            "by_type": [{"session_type": r["session_type"], "count": r["count"]} for r in chats_by_type],
        },
        "subscriptions": {
            "active": active_subs or 0,
            "trialing": trialing_subs or 0,
            "canceled": canceled_subs or 0,
            "total": total_subs or 0,
            "mrr_cents": mrr_cents,
            "mrr_display": f"${mrr_cents / 100:,.2f}",
            "trial_conversion_rate": trial_conversion_rate,
            "conversions_7d": conversions_7d,
            "conversions_30d": conversions_30d,
            "conversion_cohorts": conversion_cohorts,
            "churn_rate_30d": churn_rate,
            "churned_30d": churned_30d or 0,
            "churned_30d_paid": churned_30d_paid or 0,
            "churned_30d_trial": churned_30d_trial or 0,
            "by_status": [{"status": r["status"], "count": r["count"]} for r in subs_by_status],
            "by_source": [{"source": r["source"], "total": r["count"], "active": r["active_count"]} for r in subs_by_source],
            "mrr_by_source": [
                {"source": r["source"], "mrr_cents": (r["mrr_monthly"] or 0) + (r["mrr_annual"] or 0)}
                for r in mrr_by_source
            ],
            "mrr_trend": [
                {"period": r["period"], "mrr_cents": (r["monthly_total"] or 0) + (r["annual_total"] or 0)}
                for r in mrr_trend_weekly
            ],
            "mrr_trend_monthly": [
                {"period": r["period"], "mrr_cents": (r["monthly_total"] or 0) + (r["annual_total"] or 0)}
                for r in mrr_trend_monthly
            ],
            "arr_cents": arr_cents,
            "arr_display": f"${arr_cents / 100:,.2f}",
            "avg_sub_age_days": round(float(avg_sub_age or 0), 1),
            "avg_monthly_per_sub_cents": round(float(avg_monthly_per_sub or 0)),
            "est_ltv_cents": round(float(avg_monthly_per_sub or 0) * max(float(avg_sub_age or 30) / 30, 1)),
            "est_ltv_display": f"${round(float(avg_monthly_per_sub or 0) * max(float(avg_sub_age or 30) / 30, 1)) / 100:,.2f}",
            "recent_events": [
                {
                    "event_type": r["event_type"],
                    "customer": r["stripe_customer_id"],
                    "source": r.get("source", "stripe"),
                    "created_at": str(r["created_at"]),
                }
                for r in recent_events
            ],
        },
        "conversion_funnel": conversion_funnel,
        "ad_spend": [
            {"month": r["month"], "channel": r["channel"], "amount_cents": r["amount_cents"]}
            for r in ad_spend_rows
        ],
        "subscription_utm": {
            "total_with_utm": subs_with_utm or 0,
            "by_source": [{"channel": r["channel"], "count": r["count"]} for r in subs_by_utm_source],
            "by_medium": [{"medium": r["medium"], "count": r["count"]} for r in subs_by_utm_medium],
            "by_campaign": [{"campaign": r["campaign"], "count": r["count"]} for r in subs_by_utm_campaign],
        },
        "last_sync": last_sync,
    }


@app.get("/api/admin/leads-csv")
async def admin_leads_csv(request: Request):
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)

    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM leads ORDER BY created_at DESC")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "first_name", "email", "experience_level", "goals", "referral_source", "recommended_plan", "extra", "utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content", "ym_source", "created_at"])
    for r in rows:
        writer.writerow([r["id"], r["first_name"], r["email"], r["experience_level"], r["goals"], r["referral_source"], r["recommended_plan"], r["extra"], r.get("utm_source",""), r.get("utm_medium",""), r.get("utm_campaign",""), r.get("utm_term",""), r.get("utm_content",""), r.get("ym_source",""), str(r["created_at"])])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=mm-leads.csv"}
    )


@app.get("/api/health")
async def health():
    db_status = "connected" if db_pool else "not connected"
    stripe_status = "configured" if STRIPE_SECRET_KEY else "not configured"
    digest_status = "active" if (RESEND_API_KEY and DIGEST_RECIPIENTS and scheduler) else "disabled"
    shadow_sync_status = "active" if (YMOVE_API_KEY and scheduler) else "disabled"
    return {
        "status": "ok",
        "service": "Movement & Miles",
        "version": "23.2.0",
        "database": db_status,
        "stripe": stripe_status,
        "daily_digest": digest_status,
        "digest_recipients": DIGEST_RECIPIENTS if DIGEST_RECIPIENTS else "none",
        "daily_shadow_sync": shadow_sync_status,
        "shadow_sync_schedule": "8:00 AM ET daily" if shadow_sync_status == "active" else "disabled",
    }



# --- Session 16: Dedup Active Subscriptions ---

@app.post("/api/admin/dedup-active")
async def dedup_active(request: Request):
    """Find and resolve duplicate active subscriptions for the same email.
    Without confirm=true: preview. With confirm=true: cancel the inferior duplicate."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    body = {}
    try:
        body = await request.json()
    except Exception:
        pass
    confirm = body.get("confirm", False)

    async with db_pool.acquire() as conn:
        dups = await conn.fetch("""
            SELECT lower(email) as em,
                   array_agg(id ORDER BY updated_at DESC) as ids,
                   array_agg(source ORDER BY updated_at DESC) as sources,
                   array_agg(stripe_subscription_id ORDER BY updated_at DESC) as sub_ids,
                   array_agg(plan_amount ORDER BY updated_at DESC) as amounts
            FROM subscriptions
            WHERE status IN ('active', 'trialing') AND email != '' AND email IS NOT NULL
            GROUP BY lower(email) HAVING COUNT(*) > 1
        """)

        results = []
        canceled_count = 0
        for d in dups:
            # Pick keeper: prefer stripe (live webhooks), then most recently updated
            keep_idx = 0
            for i, src in enumerate(d["sources"]):
                if src == "stripe" and d["sources"][keep_idx] != "stripe":
                    keep_idx = i
                    break

            entries = []
            for i in range(len(d["ids"])):
                action = "KEEP" if i == keep_idx else "CANCEL"
                entries.append({"id": d["ids"][i], "sub_id": d["sub_ids"][i], "source": d["sources"][i], "amount": d["amounts"][i], "action": action})
                if confirm and action == "CANCEL":
                    await conn.execute(
                        "UPDATE subscriptions SET status = 'canceled', canceled_at = NOW(), updated_at = NOW() WHERE id = $1",
                        d["ids"][i]
                    )
                    canceled_count += 1
            results.append({"email": d["em"], "records": entries})

    if confirm:
        return {"status": "ok", "duplicates_found": len(dups), "canceled": canceled_count, "details": results}
    return {"status": "preview", "duplicates_found": len(dups), "plan": results}


# --- Session 16: Meg Apple/Google Import ---

@app.post("/api/admin/import-meg-apple-google")
async def import_meg_apple_google(request: Request, file: UploadFile = File(...)):
    """Import from Meg Apple & Google Members sheet (headerless: first,last,email,date,source).
    ?preview=true for dry run. All imports tagged with batch_id for revert."""

    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    preview_mode = request.query_params.get("preview", "false").lower() == "true"
    skip_reactivate = request.query_params.get("skip_reactivate", "false").lower() == "true"
    contents = await file.read()

    wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)

    target_sheet = None
    for name in wb.sheetnames:
        if "apple" in name.lower() and "google" in name.lower():
            target_sheet = name
            break
    if not target_sheet:
        wb.close()
        raise HTTPException(status_code=400, detail=f"Apple & Google sheet not found. Sheets: {wb.sheetnames}")

    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    parsed = []
    for r in rows:
        if not any(r):
            continue
        email = str(r[2] or "").strip().lower() if len(r) > 2 else ""
        if not email or "@" not in email:
            continue
        first_name = str(r[0] or "").strip() if len(r) > 0 else ""
        last_name = str(r[1] or "").strip() if len(r) > 1 else ""
        date_val = r[3] if len(r) > 3 else None
        source_val = str(r[4] or "").strip().lower() if len(r) > 4 else "apple"
        if source_val not in ("apple", "google"):
            source_val = "apple"
        period_start = None
        if hasattr(date_val, "strftime"):
            period_start = date_val.replace(tzinfo=timezone.utc) if date_val.tzinfo is None else date_val
        parsed.append({"first_name": first_name, "last_name": last_name, "email": email, "date": period_start, "source": source_val})

    batch_id = f"meg_ag_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

    async with db_pool.acquire() as conn:
        active_emails = set(r["em"] for r in await conn.fetch(
            "SELECT DISTINCT lower(email) as em FROM subscriptions WHERE email != '' AND status IN ('active', 'trialing')"
        ))
        cancelled_emails = set(r["em"] for r in await conn.fetch(
            "SELECT DISTINCT lower(email) as em FROM subscriptions WHERE email != '' AND status = 'canceled'"
        ))
        all_emails = set(r["em"] for r in await conn.fetch(
            "SELECT DISTINCT lower(email) as em FROM subscriptions WHERE email != ''"
        ))

        new_imports = []
        skip_active = []
        reactivate_candidates = []
        seen = set()

        for p in parsed:
            if p["email"] in seen:
                continue
            seen.add(p["email"])
            if p["email"] in active_emails:
                skip_active.append(p)
            elif p["email"] in cancelled_emails:
                reactivate_candidates.append(p)
            else:
                new_imports.append(p)

        if preview_mode:
            return {
                "status": "preview",
                "total_parsed": len(parsed),
                "unique_emails": len(seen),
                "would_import_new": len(new_imports),
                "would_skip_already_active": len(skip_active),
                "reactivate_candidates": len(reactivate_candidates),
                "note_reactivate": "These are in DB as cancelled but Meg says active. Only Apple/Google source subs will be reactivated (not Stripe).",
                "sample_new": [{"email": p["email"], "source": p["source"], "date": str(p["date"] or "")} for p in new_imports[:15]],
                "sample_reactivate": [{"email": p["email"], "source": p["source"]} for p in reactivate_candidates[:15]],
            }

        count_before = await conn.fetchval("SELECT COUNT(*) FROM subscriptions")
        imported = 0
        reactivated = 0
        errors = 0

        for p in new_imports:
            email_hash = hashlib.md5(p["email"].encode()).hexdigest()[:16]
            syn_id = f"meg_{p['source']}_{email_hash}"
            try:
                await conn.execute("""
                    INSERT INTO subscriptions (
                        stripe_customer_id, stripe_subscription_id, email, status,
                        plan_interval, plan_amount, currency, source,
                        current_period_start, created_at, updated_at, import_batch,
                        first_name, last_name
                    ) VALUES ('', $1, $2, 'active', 'month', 1999, 'usd', $3, $4, COALESCE($5, NOW()), NOW(), $6, $7, $8)
                    ON CONFLICT (stripe_subscription_id) DO NOTHING
                """,
                    syn_id, p["email"], p["source"],
                    p["date"], p["date"], batch_id,
                    p.get("first_name", ""), p.get("last_name", "")
                )
                imported += 1
            except Exception as e:
                print(f"[Meg import] Error: {e}")
                errors += 1

        # Reactivate: only Apple/Google source subs (Stripe webhook is authoritative)
        if skip_reactivate:
            reactivate_candidates = []  # skip all reactivations
        for p in reactivate_candidates:
            try:
                result = await conn.execute("""
                    UPDATE subscriptions SET status = 'active', canceled_at = NULL,
                        reactivated_at = NOW(), updated_at = NOW(), import_batch = $1
                    WHERE id = (
                        SELECT id FROM subscriptions
                        WHERE lower(email) = $2 AND status = 'canceled' AND source IN ('apple', 'google')
                        ORDER BY created_at DESC LIMIT 1
                    )
                """, batch_id, p["email"])
                if result and result.endswith("1"):
                    reactivated += 1
            except Exception as e:
                print(f"[Meg import] Reactivate error: {e}")
                errors += 1

        count_after = await conn.fetchval("SELECT COUNT(*) FROM subscriptions")

    return {
        "status": "ok",
        "batch_id": batch_id,
        "imported_new": imported,
        "reactivated": reactivated,
        "skipped_already_active": len(skip_active),
        "errors": errors,
        "count_before": count_before,
        "count_after": count_after,
        "revert_info": f"To undo: POST /api/admin/revert-batch with batch_id={batch_id}. Note: revert deletes new imports and re-cancels reactivated subs."
    }


# --- Session 16: Export Reactivation Candidates ---

@app.post("/api/admin/reactivation-candidates")
async def reactivation_candidates(request: Request, file: UploadFile = File(...)):
    """Upload Meg Apple/Google XLSX, returns full list of reactivation candidates (cancelled in DB, active in Meg sheet)."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    contents = await file.read()
    wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)

    target_sheet = None
    for name in wb.sheetnames:
        if "apple" in name.lower() and "google" in name.lower():
            target_sheet = name
            break
    if not target_sheet:
        wb.close()
        raise HTTPException(status_code=400, detail="Apple & Google sheet not found")

    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    meg_people = {}
    for r in rows:
        if not any(r):
            continue
        email = str(r[2] or "").strip().lower() if len(r) > 2 else ""
        if not email or "@" not in email:
            continue
        first_name = str(r[0] or "").strip() if len(r) > 0 else ""
        last_name = str(r[1] or "").strip() if len(r) > 1 else ""
        source_val = str(r[4] or "").strip().upper() if len(r) > 4 else "APPLE"
        date_val = r[3].strftime("%Y-%m-%d") if len(r) > 3 and hasattr(r[3], "strftime") else ""
        meg_people[email] = {"first_name": first_name, "last_name": last_name, "source": source_val or "APPLE", "date": date_val}

    async with db_pool.acquire() as conn:
        active_emails = set(r["em"] for r in await conn.fetch(
            "SELECT DISTINCT lower(email) as em FROM subscriptions WHERE email != '' AND status IN ('active', 'trialing')"
        ))
        cancelled_rows = await conn.fetch("""
            SELECT lower(email) as em, source, status, canceled_at, created_at
            FROM subscriptions
            WHERE email != '' AND status = 'canceled'
            AND lower(email) IN (SELECT unnest($1::text[]))
            ORDER BY email, created_at DESC
        """, list(meg_people.keys()))

    # Build cancelled lookup (most recent per email)
    cancelled_map = {}
    for r in cancelled_rows:
        if r["em"] not in cancelled_map:
            cancelled_map[r["em"]] = {"source": r["source"], "canceled_at": str(r["canceled_at"] or ""), "created_at": str(r["created_at"] or "")}

    candidates = []
    for email, info in meg_people.items():
        if email in active_emails:
            continue
        if email in cancelled_map:
            db_info = cancelled_map[email]
            candidates.append({
                "email": email,
                "first_name": info["first_name"],
                "last_name": info["last_name"],
                "meg_source": info["source"],
                "meg_date": info["date"],
                "db_source": db_info["source"],
                "db_canceled_at": db_info["canceled_at"],
                "db_created_at": db_info["created_at"],
            })

    candidates.sort(key=lambda x: x["email"])

    return {
        "total_meg_entries": len(meg_people),
        "already_active_in_db": len([e for e in meg_people if e in active_emails]),
        "reactivation_candidates": len(candidates),
        "candidates": candidates
    }


# --- Session 16: Comprehensive Data Audit ---

@app.get("/api/admin/data-audit")
async def data_audit(request: Request):
    """Run 15+ data quality checks across all tables. Returns issues that affect financial reports."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    audit = {"checks": {}, "summary": {}, "critical_issues": [], "warnings": []}

    async with db_pool.acquire() as conn:

        # ============================================================
        # CHECK 1: Overall counts
        # ============================================================
        total = await conn.fetchval("SELECT COUNT(*) FROM subscriptions")
        active = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE status = 'active'")
        trialing = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE status = 'trialing'")
        canceled = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE status = 'canceled'")
        other_status = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE status NOT IN ('active', 'trialing', 'canceled')")
        audit["checks"]["totals"] = {"total": total, "active": active, "trialing": trialing, "canceled": canceled, "other_status": other_status}

        # ============================================================
        # CHECK 2: Subs with NO email (cannot attribute, cannot match leads)
        # ============================================================
        no_email = await conn.fetch(
            "SELECT source, status, COUNT(*) as cnt FROM subscriptions WHERE email IS NULL OR email = '' GROUP BY source, status ORDER BY cnt DESC"
        )
        no_email_total = sum(r["cnt"] for r in no_email)
        no_email_active = sum(r["cnt"] for r in no_email if r["status"] in ("active", "trialing"))
        audit["checks"]["no_email"] = {
            "total_without_email": no_email_total,
            "active_without_email": no_email_active,
            "by_source_status": [{"source": r["source"], "status": r["status"], "count": r["cnt"]} for r in no_email]
        }
        if no_email_active > 0:
            audit["critical_issues"].append(f"{no_email_active} active/trialing subs have NO email. Cannot attribute to leads, cannot match cancellations, invisible in funnel.")

        # ============================================================
        # CHECK 3: Duplicate active subs (same email, multiple active records = double MRR)
        # ============================================================
        dup_active = await conn.fetch("""
            SELECT lower(email) as em, COUNT(*) as cnt,
                   array_agg(source) as sources,
                   SUM(CASE WHEN plan_interval='month' THEN plan_amount ELSE plan_amount/12 END) as total_mrr_cents
            FROM subscriptions
            WHERE status IN ('active', 'trialing') AND email != '' AND email IS NOT NULL
            GROUP BY lower(email) HAVING COUNT(*) > 1
            ORDER BY cnt DESC LIMIT 50
        """)
        dup_mrr_inflate = sum(r["total_mrr_cents"] - (1999 if r["total_mrr_cents"] else 0) for r in dup_active)
        audit["checks"]["duplicate_active_emails"] = {
            "count": len(dup_active),
            "estimated_mrr_inflation_cents": dup_mrr_inflate,
            "top_duplicates": [{"email": r["em"], "active_records": r["cnt"], "sources": list(r["sources"]), "combined_mrr_cents": r["total_mrr_cents"]} for r in dup_active[:20]]
        }
        if len(dup_active) > 0:
            audit["critical_issues"].append(f"{len(dup_active)} emails have MULTIPLE active subscriptions. This inflates MRR by ~${dup_mrr_inflate/100:,.2f}. Each person should only have 1 active sub.")

        # ============================================================
        # CHECK 4: Plan amount anomalies (wrong amounts = wrong MRR)
        # ============================================================
        plan_dist = await conn.fetch("""
            SELECT plan_amount, plan_interval, source, status, COUNT(*) as cnt
            FROM subscriptions
            WHERE status IN ('active', 'trialing')
            GROUP BY plan_amount, plan_interval, source, status
            ORDER BY cnt DESC
        """)
        zero_amount_active = sum(r["cnt"] for r in plan_dist if (r["plan_amount"] or 0) == 0)
        weird_amounts = [r for r in plan_dist if r["plan_amount"] not in (0, 1999, 17999) and r["plan_amount"] is not None]
        audit["checks"]["plan_amounts"] = {
            "active_with_zero_amount": zero_amount_active,
            "distribution": [{"amount": r["plan_amount"], "interval": r["plan_interval"], "source": r["source"], "status": r["status"], "count": r["cnt"]} for r in plan_dist],
            "unexpected_amounts": [{"amount": r["plan_amount"], "interval": r["plan_interval"], "source": r["source"], "count": r["cnt"]} for r in weird_amounts]
        }
        if zero_amount_active > 0:
            audit["warnings"].append(f"{zero_amount_active} active subs have $0 plan_amount. They count as active but contribute nothing to MRR.")
        if weird_amounts:
            audit["warnings"].append(f"{len(weird_amounts)} plan_amount values are not standard ($19.99/mo or $179.99/yr). Could indicate data corruption.")

        # ============================================================
        # CHECK 5: Stale trialing subs (trial_end in the past but still trialing)
        # ============================================================
        stale_trials = await conn.fetchval("""
            SELECT COUNT(*) FROM subscriptions
            WHERE status = 'trialing' AND trial_end IS NOT NULL AND trial_end < NOW()
        """)
        audit["checks"]["stale_trialing"] = {"count": stale_trials or 0}
        if stale_trials and stale_trials > 0:
            audit["critical_issues"].append(f"{stale_trials} subs are still marked 'trialing' but their trial has ended. These inflate trial count and may be hiding cancellations or conversions.")

        # ============================================================
        # CHECK 6: Future dates (created_at or trial_start in the future)
        # ============================================================
        future_created = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE created_at > NOW() + INTERVAL '1 day'")
        future_trial = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE trial_start > NOW() + INTERVAL '1 day'")
        future_period = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE current_period_start > NOW() + INTERVAL '1 year'")
        audit["checks"]["future_dates"] = {
            "future_created_at": future_created or 0,
            "future_trial_start": future_trial or 0,
            "far_future_period_start": future_period or 0
        }
        if (future_created or 0) > 0:
            audit["critical_issues"].append(f"{future_created} subs have created_at in the future. Distorts daily metrics and cohort analysis.")

        # ============================================================
        # CHECK 7: Source integrity (source vs subscription_id pattern)
        # ============================================================
        stripe_no_sub = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE source = 'stripe' AND stripe_subscription_id NOT LIKE 'sub_%' AND stripe_subscription_id NOT LIKE 'import_%'")
        import_tagged = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE stripe_subscription_id LIKE 'import_%'")
        apple_count = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE source = 'apple'")
        google_count = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE source = 'google'")
        stripe_count = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE source = 'stripe'")
        audit["checks"]["source_integrity"] = {
            "stripe_total": stripe_count or 0,
            "apple_total": apple_count or 0,
            "google_total": google_count or 0,
            "import_tagged_records": import_tagged or 0,
            "stripe_without_sub_prefix": stripe_no_sub or 0
        }

        # ============================================================
        # CHECK 8: MRR accuracy (calculated vs what dashboard would show)
        # ============================================================
        mrr_monthly = await conn.fetchval("SELECT COALESCE(SUM(plan_amount), 0) FROM subscriptions WHERE status = 'active' AND plan_interval = 'month'")
        mrr_annual = await conn.fetchval("SELECT COALESCE(SUM(plan_amount / 12), 0) FROM subscriptions WHERE status = 'active' AND plan_interval = 'year'")
        mrr_total = (mrr_monthly or 0) + (mrr_annual or 0)
        mrr_by_src = await conn.fetch("""
            SELECT source,
                COUNT(*) as active_count,
                COALESCE(SUM(CASE WHEN plan_interval='month' THEN plan_amount ELSE 0 END), 0) as monthly,
                COALESCE(SUM(CASE WHEN plan_interval='year' THEN plan_amount/12 ELSE 0 END), 0) as annual_equiv
            FROM subscriptions WHERE status = 'active'
            GROUP BY source
        """)
        audit["checks"]["mrr_breakdown"] = {
            "total_mrr_cents": mrr_total,
            "total_mrr_display": f"${mrr_total/100:,.2f}",
            "from_monthly": mrr_monthly or 0,
            "from_annual": mrr_annual or 0,
            "by_source": [{"source": r["source"], "active_count": r["active_count"], "mrr_cents": (r["monthly"] or 0) + (r["annual_equiv"] or 0)} for r in mrr_by_src]
        }

        # ============================================================
        # CHECK 9: Missing conversion stamps
        # ============================================================
        should_have_converted = await conn.fetchval("""
            SELECT COUNT(*) FROM subscriptions
            WHERE converted_at IS NULL
            AND trial_end IS NOT NULL AND current_period_start IS NOT NULL
            AND current_period_start > trial_end
            AND status = 'active'
        """)
        has_converted = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE converted_at IS NOT NULL")
        trial_with_data = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE trial_start IS NOT NULL")
        trial_no_data = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE trial_start IS NULL")
        audit["checks"]["conversion_tracking"] = {
            "has_converted_at": has_converted or 0,
            "missing_converted_at_but_clearly_converted": should_have_converted or 0,
            "subs_with_trial_data": trial_with_data or 0,
            "subs_without_trial_data": trial_no_data or 0
        }
        if should_have_converted and should_have_converted > 10:
            audit["warnings"].append(f"{should_have_converted} active subs clearly converted from trial but lack converted_at stamp. Run backfill-conversions to fix.")

        # ============================================================
        # CHECK 10: Active subs that Meg lists as cancelled
        # (Cross-reference: how many of our "active" subs might actually be cancelled?)
        # ============================================================
        active_stripe_count = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE source = 'stripe' AND status = 'active'")
        active_apple_count = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE source = 'apple' AND status IN ('active', 'trialing')")
        active_google_count = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE source = 'google' AND status IN ('active', 'trialing')")
        audit["checks"]["active_by_source"] = {
            "stripe_active": active_stripe_count or 0,
            "apple_active": active_apple_count or 0,
            "google_active": active_google_count or 0,
            "note": "Compare these against Meg spreadsheet: 866 Stripe active, 796 Apple/Google active. Large discrepancies = stale data."
        }

        # ============================================================
        # CHECK 11: Subscription age analysis (are old subs still marked active?)
        # ============================================================
        old_active = await conn.fetch("""
            SELECT source, COUNT(*) as cnt,
                MIN(created_at) as oldest,
                MAX(created_at) as newest
            FROM subscriptions
            WHERE status = 'active'
            GROUP BY source
        """)
        very_old = await conn.fetchval("""
            SELECT COUNT(*) FROM subscriptions
            WHERE status = 'active' AND created_at < NOW() - INTERVAL '2 years'
        """)
        audit["checks"]["active_age"] = {
            "by_source": [{"source": r["source"], "active_count": r["cnt"], "oldest": str(r["oldest"]), "newest": str(r["newest"])} for r in old_active],
            "active_older_than_2_years": very_old or 0
        }
        if very_old and very_old > 50:
            audit["warnings"].append(f"{very_old} subs marked active are over 2 years old. Some may be zombies that cancelled outside our tracking.")

        # ============================================================
        # CHECK 12: period_end in the past for active subs (expired but not flipped)
        # ============================================================
        expired_active = await conn.fetchval("""
            SELECT COUNT(*) FROM subscriptions
            WHERE status = 'active'
            AND current_period_end IS NOT NULL
            AND current_period_end < NOW() - INTERVAL '7 days'
        """)
        audit["checks"]["expired_but_active"] = {"count": expired_active or 0}
        if expired_active and expired_active > 20:
            audit["critical_issues"].append(f"{expired_active} subs are marked 'active' but their current_period_end is over 7 days ago. These are likely cancelled/expired and inflate MRR.")

        # ============================================================
        # CHECK 13: Leads vs Subs email match rate
        # ============================================================
        total_leads = await conn.fetchval("SELECT COUNT(*) FROM leads WHERE email != ''")
        leads_with_sub = await conn.fetchval("""
            SELECT COUNT(DISTINCT lower(l.email)) FROM leads l
            INNER JOIN subscriptions s ON lower(l.email) = lower(s.email)
            WHERE l.email != '' AND s.email != ''
        """)
        subs_with_lead = await conn.fetchval("""
            SELECT COUNT(DISTINCT lower(s.email)) FROM subscriptions s
            INNER JOIN leads l ON lower(s.email) = lower(l.email)
            WHERE s.email != '' AND l.email != '' AND s.status IN ('active', 'trialing')
        """)
        active_with_email = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE status IN ('active', 'trialing') AND email != '' AND email IS NOT NULL")
        audit["checks"]["lead_attribution"] = {
            "total_leads": total_leads or 0,
            "leads_who_subscribed": leads_with_sub or 0,
            "active_subs_with_matching_lead": subs_with_lead or 0,
            "active_subs_with_email": active_with_email or 0,
            "attribution_rate": round(((subs_with_lead or 0) / max(active_with_email or 1, 1)) * 100, 1)
        }

        # ============================================================
        # CHECK 14: Cancelled subs with NO canceled_at date
        # ============================================================
        cancel_no_date = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE status = 'canceled' AND canceled_at IS NULL")
        audit["checks"]["cancel_date_quality"] = {
            "canceled_without_date": cancel_no_date or 0,
            "note": "Missing canceled_at means churn timing is unknown. Affects churn rate calculations."
        }

        # ============================================================
        # CHECK 15: Net MRR risk score
        # ============================================================
        risk_factors = []
        risk_score = 0
        if (expired_active or 0) > 20:
            risk_factors.append(f"~{expired_active} expired-but-active subs inflating MRR")
            risk_score += 3
        if len(dup_active) > 10:
            risk_factors.append(f"~{len(dup_active)} duplicate active emails inflating MRR")
            risk_score += 3
        if no_email_active > 100:
            risk_factors.append(f"~{no_email_active} active subs with no email (unverifiable)")
            risk_score += 2
        if (stale_trials or 0) > 10:
            risk_factors.append(f"~{stale_trials} stale trialing subs (trial ended)")
            risk_score += 1
        if (very_old or 0) > 50:
            risk_factors.append(f"~{very_old} active subs over 2 years old (possible zombies)")
            risk_score += 2
        if zero_amount_active > 20:
            risk_factors.append(f"~{zero_amount_active} active subs with $0 plan amount")
            risk_score += 1

        if risk_score >= 6:
            confidence = "LOW"
        elif risk_score >= 3:
            confidence = "MEDIUM"
        else:
            confidence = "HIGH"

        audit["summary"] = {
            "mrr_confidence": confidence,
            "risk_score": risk_score,
            "risk_factors": risk_factors,
            "total_critical_issues": len(audit["critical_issues"]),
            "total_warnings": len(audit["warnings"])
        }

    return audit


# --- Session 25: Per-Email Diff vs ymove (READ-ONLY) ---

@app.post("/api/admin/ymove-diff")
async def ymove_diff(request: Request):
    """S25: Read-only diff between our active+trialing subs and ymove's subscribed list.
    No writes, no corrections. Pulls ymove fresh each call.
    Returns 4 buckets + a preflight section to test the 'cancelled coming back' theory."""
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)

    if not YMOVE_API_KEY:
        return JSONResponse(status_code=500, content={"error": "YMOVE_API_KEY not set"})
    if not db_pool:
        return JSONResponse(status_code=500, content={"error": "No database connected"})

    CAP = 100  # max records per bucket in response

    # ---------- PRE-FLIGHT: test the "cancelled coming back" theory ----------
    preflight = {}
    async with db_pool.acquire() as conn:
        # Check 1: Stripe-source records by ID prefix. Real Stripe webhooks ALWAYS start with sub_.
        # Anything else under source='stripe' is a reconciliation artifact and a prime suspect.
        rows = await conn.fetch("""
            SELECT
              CASE
                WHEN stripe_subscription_id LIKE 'sub_%' THEN 'sub_ (real Stripe webhook)'
                WHEN stripe_subscription_id LIKE 'ymove_new_stripe_%' THEN 'ymove_new_stripe_ (synthetic)'
                WHEN stripe_subscription_id LIKE 'ymove_switch_stripe_%' THEN 'ymove_switch_stripe_ (synthetic)'
                WHEN stripe_subscription_id LIKE 'import_%' THEN 'import_ (Meg spreadsheet)'
                WHEN stripe_subscription_id LIKE 'meg_%' THEN 'meg_ (Meg spreadsheet)'
                ELSE 'OTHER: ' || COALESCE(LEFT(stripe_subscription_id, 20), 'NULL')
              END AS prefix_bucket,
              COUNT(*) AS n
            FROM subscriptions
            WHERE source = 'stripe' AND status IN ('active', 'trialing')
            GROUP BY 1
            ORDER BY n DESC
        """)
        preflight["stripe_id_prefix_breakdown"] = [{"prefix": r["prefix_bucket"], "count": r["n"]} for r in rows]

        # Check 2: Active records whose email also has a cancelled S23/S24/S25 cleanup-batch record.
        # This is the "removed but came back" pattern.
        rows = await conn.fetch("""
            SELECT
              s_active.email,
              s_active.source AS active_source,
              s_active.stripe_subscription_id AS active_sub_id,
              s_active.created_at AS active_created_at,
              s_cancelled.import_batch AS cancelled_batch,
              s_cancelled.stripe_subscription_id AS cancelled_sub_id
            FROM subscriptions s_active
            JOIN subscriptions s_cancelled ON LOWER(s_cancelled.email) = LOWER(s_active.email)
            WHERE s_active.status IN ('active', 'trialing')
              AND s_cancelled.status = 'canceled'
              AND (
                s_cancelled.import_batch LIKE 's23_%' OR
                s_cancelled.import_batch LIKE 's24_%' OR
                s_cancelled.import_batch LIKE 's25_%'
              )
              AND s_active.id != s_cancelled.id
            ORDER BY s_active.email
            LIMIT 200
        """)
        preflight["active_with_cancelled_cleanup_sibling"] = {
            "count": len(rows),
            "records": [
                {
                    "email": r["email"],
                    "active_source": r["active_source"],
                    "active_sub_id": r["active_sub_id"],
                    "active_created_at": r["active_created_at"].isoformat() if r["active_created_at"] else None,
                    "cancelled_batch": r["cancelled_batch"],
                    "cancelled_sub_id": r["cancelled_sub_id"],
                }
                for r in rows[:CAP]
            ],
            "truncated": len(rows) > CAP,
        }

    # ---------- PHASE 1: Pull ymove fresh ----------
    ymove_emails = {}  # email -> provider
    pull_status = "starting"
    pages_pulled = 0
    try:
        page = 1
        async with httpx.AsyncClient(timeout=30.0) as client:
            while True:
                try:
                    resp = await client.get(
                        f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup/all",
                        headers={"X-Authorization": YMOVE_API_KEY},
                        params={"status": "subscribed", "page": str(page)}
                    )
                except httpx.TimeoutException:
                    pull_status = f"timeout_page_{page}"
                    break

                if resp.status_code == 429:
                    retry_after = float(resp.headers.get("retry-after", "5"))
                    await asyncio.sleep(retry_after)
                    continue

                if resp.status_code != 200:
                    pull_status = f"error_page_{page}_http_{resp.status_code}"
                    break

                data = resp.json()
                users = data.get("users", [])
                for u in users:
                    em = (u.get("email") or "").strip().lower()
                    if em:
                        provider = (u.get("subscriptionProvider") or "undetermined").lower()
                        if provider not in ("apple", "google", "stripe", "manual", "undetermined"):
                            provider = "undetermined"
                        ymove_emails[em] = provider

                total_pages = data.get("totalPages", 1)
                pages_pulled = page
                if page >= total_pages:
                    pull_status = "success"
                    break
                page += 1
                await asyncio.sleep(0.5)
    except Exception as e:
        pull_status = f"error: {str(e)[:200]}"

    if pull_status != "success":
        return JSONResponse(status_code=502, content={
            "error": "ymove pull failed",
            "pull_status": pull_status,
            "pages_pulled": pages_pulled,
            "preflight": preflight,
        })

    # ---------- PHASE 2: Pull our active+trialing ----------
    our_subs = {}  # email -> dict
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT email, source, status, stripe_subscription_id, created_at, import_batch
            FROM subscriptions
            WHERE status IN ('active', 'trialing')
              AND email IS NOT NULL AND email != ''
        """)
        for r in rows:
            em = r["email"].strip().lower()
            our_subs[em] = {
                "source": r["source"],
                "status": r["status"],
                "sub_id": r["stripe_subscription_id"],
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "import_batch": r["import_batch"],
            }

    # ---------- PHASE 3: Compute the 4 diff buckets ----------
    only_in_ours = []      # active in ours, not in ymove subscribed
    only_in_ymove = []     # in ymove subscribed, not active in ours
    provider_mismatch = [] # in both, source != provider (excluding stripe<->stripe)
    in_both_aligned = 0    # for sanity-check counting

    our_emails_set = set(our_subs.keys())
    ymove_emails_set = set(ymove_emails.keys())

    for em in our_emails_set - ymove_emails_set:
        rec = our_subs[em]
        only_in_ours.append({
            "email": em,
            "our_source": rec["source"],
            "our_status": rec["status"],
            "our_sub_id": rec["sub_id"],
            "created_at": rec["created_at"],
            "import_batch": rec["import_batch"],
        })

    for em in ymove_emails_set - our_emails_set:
        only_in_ymove.append({
            "email": em,
            "ymove_provider": ymove_emails[em],
        })

    for em in our_emails_set & ymove_emails_set:
        rec = our_subs[em]
        ymove_provider = ymove_emails[em]
        our_source = rec["source"]
        # ymove "undetermined" means their API returned null for this user — not a real mismatch
        if ymove_provider == "undetermined":
            in_both_aligned += 1
            continue
        if our_source == ymove_provider:
            in_both_aligned += 1
        else:
            provider_mismatch.append({
                "email": em,
                "our_source": our_source,
                "ymove_provider": ymove_provider,
                "our_sub_id": rec["sub_id"],
                "our_status": rec["status"],
                "created_at": rec["created_at"],
            })

    # Sort for stable output
    only_in_ours.sort(key=lambda x: (x["our_source"] or "", x["email"]))
    only_in_ymove.sort(key=lambda x: (x["ymove_provider"], x["email"]))
    provider_mismatch.sort(key=lambda x: (x["our_source"] or "", x["ymove_provider"], x["email"]))

    return {
        "preflight": preflight,
        "summary": {
            "our_active_trialing_total": len(our_subs),
            "ymove_subscribed_total": len(ymove_emails),
            "in_both_aligned": in_both_aligned,
            "only_in_ours_count": len(only_in_ours),
            "only_in_ymove_count": len(only_in_ymove),
            "provider_mismatch_count": len(provider_mismatch),
            "ymove_pages_pulled": pages_pulled,
        },
        "only_in_ours": {
            "count": len(only_in_ours),
            "records": only_in_ours[:CAP],
            "truncated": len(only_in_ours) > CAP,
        },
        "only_in_ymove": {
            "count": len(only_in_ymove),
            "records": only_in_ymove[:CAP],
            "truncated": len(only_in_ymove) > CAP,
        },
        "provider_mismatch": {
            "count": len(provider_mismatch),
            "records": provider_mismatch[:CAP],
            "truncated": len(provider_mismatch) > CAP,
        },
    }


# --- Session 25: Verify Isabella + Jess against Stripe API ---

@app.post("/api/admin/verify-isabella-jess-s25")
async def verify_isabella_jess_s25(request: Request):
    """S25: Read-only Stripe API check for the 2 only_in_ymove records.
    Both have cancelled Stripe records in our DB. We need to know if Stripe agrees.
    """
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)
    if not db_pool:
        return JSONResponse(status_code=500, content={"error": "No database connected"})

    EMAILS = ["isabella.marovich-tadic@menzies.edu.au", "jessica.mullen@yahoo.com"]

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT id, email, source, status, stripe_subscription_id, stripe_customer_id,
                   plan_amount, created_at, canceled_at
            FROM subscriptions
            WHERE LOWER(email) = ANY($1::text[])
              AND source = 'stripe'
            ORDER BY email, created_at DESC
        """, [e.lower() for e in EMAILS])

    results = []
    for r in rows:
        sub_id = r["stripe_subscription_id"]
        record = {
            "id": r["id"],
            "email": r["email"],
            "our_status": r["status"],
            "our_sub_id": sub_id,
            "stripe_customer_id": r["stripe_customer_id"],
            "our_canceled_at": r["canceled_at"].isoformat() if r["canceled_at"] else None,
        }
        # Check this specific sub_id
        try:
            sub = stripe.Subscription.retrieve(sub_id)
            record["this_sub_status"] = sub.status
            record["this_sub_canceled_at"] = sub.canceled_at
        except Exception as e:
            record["this_sub_error"] = str(e)[:200]

        # Also check if customer has ANY other active sub (the resubscribe theory)
        try:
            customers = stripe.Customer.list(email=r["email"], limit=10)
            all_subs = []
            for cust in customers.data:
                cust_subs = stripe.Subscription.list(customer=cust.id, status="all", limit=20)
                for s in cust_subs.data:
                    all_subs.append({
                        "sub_id": s.id,
                        "status": s.status,
                        "created": s.created,
                        "current_period_end": s.current_period_end,
                        "customer_id": cust.id,
                    })
            record["all_stripe_subs_for_email"] = all_subs
            record["any_currently_active"] = any(s["status"] in ("active", "trialing") for s in all_subs)
        except Exception as e:
            record["customer_lookup_error"] = str(e)[:200]

        results.append(record)

    return {"results": results, "count": len(results)}


# --- Session 25: Investigate 2 new only_in_ymove + false-cancelled Stripe scan ---

S25_NEW_YMOVE_EMAILS = [
    "isabella.marovich-tadic@menzies.edu.au",
    "jessica.mullen@yahoo.com",
]

@app.post("/api/admin/investigate-stripe-gaps-s25")
async def investigate_stripe_gaps_s25(request: Request):
    """S25: Two-part read-only diagnostic.
    Part 1: For the 2 specific only_in_ymove emails, dump all DB records (any status).
    Part 2: Find cancelled Stripe records where ymove still considers the user subscribed.
    """
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)
    if not YMOVE_API_KEY or not db_pool:
        return JSONResponse(status_code=500, content={"error": "Missing config"})

    # Part 1: dump all records for the 2 specific emails
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT id, email, source, status, stripe_subscription_id, plan_amount,
                   created_at, canceled_at, import_batch
            FROM subscriptions
            WHERE LOWER(email) = ANY($1::text[])
            ORDER BY email, created_at DESC
        """, [e.lower() for e in S25_NEW_YMOVE_EMAILS])
    part1 = [
        {
            "id": r["id"], "email": r["email"], "source": r["source"], "status": r["status"],
            "sub_id": r["stripe_subscription_id"], "plan_amount": r["plan_amount"],
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            "canceled_at": r["canceled_at"].isoformat() if r["canceled_at"] else None,
            "import_batch": r["import_batch"],
        } for r in rows
    ]

    # Part 2: pull ymove subscribed list, then find cancelled Stripe records that ymove says are active
    ymove_emails = set()
    pull_status = "starting"
    pages_pulled = 0
    try:
        page = 1
        async with httpx.AsyncClient(timeout=30.0) as client:
            while True:
                try:
                    resp = await client.get(
                        f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup/all",
                        headers={"X-Authorization": YMOVE_API_KEY},
                        params={"status": "subscribed", "page": str(page)}
                    )
                except httpx.TimeoutException:
                    pull_status = f"timeout_page_{page}"; break
                if resp.status_code == 429:
                    await asyncio.sleep(float(resp.headers.get("retry-after", "5"))); continue
                if resp.status_code != 200:
                    pull_status = f"error_page_{page}_http_{resp.status_code}"; break
                data = resp.json()
                for u in data.get("users", []):
                    em = (u.get("email") or "").strip().lower()
                    if em:
                        ymove_emails.add(em)
                total_pages = data.get("totalPages", 1)
                pages_pulled = page
                if page >= total_pages:
                    pull_status = "success"; break
                page += 1
                await asyncio.sleep(0.5)
    except Exception as e:
        pull_status = f"error: {str(e)[:200]}"

    if pull_status != "success":
        return JSONResponse(status_code=502, content={
            "error": "ymove pull failed", "pull_status": pull_status,
            "pages_pulled": pages_pulled, "part1": part1,
        })

    # Get our cancelled Stripe records + active emails (to filter normal churn-and-rejoin)
    async with db_pool.acquire() as conn:
        cancelled_stripe = await conn.fetch("""
            SELECT id, email, source, status, stripe_subscription_id, plan_amount,
                   created_at, canceled_at, import_batch
            FROM subscriptions
            WHERE status = 'canceled' AND source = 'stripe'
              AND email IS NOT NULL AND email != ''
              AND stripe_subscription_id LIKE 'sub_%'
            ORDER BY email, canceled_at DESC NULLS LAST
        """)
        active_emails_rows = await conn.fetch("""
            SELECT LOWER(email) AS email FROM subscriptions
            WHERE status IN ('active', 'trialing') AND email IS NOT NULL AND email != ''
        """)
    our_active_emails = {r["email"] for r in active_emails_rows}

    false_cancel_stripe = []
    seen = set()
    for r in cancelled_stripe:
        em = r["email"].strip().lower()
        if em in seen: continue
        if em not in ymove_emails: continue
        if em in our_active_emails: continue
        seen.add(em)
        false_cancel_stripe.append({
            "id": r["id"], "email": r["email"], "sub_id": r["stripe_subscription_id"],
            "plan_amount": r["plan_amount"],
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            "canceled_at": r["canceled_at"].isoformat() if r["canceled_at"] else None,
            "import_batch": r["import_batch"],
        })

    return {
        "part1_specific_emails": {
            "queried": S25_NEW_YMOVE_EMAILS,
            "found_records": part1,
            "found_count": len(part1),
        },
        "part2_false_cancelled_stripe": {
            "count": len(false_cancel_stripe),
            "records": false_cancel_stripe[:100],
            "truncated": len(false_cancel_stripe) > 100,
        },
        "ymove_subscribed_total": len(ymove_emails),
        "our_cancelled_stripe_total": len(cancelled_stripe),
    }


# --- Session 25: Verify 8 historical Stripe records against live Stripe API ---

S25_HISTORICAL_STRIPE_EMAILS = [
    "abbey.e.baier@gmail.com",
    "ahfouch@gmail.com",
    "alessandraclelia.volpato@gmail.com",
    "amets30@yahoo.com",
    "andreedesrochers@gmail.com",
    "cassyroop@gmail.com",
    "chloe.levray@gmail.com",
    "hstrandness@gmail.com",
]

@app.post("/api/admin/verify-historical-stripe-s25")
async def verify_historical_stripe_s25(request: Request):
    """S25 Step 5: Read-only. Hits Stripe API for each of the 8 historical Stripe records
    that ymove has no record of. Returns Stripe's authoritative status per record.
    No writes."""
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)
    if not db_pool:
        return JSONResponse(status_code=500, content={"error": "No database connected"})

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT id, email, source, status, stripe_subscription_id, plan_amount,
                   created_at, current_period_end
            FROM subscriptions
            WHERE LOWER(email) = ANY($1::text[])
              AND status IN ('active', 'trialing')
              AND source = 'stripe'
            ORDER BY email
        """, [e.lower() for e in S25_HISTORICAL_STRIPE_EMAILS])

    results = []
    for r in rows:
        sub_id = r["stripe_subscription_id"]
        record = {
            "id": r["id"],
            "email": r["email"],
            "our_status": r["status"],
            "our_sub_id": sub_id,
            "plan_amount": r["plan_amount"],
            "our_created_at": r["created_at"].isoformat() if r["created_at"] else None,
            "our_current_period_end": r["current_period_end"].isoformat() if r["current_period_end"] else None,
        }
        if not sub_id or not sub_id.startswith("sub_"):
            record["stripe_check"] = "skipped_not_sub_prefix"
            results.append(record)
            continue
        try:
            real_sub = stripe.Subscription.retrieve(sub_id)
            record["stripe_status"] = real_sub.status
            record["stripe_current_period_end"] = real_sub.current_period_end
            record["stripe_cancel_at_period_end"] = real_sub.cancel_at_period_end
            record["stripe_canceled_at"] = real_sub.canceled_at
            record["agreement"] = "MATCH" if real_sub.status == r["status"] else "MISMATCH"
        except stripe.error.InvalidRequestError as e:
            record["stripe_check"] = "not_found_in_stripe"
            record["stripe_error"] = str(e)[:200]
        except Exception as e:
            record["stripe_check"] = "error"
            record["stripe_error"] = str(e)[:200]
        results.append(record)

    summary = {
        "checked": len(results),
        "expected": len(S25_HISTORICAL_STRIPE_EMAILS),
        "by_stripe_status": {},
        "by_agreement": {},
    }
    for r in results:
        s = r.get("stripe_status", r.get("stripe_check", "unknown"))
        summary["by_stripe_status"][s] = summary["by_stripe_status"].get(s, 0) + 1
        a = r.get("agreement", "n/a")
        summary["by_agreement"][a] = summary["by_agreement"].get(a, 0) + 1

    return {"summary": summary, "results": results}


# --- Session 25: Reactivate kelsey (id=11315) ---

@app.post("/api/admin/reactivate-kelsey-s25")
async def reactivate_kelsey_s25(request: Request):
    """S25 Step 4 (revised): Reactivate the existing cancelled Apple record for kelseymsimms.
    Hardcoded id=11315, with defensive checks that the record matches what we expect.
    """
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)
    if not db_pool:
        return JSONResponse(status_code=500, content={"error": "No database connected"})

    try:
        body = await request.json()
    except Exception:
        body = {}
    apply = bool(body.get("apply", False))

    EXPECTED_ID = 11315
    EXPECTED_EMAIL = "kelseymsimms@gmail.com"
    EXPECTED_SOURCE = "apple"
    EXPECTED_STATUS_BEFORE = "canceled"
    EXPECTED_BATCH_PREFIX = "shadow_6_"

    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, email, source, status, stripe_subscription_id, plan_amount, import_batch, canceled_at FROM subscriptions WHERE id = $1",
            EXPECTED_ID
        )
        if not row:
            return JSONResponse(status_code=404, content={"error": f"id={EXPECTED_ID} not found"})

        # Defensive checks
        checks = {
            "email_matches": row["email"].lower() == EXPECTED_EMAIL,
            "source_matches": row["source"] == EXPECTED_SOURCE,
            "status_matches": row["status"] == EXPECTED_STATUS_BEFORE,
            "batch_matches": (row["import_batch"] or "").startswith(EXPECTED_BATCH_PREFIX),
        }
        all_passed = all(checks.values())

        record = {
            "id": row["id"],
            "email": row["email"],
            "source": row["source"],
            "status": row["status"],
            "sub_id": row["stripe_subscription_id"],
            "plan_amount": row["plan_amount"],
            "import_batch": row["import_batch"],
            "canceled_at": row["canceled_at"].isoformat() if row["canceled_at"] else None,
        }

        if not all_passed:
            return {
                "mode": "blocked",
                "reason": "Defensive check failed — record does not match expected state",
                "checks": checks,
                "record": record,
            }

        if not apply:
            return {
                "mode": "preview",
                "checks": checks,
                "record": record,
                "action": "Will set status='active', canceled_at=NULL, append S25 batch tag",
                "next_step": 'POST again with {"apply": true} to commit',
            }

        from datetime import datetime as _dt
        batch_tag = f"s25_kelsey_reactivate_{_dt.utcnow().strftime('%Y%m%d_%H%M%S')}"
        await conn.execute(
            """UPDATE subscriptions
               SET status = 'active',
                   canceled_at = NULL,
                   reactivated_at = NOW(),
                   updated_at = NOW(),
                   import_batch = COALESCE(import_batch, '') || ' ' || $1
               WHERE id = $2 AND status = 'canceled'""",
            batch_tag, EXPECTED_ID
        )

        return {
            "mode": "applied",
            "batch_tag": batch_tag,
            "id": EXPECTED_ID,
            "record_before": record,
            "reversal_hint": f"To revert: UPDATE subscriptions SET status='canceled', canceled_at=NOW() WHERE id={EXPECTED_ID}",
        }


# --- Session 25: Diagnostic — find false-cancelled records ---

@app.post("/api/admin/find-false-cancelled-s25")
async def find_false_cancelled_s25(request: Request):
    """S25 Step 4b: Read-only diagnostic.
    Finds records where status='canceled' AND source IN ('apple','google','undetermined')
    BUT the email is currently active in ymove's subscribed list, AND we don't already
    have a separate active record for that email (which would be normal churn-and-rejoin).
    These are 'false cancels' — likely victims of Gap 1, Bug 2, or S23/S24 cleanup overreach.
    """
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)

    if not YMOVE_API_KEY:
        return JSONResponse(status_code=500, content={"error": "YMOVE_API_KEY not set"})
    if not db_pool:
        return JSONResponse(status_code=500, content={"error": "No database connected"})

    # Phase 1: pull ymove subscribed list fresh
    ymove_emails = {}
    pull_status = "starting"
    pages_pulled = 0
    try:
        page = 1
        async with httpx.AsyncClient(timeout=30.0) as client:
            while True:
                try:
                    resp = await client.get(
                        f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup/all",
                        headers={"X-Authorization": YMOVE_API_KEY},
                        params={"status": "subscribed", "page": str(page)}
                    )
                except httpx.TimeoutException:
                    pull_status = f"timeout_page_{page}"
                    break
                if resp.status_code == 429:
                    await asyncio.sleep(float(resp.headers.get("retry-after", "5")))
                    continue
                if resp.status_code != 200:
                    pull_status = f"error_page_{page}_http_{resp.status_code}"
                    break
                data = resp.json()
                for u in data.get("users", []):
                    em = (u.get("email") or "").strip().lower()
                    if em:
                        provider = (u.get("subscriptionProvider") or "undetermined").lower()
                        ymove_emails[em] = provider
                total_pages = data.get("totalPages", 1)
                pages_pulled = page
                if page >= total_pages:
                    pull_status = "success"
                    break
                page += 1
                await asyncio.sleep(0.5)
    except Exception as e:
        pull_status = f"error: {str(e)[:200]}"

    if pull_status != "success":
        return JSONResponse(status_code=502, content={
            "error": "ymove pull failed",
            "pull_status": pull_status,
            "pages_pulled": pages_pulled,
        })

    # Phase 2: pull our cancelled Apple/Google/Undetermined records
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT id, email, source, status, stripe_subscription_id, plan_amount,
                   created_at, canceled_at, import_batch
            FROM subscriptions
            WHERE status = 'canceled'
              AND source IN ('apple', 'google', 'undetermined')
              AND email IS NOT NULL AND email != ''
            ORDER BY email, canceled_at DESC NULLS LAST
        """)
        active_emails_rows = await conn.fetch("""
            SELECT LOWER(email) AS email FROM subscriptions
            WHERE status IN ('active', 'trialing')
              AND email IS NOT NULL AND email != ''
        """)
    our_active_emails = {r["email"] for r in active_emails_rows}

    false_cancels = []
    seen_emails = set()
    for r in rows:
        em = r["email"].strip().lower()
        if em in seen_emails:
            continue
        if em not in ymove_emails:
            continue
        if em in our_active_emails:
            continue
        seen_emails.add(em)
        false_cancels.append({
            "id": r["id"],
            "email": r["email"],
            "our_source": r["source"],
            "our_sub_id": r["stripe_subscription_id"],
            "plan_amount": r["plan_amount"],
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            "canceled_at": r["canceled_at"].isoformat() if r["canceled_at"] else None,
            "import_batch": r["import_batch"],
            "ymove_provider": ymove_emails[em],
        })

    by_batch = {}
    by_source = {}
    for fc in false_cancels:
        b = fc["import_batch"] or "(none)"
        by_batch[b] = by_batch.get(b, 0) + 1
        by_source[fc["our_source"]] = by_source.get(fc["our_source"], 0) + 1

    return {
        "summary": {
            "false_cancel_count": len(false_cancels),
            "ymove_subscribed_total": len(ymove_emails),
            "our_cancelled_ag_total": len(rows),
            "by_source": by_source,
            "by_import_batch": by_batch,
        },
        "false_cancels": false_cancels[:200],
        "truncated": len(false_cancels) > 200,
    }


# --- Session 25: Backfill missing Apple sub (kelseymsimms) ---

@app.post("/api/admin/backfill-kelsey-s25")
async def backfill_kelsey_s25(request: Request):
    """S25 Step 4: Insert kelseymsimms@gmail.com — Apple sub ymove has but we don't.
    Hardcoded single record. Verified live against ymove inspect-user before building.
    Body: {"apply": false} for preview, {"apply": true} to commit.
    """
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)

    if not db_pool:
        return JSONResponse(status_code=500, content={"error": "No database connected"})

    try:
        body = await request.json()
    except Exception:
        body = {}
    apply = bool(body.get("apply", False))

    record = {
        "email": "kelseymsimms@gmail.com",
        "source": "apple",
        "status": "active",
        "stripe_subscription_id": "s25_backfill_kelseymsimms_apple",
        "plan_amount": 1999,  # $19.99/mo monthly
        "created_at": "2024-07-04T21:10:43+00:00",
        "first_name": "Kelsey",
        "last_name": "Simms",
        "ymove_user_id": 991697970,
    }

    async with db_pool.acquire() as conn:
        # Safety: confirm she's not already in the DB (in any status)
        existing = await conn.fetch(
            "SELECT id, status, source, stripe_subscription_id FROM subscriptions WHERE LOWER(email) = $1",
            record["email"].lower()
        )

        if existing:
            return {
                "mode": "blocked",
                "reason": "Email already exists in DB",
                "existing_records": [dict(r) for r in existing],
            }

        if not apply:
            return {
                "mode": "preview",
                "to_insert": record,
                "next_step": 'POST again with {"apply": true} to commit',
            }

        from datetime import datetime as _dt
        batch_tag = f"s25_kelsey_backfill_{_dt.utcnow().strftime('%Y%m%d_%H%M%S')}"
        new_id = await conn.fetchval(
            """INSERT INTO subscriptions
               (email, source, status, stripe_subscription_id, plan_amount, plan_interval,
                created_at, updated_at, import_batch, first_name, last_name)
               VALUES ($1, $2, $3, $4, $5, 'month', $6::timestamptz, NOW(), $7, $8, $9)
               RETURNING id""",
            record["email"], record["source"], record["status"],
            record["stripe_subscription_id"], record["plan_amount"],
            record["created_at"], batch_tag,
            record["first_name"], record["last_name"]
        )

        return {
            "mode": "applied",
            "batch_tag": batch_tag,
            "inserted_id": new_id,
            "record": record,
            "reversal_hint": f"To revert: DELETE FROM subscriptions WHERE id = {new_id}",
        }


# --- Session 25: Cancel test/junk Stripe records ---

S25_TEST_CANCEL_EMAILS = [
    "sfdasafsaffas@ymove.app",
    "sfdfdssfdfsdfsdasfad@ymove.app",
    "utm_sourceemail@ymove.app",
    "markus.zwigart@gmx.dr",
    "tosh.koevoets@gmail.com",
]

@app.post("/api/admin/cancel-test-stripe-s25")
async def cancel_test_stripe_s25(request: Request):
    """S25 Step 3: Soft-cancel 5 specific test/junk Stripe records identified in the ymove diff.
    Hardcoded allowlist — endpoint cannot cancel anything else.
    Body: {"apply": false} for preview, {"apply": true} to commit.
    Reversible via batch tag.
    """
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)

    if not db_pool:
        return JSONResponse(status_code=500, content={"error": "No database connected"})

    try:
        body = await request.json()
    except Exception:
        body = {}
    apply = bool(body.get("apply", False))

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT id, email, source, status, stripe_subscription_id, plan_amount, created_at
            FROM subscriptions
            WHERE LOWER(email) = ANY($1::text[])
              AND status IN ('active', 'trialing')
              AND source = 'stripe'
            ORDER BY email
        """, [e.lower() for e in S25_TEST_CANCEL_EMAILS])

        candidates = [
            {
                "id": r["id"],
                "email": r["email"],
                "source": r["source"],
                "status": r["status"],
                "sub_id": r["stripe_subscription_id"],
                "plan_amount": float(r["plan_amount"]) if r["plan_amount"] is not None else None,
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            }
            for r in rows
        ]

        total_mrr_removed = sum(c["plan_amount"] or 0 for c in candidates if c["status"] == "active")

        if not apply:
            return {
                "mode": "preview",
                "candidate_count": len(candidates),
                "expected_count": len(S25_TEST_CANCEL_EMAILS),
                "mrr_to_remove": round(total_mrr_removed, 2),
                "candidates": candidates,
                "next_step": 'POST again with {"apply": true} to commit',
            }

        from datetime import datetime as _dt
        batch_tag = f"s25_test_cancel_{_dt.utcnow().strftime('%Y%m%d_%H%M%S')}"
        cancelled_ids = []
        for c in candidates:
            await conn.execute(
                """UPDATE subscriptions
                   SET status = 'canceled',
                       canceled_at = NOW(),
                       updated_at = NOW(),
                       import_batch = COALESCE(import_batch, '') || ' ' || $1
                   WHERE id = $2 AND status IN ('active', 'trialing')""",
                batch_tag, c["id"]
            )
            cancelled_ids.append(c["id"])

        return {
            "mode": "applied",
            "batch_tag": batch_tag,
            "cancelled_count": len(cancelled_ids),
            "cancelled_ids": cancelled_ids,
            "mrr_removed": round(total_mrr_removed, 2),
            "candidates": candidates,
            "reversal_hint": f"To revert: UPDATE subscriptions SET status = 'active', canceled_at = NULL WHERE import_batch LIKE '%{batch_tag}%'",
        }


# --- Session 25: Fix mislabeled manual->apple/google records (broad pass) ---

@app.post("/api/admin/fix-manual-import-labels")
async def fix_manual_import_labels(request: Request):
    """S25 Step 2: Targeted fix for manual-labeled records whose ID prefix proves
    they came from Meg's Apple/Google imports. Sub_id prefix determines new source.

    Body: {"apply": false} for preview, {"apply": true} to commit.
    Scope: source='manual' AND status IN ('active','trialing')
           AND sub_id LIKE 'import_apple_%' OR 'meg_apple_%' OR 'import_google_%' OR 'meg_google_%'
    """
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)

    if not db_pool:
        return JSONResponse(status_code=500, content={"error": "No database connected"})

    try:
        body = await request.json()
    except Exception:
        body = {}
    apply = bool(body.get("apply", False))

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT id, email, source, status, stripe_subscription_id, created_at, import_batch
            FROM subscriptions
            WHERE source = 'manual'
              AND status IN ('active', 'trialing')
              AND (
                stripe_subscription_id LIKE 'import_apple_%' OR
                stripe_subscription_id LIKE 'meg_apple_%' OR
                stripe_subscription_id LIKE 'import_google_%' OR
                stripe_subscription_id LIKE 'meg_google_%'
              )
            ORDER BY email
        """)

        candidates = []
        for r in rows:
            sub_id = r["stripe_subscription_id"] or ""
            if "apple" in sub_id:
                new_source = "apple"
            elif "google" in sub_id:
                new_source = "google"
            else:
                new_source = "?"  # shouldn't hit, but defensive
            candidates.append({
                "id": r["id"],
                "email": r["email"],
                "current_source": r["source"],
                "new_source": new_source,
                "status": r["status"],
                "sub_id": sub_id,
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "import_batch": r["import_batch"],
            })

        if not apply:
            by_target = {}
            for c in candidates:
                by_target[c["new_source"]] = by_target.get(c["new_source"], 0) + 1
            return {
                "mode": "preview",
                "candidate_count": len(candidates),
                "by_new_source": by_target,
                "candidates": candidates,
                "next_step": 'POST again with {"apply": true} to commit',
            }

        from datetime import datetime as _dt
        batch_tag = f"s25_relabel_manual_to_import_{_dt.utcnow().strftime('%Y%m%d_%H%M%S')}"
        updated = 0
        for c in candidates:
            if c["new_source"] not in ("apple", "google"):
                continue
            await conn.execute(
                """UPDATE subscriptions
                   SET source = $1,
                       import_batch = COALESCE(import_batch, '') || ' ' || $2
                   WHERE id = $3 AND source = 'manual'""",
                c["new_source"], batch_tag, c["id"]
            )
            updated += 1

        return {
            "mode": "applied",
            "batch_tag": batch_tag,
            "updated_count": updated,
            "candidates": candidates,
        }


# --- Session 25: Fix mislabeled manual->apple records ---

@app.post("/api/admin/fix-manual-apple-labels")
async def fix_manual_apple_labels(request: Request):
    """S25: Targeted fix for the Bug 2 fallout from S24.
    Records where source='manual' but stripe_subscription_id starts with 'ymove_new_apple_'
    are misclassified — the ID prefix proves they're Apple. This relabels them.

    Body: {"apply": false} for preview, {"apply": true} to commit.
    Scope: source='manual' AND status IN ('active','trialing') AND sub_id LIKE 'ymove_new_apple_%'
    Effect: source='manual' -> source='apple'. Nothing else changes.
    """
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)

    if not db_pool:
        return JSONResponse(status_code=500, content={"error": "No database connected"})

    try:
        body = await request.json()
    except Exception:
        body = {}
    apply = bool(body.get("apply", False))

    async with db_pool.acquire() as conn:
        # Find candidates
        rows = await conn.fetch("""
            SELECT id, email, source, status, stripe_subscription_id, created_at, import_batch
            FROM subscriptions
            WHERE source = 'manual'
              AND status IN ('active', 'trialing')
              AND stripe_subscription_id LIKE 'ymove_new_apple_%'
            ORDER BY email
        """)

        candidates = [
            {
                "id": r["id"],
                "email": r["email"],
                "current_source": r["source"],
                "new_source": "apple",
                "status": r["status"],
                "sub_id": r["stripe_subscription_id"],
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "import_batch": r["import_batch"],
            }
            for r in rows
        ]

        if not apply:
            return {
                "mode": "preview",
                "candidate_count": len(candidates),
                "candidates": candidates,
                "next_step": 'POST again with {"apply": true} to commit',
            }

        # Apply: tag with batch for reversibility
        from datetime import datetime as _dt
        batch_tag = f"s25_relabel_manual_to_apple_{_dt.utcnow().strftime('%Y%m%d_%H%M%S')}"
        updated_ids = []
        for c in candidates:
            await conn.execute(
                """UPDATE subscriptions
                   SET source = 'apple',
                       import_batch = COALESCE(import_batch, '') || ' ' || $1
                   WHERE id = $2 AND source = 'manual'
                     AND stripe_subscription_id LIKE 'ymove_new_apple_%'""",
                batch_tag, c["id"]
            )
            updated_ids.append(c["id"])

        return {
            "mode": "applied",
            "batch_tag": batch_tag,
            "updated_count": len(updated_ids),
            "updated_ids": updated_ids,
            "candidates": candidates,
        }


# --- Session 23: Deep Reconciliation Audit ---

@app.get("/api/admin/reconciliation-audit")
async def reconciliation_audit(request: Request):
    """S23: Deep audit of all batch operations, duplicates, and data integrity.
    Answers: What manual operations have been done? Are there duplicates? What are we missing?"""
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    report = {"batch_history": {}, "duplicates": {}, "source_accuracy": {}, "gaps": {}, "action_items": []}

    async with db_pool.acquire() as conn:

        # ============================================================
        # SECTION 1: Every batch operation ever performed
        # ============================================================
        batches = await conn.fetch("""
            SELECT import_batch, status, source, COUNT(*) as cnt,
                   MIN(created_at) as earliest, MAX(updated_at) as latest_update
            FROM subscriptions
            WHERE import_batch IS NOT NULL AND import_batch != ''
            GROUP BY import_batch, status, source
            ORDER BY latest_update DESC
        """)
        batch_summary = {}
        for b in batches:
            key = b["import_batch"]
            if key not in batch_summary:
                batch_summary[key] = {"records": [], "total_count": 0}
            batch_summary[key]["records"].append({
                "status": b["status"],
                "source": b["source"],
                "count": b["cnt"],
                "earliest": str(b["earliest"]),
                "latest_update": str(b["latest_update"])
            })
            batch_summary[key]["total_count"] += b["cnt"]
        report["batch_history"] = batch_summary

        # ============================================================
        # SECTION 2: Subscription ID prefix analysis
        # Shows exactly where every record came from
        # ============================================================
        id_patterns = await conn.fetch("""
            SELECT
                CASE
                    WHEN stripe_subscription_id LIKE 'sub_%' THEN 'sub_* (Stripe direct)'
                    WHEN stripe_subscription_id LIKE 'import_apple_%' THEN 'import_apple_* (Meg import)'
                    WHEN stripe_subscription_id LIKE 'import_google_%' THEN 'import_google_* (Meg import)'
                    WHEN stripe_subscription_id LIKE 'meg_apple_%' THEN 'meg_apple_* (Meg import v2)'
                    WHEN stripe_subscription_id LIKE 'meg_google_%' THEN 'meg_google_* (Meg import v2)'
                    WHEN stripe_subscription_id LIKE 'ymove_new_%' THEN 'ymove_new_* (Shadow sync auto-import)'
                    WHEN stripe_subscription_id LIKE 'ymove_switch_%' THEN 'ymove_switch_* (Cross-platform switch)'
                    WHEN stripe_subscription_id LIKE 'ym_google_%' THEN 'ym_google_* (ymove webhook Google)'
                    WHEN stripe_subscription_id ~ '^[0-9]+$' THEN 'numeric (Apple transaction ID)'
                    ELSE 'other: ' || LEFT(stripe_subscription_id, 20)
                END as id_pattern,
                source, status, COUNT(*) as cnt
            FROM subscriptions
            GROUP BY id_pattern, source, status
            ORDER BY id_pattern, cnt DESC
        """)
        report["id_patterns"] = [{"pattern": r["id_pattern"], "source": r["source"], "status": r["status"], "count": r["cnt"]} for r in id_patterns]

        # ============================================================
        # SECTION 3: Duplicate active emails (detailed)
        # ============================================================
        dup_details = await conn.fetch("""
            SELECT s.id, lower(s.email) as email, s.stripe_subscription_id, s.source, s.status,
                   s.plan_interval, s.plan_amount, s.import_batch, s.created_at, s.updated_at
            FROM subscriptions s
            INNER JOIN (
                SELECT lower(email) as em FROM subscriptions
                WHERE status IN ('active', 'trialing') AND email != '' AND email IS NOT NULL
                GROUP BY lower(email) HAVING COUNT(*) > 1
            ) dups ON lower(s.email) = dups.em
            WHERE s.status IN ('active', 'trialing')
            ORDER BY lower(s.email), s.created_at
        """)
        dup_grouped = {}
        for d in dup_details:
            em = d["email"]
            if em not in dup_grouped:
                dup_grouped[em] = []
            dup_grouped[em].append({
                "id": d["id"],
                "sub_id": d["stripe_subscription_id"],
                "source": d["source"],
                "status": d["status"],
                "plan": f"{d['plan_interval']}/${d['plan_amount']}",
                "import_batch": d["import_batch"],
                "created": str(d["created_at"]),
                "updated": str(d["updated_at"])
            })
        report["duplicates"] = {
            "count": len(dup_grouped),
            "emails": dup_grouped
        }
        if len(dup_grouped) > 0:
            report["action_items"].append(f"CRITICAL: {len(dup_grouped)} emails have multiple active subs. Each inflates MRR. Review and cancel duplicates.")

        # ============================================================
        # SECTION 4: Source vs ID pattern mismatches
        # (e.g., source='apple' but ID is 'ymove_new_undetermined_*')
        # ============================================================
        mismatches = await conn.fetch("""
            SELECT id, email, stripe_subscription_id, source, status, import_batch
            FROM subscriptions
            WHERE status IN ('active', 'trialing')
            AND (
                (source = 'apple' AND stripe_subscription_id LIKE 'ymove_new_google_%')
                OR (source = 'google' AND stripe_subscription_id LIKE 'ymove_new_apple_%')
                OR (source = 'apple' AND stripe_subscription_id LIKE 'ym_google_%')
                OR (source = 'google' AND stripe_subscription_id LIKE '%apple%')
                OR (source = 'stripe' AND stripe_subscription_id NOT LIKE 'sub_%')
                OR (source NOT IN ('stripe', 'apple', 'google', 'undetermined'))
            )
        """)
        report["source_accuracy"] = {
            "mismatched_records": len(mismatches),
            "details": [{"id": r["id"], "email": r["email"], "sub_id": r["stripe_subscription_id"],
                         "source": r["source"], "import_batch": r["import_batch"]} for r in mismatches[:30]]
        }
        if len(mismatches) > 0:
            report["action_items"].append(f"WARNING: {len(mismatches)} active subs have source/ID pattern mismatches. Provider may be wrong.")

        # ============================================================
        # SECTION 5: Records that came from shadow sync vs webhooks vs manual
        # ============================================================
        origin_counts = await conn.fetch("""
            SELECT
                CASE
                    WHEN stripe_subscription_id LIKE 'sub_%' THEN 'stripe_webhook'
                    WHEN stripe_subscription_id ~ '^[0-9]+$' THEN 'ymove_webhook_apple'
                    WHEN stripe_subscription_id LIKE 'ym_google_%' THEN 'ymove_webhook_google'
                    WHEN stripe_subscription_id LIKE 'ymove_new_%' THEN 'shadow_sync_import'
                    WHEN stripe_subscription_id LIKE 'ymove_switch_%' THEN 'shadow_sync_switch'
                    WHEN stripe_subscription_id LIKE 'import_%' OR stripe_subscription_id LIKE 'meg_%' THEN 'manual_import'
                    ELSE 'unknown'
                END as origin,
                status,
                COUNT(*) as cnt
            FROM subscriptions
            GROUP BY origin, status
            ORDER BY origin, status
        """)
        report["data_origins"] = [{"origin": r["origin"], "status": r["status"], "count": r["cnt"]} for r in origin_counts]

        # ============================================================
        # SECTION 6: Undetermined source subs (the bug we just fixed)
        # ============================================================
        undetermined = await conn.fetch("""
            SELECT id, email, stripe_subscription_id, status, import_batch, created_at
            FROM subscriptions
            WHERE source = 'undetermined'
            ORDER BY created_at DESC
            LIMIT 50
        """)
        # Also check for subs that SHOULD be undetermined (ymove_new_apple but no real apple transaction)
        suspicious_apple = await conn.fetch("""
            SELECT id, email, stripe_subscription_id, source, status, import_batch
            FROM subscriptions
            WHERE source = 'apple'
            AND stripe_subscription_id LIKE 'ymove_new_apple_%'
            AND status IN ('active', 'trialing')
        """)
        report["gaps"] = {
            "undetermined_subs": [{"id": r["id"], "email": r["email"], "sub_id": r["stripe_subscription_id"],
                                   "status": r["status"], "batch": r["import_batch"]} for r in undetermined],
            "suspicious_apple_from_sync": {
                "count": len(suspicious_apple),
                "note": "These were imported by shadow sync with source=apple, but ymove API returns null provider. They may actually be Google or Manual users.",
                "records": [{"id": r["id"], "email": r["email"], "sub_id": r["stripe_subscription_id"],
                            "batch": r["import_batch"]} for r in suspicious_apple[:30]]
            }
        }
        if len(suspicious_apple) > 0:
            report["action_items"].append(f"WARNING: {len(suspicious_apple)} active subs were auto-imported as 'apple' by shadow sync but provider was actually null. These need provider correction once Tosh fixes the API.")

        # ============================================================
        # SECTION 7: Recent ymove sync run results
        # ============================================================
        try:
            recent_syncs = await conn.fetch("""
                SELECT id, phase, started_at, completed_at, our_active_count, ymove_active_count,
                       deactivated, reactivated, new_imported, errors
                FROM ymove_sync_runs
                ORDER BY id DESC LIMIT 10
            """)
            report["recent_syncs"] = [
                {
                    "run_id": r["id"], "phase": r["phase"],
                    "started": str(r["started_at"]), "completed": str(r["completed_at"]) if r["completed_at"] else None,
                    "our_active": r["our_active_count"], "ymove_active": r["ymove_active_count"],
                    "deactivated": r["deactivated"], "reactivated": r["reactivated"],
                    "new_imported": r["new_imported"], "errors": r["errors"]
                } for r in recent_syncs
            ]
        except Exception:
            report["recent_syncs"] = "ymove_sync_runs table not found or empty"

        # ============================================================
        # SECTION 8: Active subs with NO email (can never be reconciled)
        # ============================================================
        no_email_active = await conn.fetch("""
            SELECT id, stripe_subscription_id, source, status, plan_amount, created_at, import_batch
            FROM subscriptions
            WHERE status IN ('active', 'trialing')
            AND (email IS NULL OR email = '')
            ORDER BY created_at DESC
        """)
        report["no_email_active"] = {
            "count": len(no_email_active),
            "records": [{"id": r["id"], "sub_id": r["stripe_subscription_id"], "source": r["source"],
                         "plan_amount": r["plan_amount"], "batch": r["import_batch"]} for r in no_email_active[:30]]
        }
        if len(no_email_active) > 0:
            report["action_items"].append(f"CRITICAL: {len(no_email_active)} active subs have NO email. Cannot reconcile against ymove, cannot detect cancellations, inflate counts silently.")

        # ============================================================
        # SUMMARY
        # ============================================================
        total_active = await conn.fetchval("SELECT COUNT(*) FROM subscriptions WHERE status IN ('active', 'trialing')")
        report["summary"] = {
            "total_active_subs": total_active,
            "total_batch_operations": len(batch_summary),
            "total_duplicate_emails": len(dup_grouped),
            "total_source_mismatches": len(mismatches),
            "total_no_email_active": len(no_email_active),
            "total_suspicious_apple_sync": len(suspicious_apple),
            "total_action_items": len(report["action_items"])
        }

    return report


# --- Session 16: Database Health Check ---

@app.get("/api/admin/db-check")
async def db_check(request: Request):
    """Verify all tables, columns, and indexes exist after startup refactor."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    checks = {"tables": {}, "indexes": {}, "columns": {}, "all_ok": True}

    async with db_pool.acquire() as conn:
        expected_tables = ["leads", "page_views", "chat_sessions", "subscriptions",
                          "subscription_events", "ad_spend", "platform_metrics", "ymove_webhook_log"]
        for t in expected_tables:
            exists = await conn.fetchval(
                "SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_name = $1)", t
            )
            checks["tables"][t] = exists
            if not exists:
                checks["all_ok"] = False

        critical_cols = {
            "subscriptions": ["converted_at", "readable_id", "renewal_count", "last_renewed_at", "source", "import_batch"],
            "leads": ["utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content", "ym_source"],
            "subscription_events": ["source"],
        }
        for table, cols in critical_cols.items():
            checks["columns"][table] = {}
            for col in cols:
                exists = await conn.fetchval(
                    "SELECT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_name = $1 AND column_name = $2)",
                    table, col
                )
                checks["columns"][table][col] = exists
                if not exists:
                    checks["all_ok"] = False

        expected_indexes = ["idx_leads_email_lower", "idx_subs_email_lower", "idx_subs_trial_start", "idx_subs_import_batch"]
        for idx in expected_indexes:
            exists = await conn.fetchval(
                "SELECT EXISTS (SELECT 1 FROM pg_indexes WHERE indexname = $1)", idx
            )
            checks["indexes"][idx] = exists
            if not exists:
                checks["all_ok"] = False

        checks["row_counts"] = {}
        for t in expected_tables:
            if checks["tables"].get(t):
                count = await conn.fetchval(f"SELECT COUNT(*) FROM {t}")
                checks["row_counts"][t] = count

    return checks


# --- Session 16: Safe Excel/CSV Import with Batch Tracking ---

@app.post("/api/admin/import-preview")
async def import_preview(request: Request, file: UploadFile = File(...)):
    """Dry-run import: reads Excel/CSV, shows what would be imported vs skipped."""

    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    contents = await file.read()
    filename = file.filename or ""

    rows_to_check = []
    if filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls"):
        wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            for ri, row in enumerate(ws.iter_rows(values_only=True)):
                if ri == 0:
                    headers = [str(h or "").strip().lower() for h in row]
                    continue
                if not any(row):
                    continue
                row_dict = {}
                for ci, h in enumerate(headers):
                    row_dict[h] = str(row[ci] or "").strip() if ci < len(row) else ""
                row_dict["_sheet"] = sheet_name
                rows_to_check.append(row_dict)
        wb.close()
    else:
        text = contents.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            normalized = {k.strip().lower(): (v or "").strip() for k, v in row.items()}
            normalized["_sheet"] = "csv"
            rows_to_check.append(normalized)



    async with db_pool.acquire() as conn:
        existing_emails = await conn.fetch(
            "SELECT DISTINCT lower(email) as em FROM subscriptions WHERE email != ''"
        )
        existing_sub_ids = await conn.fetch("SELECT stripe_subscription_id FROM subscriptions")
        total_before = await conn.fetchval("SELECT COUNT(*) FROM subscriptions")

    existing_email_set = set(r["em"] for r in existing_emails)
    existing_id_set = set(r["stripe_subscription_id"] for r in existing_sub_ids)

    would_import = []
    would_skip_duplicate = []
    would_skip_no_email = []
    by_sheet = {}

    for row in rows_to_check:
        email = _get_email(row)
        source = _get_source(row)
        sheet = row.get("_sheet", "unknown")
        if sheet not in by_sheet:
            by_sheet[sheet] = {"total": 0, "import": 0, "skip_dup": 0, "skip_no_email": 0}
        by_sheet[sheet]["total"] += 1
        if not email:
            would_skip_no_email.append({"sheet": sheet})
            by_sheet[sheet]["skip_no_email"] += 1
            continue
        email_hash = hashlib.md5(email.encode()).hexdigest()[:16]
        syn_id = f"import_{source}_{email_hash}"
        if email in existing_email_set or syn_id in existing_id_set:
            would_skip_duplicate.append({"email": email, "source": source, "sheet": sheet})
            by_sheet[sheet]["skip_dup"] += 1
        else:
            would_import.append({"email": email, "source": source, "sheet": sheet})
            by_sheet[sheet]["import"] += 1
            existing_email_set.add(email)
            existing_id_set.add(syn_id)

    return {
        "status": "preview",
        "filename": filename,
        "total_rows_parsed": len(rows_to_check),
        "would_import": len(would_import),
        "would_skip_duplicate": len(would_skip_duplicate),
        "would_skip_no_email": len(would_skip_no_email),
        "current_db_subscriptions": total_before,
        "after_import_estimate": total_before + len(would_import),
        "by_sheet": by_sheet,
        "sample_imports": would_import[:20],
        "sample_duplicates": would_skip_duplicate[:20],
    }


@app.post("/api/admin/import-batch")
async def import_batch(request: Request, file: UploadFile = File(...)):
    """Import subscribers with batch ID for safe revert. Use /api/admin/revert-batch to undo."""

    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    contents = await file.read()
    filename = file.filename or ""
    batch_id = f"batch_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

    rows_to_import = []
    if filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls"):
        wb = load_workbook(BytesIO(contents), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            for ri, row in enumerate(ws.iter_rows(values_only=True)):
                if ri == 0:
                    headers = [str(h or "").strip().lower() for h in row]
                    continue
                if not any(row):
                    continue
                row_dict = {}
                for ci, h in enumerate(headers):
                    row_dict[h] = str(row[ci] or "").strip() if ci < len(row) else ""
                rows_to_import.append(row_dict)
        wb.close()
    else:
        text = contents.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            normalized = {k.strip().lower(): (v or "").strip() for k, v in row.items()}
            rows_to_import.append(normalized)




    async with db_pool.acquire() as conn:
        existing_emails = await conn.fetch(
            "SELECT DISTINCT lower(email) as em FROM subscriptions WHERE email != ''"
        )
        existing_ids = await conn.fetch("SELECT stripe_subscription_id FROM subscriptions")
        existing_email_set = set(r["em"] for r in existing_emails)
        existing_id_set = set(r["stripe_subscription_id"] for r in existing_ids)
        count_before = await conn.fetchval("SELECT COUNT(*) FROM subscriptions")

        imported = 0
        skipped = 0
        errors = 0

        for row in rows_to_import:
            email = _get_email(row)
            if not email:
                skipped += 1
                continue
            source = _get_source(row, "apple")
            email_hash = hashlib.md5(email.encode()).hexdigest()[:16]
            syn_id = f"import_{source}_{email_hash}"
            if email in existing_email_set or syn_id in existing_id_set:
                skipped += 1
                continue
            period_start = _get_date(row)
            plan_amount = 1999
            plan_interval = "month"
            try:
                await conn.execute("""
                    INSERT INTO subscriptions (
                        stripe_customer_id, stripe_subscription_id, email, status,
                        plan_interval, plan_amount, currency, source,
                        current_period_start, created_at, updated_at, import_batch
                    ) VALUES ('', $1, $2, 'active', $3, $4, 'usd', $5, $6, COALESCE($7, NOW()), NOW(), $8)
                    ON CONFLICT (stripe_subscription_id) DO NOTHING
                """,
                    syn_id, email, plan_interval, plan_amount, source,
                    period_start, period_start, batch_id
                )
                imported += 1
                existing_email_set.add(email)
                existing_id_set.add(syn_id)
            except Exception as e:
                print(f"[Import] Error: {e}")
                errors += 1

        count_after = await conn.fetchval("SELECT COUNT(*) FROM subscriptions")

    return {
        "status": "ok",
        "batch_id": batch_id,
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "count_before": count_before,
        "count_after": count_after,
    }


@app.post("/api/admin/revert-batch")
async def revert_batch(request: Request):
    """Revert a batch import by deleting all records with matching import_batch tag."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    body = await request.json()
    batch_id = (body.get("batch_id") or "").strip()
    if not batch_id:
        raise HTTPException(status_code=400, detail="batch_id required")

    async with db_pool.acquire() as conn:
        count = await conn.fetchval(
            "SELECT COUNT(*) FROM subscriptions WHERE import_batch = $1", batch_id
        )
        if count == 0:
            return {"status": "ok", "deleted": 0, "message": f"No records found with batch_id '{batch_id}'"}

        confirm = body.get("confirm", False)
        if not confirm:
            sample = await conn.fetch(
                "SELECT email, source, created_at FROM subscriptions WHERE import_batch = $1 LIMIT 10",
                batch_id
            )
            return {
                "status": "confirmation_required",
                "batch_id": batch_id,
                "records_to_delete": count,
                "sample": [{"email": r["email"], "source": r["source"], "created_at": str(r["created_at"])} for r in sample],
                "message": f"This will delete {count} records. Send again with confirm: true to proceed."
            }

        result = await conn.execute(
            "DELETE FROM subscriptions WHERE import_batch = $1", batch_id
        )
        deleted = int(result.split(" ")[-1]) if result else 0

    return {"status": "ok", "batch_id": batch_id, "deleted": deleted}


@app.get("/api/admin/list-batches")
async def list_batches(request: Request):
    """List all import batches for the revert UI."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    async with db_pool.acquire() as conn:
        batches = await conn.fetch("""
            SELECT import_batch, COUNT(*) as record_count,
                   MIN(created_at) as earliest, MAX(created_at) as latest
            FROM subscriptions
            WHERE import_batch IS NOT NULL
            GROUP BY import_batch
            ORDER BY MIN(created_at) DESC
        """)

    return {"batches": [
        {"batch_id": r["import_batch"], "records": r["record_count"],
         "earliest": str(r["earliest"]), "latest": str(r["latest"])}
        for r in batches
    ]}


# --- Apple App Store Connect Integration (Session 13) ---


def generate_apple_jwt() -> str:
    """Generate a short-lived JWT for App Store Connect API."""
    if not all([APPLE_KEY_ID, APPLE_ISSUER_ID, APPLE_KEY_CONTENT]):
        raise ValueError("Apple API credentials not configured (APPLE_KEY_ID, APPLE_ISSUER_ID, APPLE_KEY_CONTENT)")
    now = int(time.time())
    payload = {
        "iss": APPLE_ISSUER_ID,
        "iat": now,
        "exp": now + (20 * 60),
        "aud": "appstoreconnect-v1",
    }
    headers = {"alg": "ES256", "kid": APPLE_KEY_ID, "typ": "JWT"}
    return pyjwt.encode(payload, APPLE_KEY_CONTENT, algorithm="ES256", headers=headers)


async def apple_api_get(path: str, params: dict = None, expect_binary: bool = False):
    """Authenticated GET to App Store Connect API."""
    token = generate_apple_jwt()
    url = f"https://api.appstoreconnect.apple.com{path}"
    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.get(url, headers={"Authorization": f"Bearer {token}"}, params=params or {})
        print(f"[Apple API] GET {path} -> {resp.status_code}")
        if expect_binary and resp.status_code == 200:
            return resp.content
        if resp.status_code != 200:
            try:
                error_data = resp.json()
            except Exception:
                error_data = {"raw": resp.text[:500]}
            print(f"[Apple API] Error: {json.dumps(error_data, indent=2)[:500]}")
            return {"error": True, "status": resp.status_code, "data": error_data}
        return resp.json()


def parse_apple_tsv(gzipped_content: bytes) -> list:
    """Decompress gzipped TSV from Apple and return list of row dicts."""
    try:
        decompressed = gzip.decompress(gzipped_content)
        text = decompressed.decode("utf-8").strip()
    except Exception:
        text = gzipped_content.decode("utf-8").strip()
    lines = text.split("\n")
    if not lines:
        return []
    headers = lines[0].split("\t")
    rows = []
    for line in lines[1:]:
        if not line.strip():
            continue
        values = line.split("\t")
        row = {}
        for i, h in enumerate(headers):
            row[h.strip()] = values[i].strip() if i < len(values) else ""
        rows.append(row)
    return rows


def aggregate_apple_subscription_rows(rows: list) -> dict:
    """Sum active subscriptions and free trials across all products/countries."""
    total_active = 0
    total_trials = 0
    for row in rows:
        try:
            total_active += int(row.get("Active Subscriptions",
                row.get("Active Standard Price Subscriptions", "0")) or 0)
            total_trials += int(row.get("Active Free Trial Introductory Offer Subscriptions",
                row.get("Active Free Trials", "0")) or 0)
        except (ValueError, TypeError):
            pass
    return {"active_subscriptions": total_active, "active_free_trials": total_trials}


def aggregate_apple_event_rows(rows: list) -> dict:
    """Categorize subscription events and sum revenue."""
    new_subs = renewals = conversions = cancellations = reactivations = 0
    revenue_cents = proceeds_cents = 0
    for row in rows:
        event = row.get("Event", "").lower()
        try:
            units = int(row.get("Quantity", row.get("Units", "0")) or 0)
        except (ValueError, TypeError):
            units = 0
        try:
            dev_proceeds = float(row.get("Developer Proceeds", row.get("Proceeds", "0")) or 0)
            cust_price = float(row.get("Customer Price", "0") or 0)
        except (ValueError, TypeError):
            dev_proceeds = cust_price = 0
        if any(x in event for x in ["new", "subscribe", "initial"]):
            new_subs += units
        if "renew" in event:
            renewals += units
        if any(x in event for x in ["convert", "paid from"]):
            conversions += units
        if any(x in event for x in ["cancel", "churn", "refund"]):
            cancellations += units
        if "reactivat" in event:
            reactivations += units
        revenue_cents += round(cust_price * 100 * units) if cust_price else 0
        proceeds_cents += round(dev_proceeds * 100 * units) if dev_proceeds else 0
    return {
        "new_subscriptions": new_subs, "renewals": renewals,
        "conversions": conversions, "cancellations": cancellations,
        "reactivations": reactivations,
        "revenue_cents": revenue_cents, "proceeds_cents": proceeds_cents,
    }


async def store_apple_subscription_metric(conn, report_date, totals: dict, raw_rows: list):
    """Upsert subscription snapshot into platform_metrics."""
    await conn.execute("""
        INSERT INTO platform_metrics (date, source, metric_type,
            active_subscriptions, active_free_trials, report_data, updated_at)
        VALUES ($1, 'apple', 'subscription', $2, $3, $4, NOW())
        ON CONFLICT (date, source, metric_type) DO UPDATE SET
            active_subscriptions = EXCLUDED.active_subscriptions,
            active_free_trials = EXCLUDED.active_free_trials,
            report_data = EXCLUDED.report_data,
            updated_at = NOW()
    """, report_date, totals["active_subscriptions"], totals["active_free_trials"],
        json.dumps(raw_rows[:50]))


async def store_apple_event_metric(conn, report_date, totals: dict, raw_rows: list):
    """Upsert event data into platform_metrics."""
    await conn.execute("""
        INSERT INTO platform_metrics (date, source, metric_type,
            new_subscriptions, renewals, conversions, cancellations,
            reactivations, revenue_cents, proceeds_cents, report_data, updated_at)
        VALUES ($1, 'apple', 'subscription_event', $2, $3, $4, $5, $6, $7, $8, $9, NOW())
        ON CONFLICT (date, source, metric_type) DO UPDATE SET
            new_subscriptions = EXCLUDED.new_subscriptions,
            renewals = EXCLUDED.renewals,
            conversions = EXCLUDED.conversions,
            cancellations = EXCLUDED.cancellations,
            reactivations = EXCLUDED.reactivations,
            revenue_cents = EXCLUDED.revenue_cents,
            proceeds_cents = EXCLUDED.proceeds_cents,
            report_data = EXCLUDED.report_data,
            updated_at = NOW()
    """, report_date, totals["new_subscriptions"], totals["renewals"],
        totals["conversions"], totals["cancellations"], totals["reactivations"],
        totals["revenue_cents"], totals["proceeds_cents"],
        json.dumps(raw_rows[:50]))


@app.get("/api/admin/apple-discover")
async def apple_discover(request: Request):
    """Discover Apple app details: ID, bundle, subscription groups, test vendor."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)

    result = {"steps": []}

    # Step 1: List apps
    apps_resp = await apple_api_get("/v1/apps")
    if isinstance(apps_resp, dict) and apps_resp.get("error"):
        return {"error": "Failed to list apps", "detail": apps_resp}

    apps = apps_resp.get("data", [])
    result["apps"] = []
    for a in apps:
        attrs = a.get("attributes", {})
        app_info = {
            "id": a.get("id"),
            "name": attrs.get("name"),
            "bundleId": attrs.get("bundleId"),
            "sku": attrs.get("sku"),
        }
        result["apps"].append(app_info)

        # Step 2: Get subscription groups
        groups_resp = await apple_api_get(f"/v1/apps/{a['id']}/subscriptionGroups")
        if isinstance(groups_resp, dict) and not groups_resp.get("error"):
            groups = groups_resp.get("data", [])
            app_info["subscription_groups"] = []
            for g in groups:
                group_info = {
                    "id": g.get("id"),
                    "name": g.get("attributes", {}).get("referenceName"),
                }
                subs_resp = await apple_api_get(f"/v1/subscriptionGroups/{g['id']}/subscriptions")
                if isinstance(subs_resp, dict) and not subs_resp.get("error"):
                    group_info["subscriptions"] = [
                        {"id": s.get("id"), "name": s.get("attributes", {}).get("name"),
                         "productId": s.get("attributes", {}).get("productId"),
                         "state": s.get("attributes", {}).get("state")}
                        for s in subs_resp.get("data", [])
                    ]
                app_info["subscription_groups"].append(group_info)

    result["steps"].append("Listed apps and subscription groups")

    # Step 3: Test vendor number if configured
    if APPLE_VENDOR_NUMBER:
        yesterday = (datetime.now(timezone.utc) - timedelta(days=2)).strftime("%Y-%m-%d")
        test_resp = await apple_api_get("/v1/salesReports", params={
            "filter[reportType]": "SALES",
            "filter[reportSubType]": "SUMMARY",
            "filter[frequency]": "DAILY",
            "filter[reportDate]": yesterday,
            "filter[vendorNumber]": APPLE_VENDOR_NUMBER,
        }, expect_binary=True)
        if isinstance(test_resp, dict) and test_resp.get("error"):
            result["vendor_test"] = {"status": "failed", "detail": test_resp}
        else:
            result["vendor_test"] = {"status": "success", "vendor": APPLE_VENDOR_NUMBER, "report_size_bytes": len(test_resp)}
        result["steps"].append(f"Tested vendor number: {APPLE_VENDOR_NUMBER}")
    else:
        result["vendor_number_help"] = "NOT SET. Set APPLE_VENDOR_NUMBER env var. Find it in App Store Connect > Payments and Financial Reports > top left gray bar."
        result["steps"].append("No vendor number configured")

    return result


@app.post("/api/admin/apple-pull-report")
async def apple_pull_report(request: Request):
    """Pull Apple subscription reports for a date and store in platform_metrics.
    Body: {"date": "2026-03-10"} â defaults to 2 days ago (Apple ~2 day lag)."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)

    if not APPLE_VENDOR_NUMBER:
        raise HTTPException(status_code=400, detail="APPLE_VENDOR_NUMBER not configured. Run /api/admin/apple-discover first.")

    body = await request.json()
    report_date_str = body.get("date", "")
    if not report_date_str:
        report_date_str = (datetime.now(timezone.utc) - timedelta(days=2)).strftime("%Y-%m-%d")

    report_date = datetime.strptime(report_date_str, "%Y-%m-%d").date()
    results = {"date": report_date_str, "reports": {}}

    # --- SUBSCRIPTION report (active counts snapshot) ---
    print(f"[Apple] Pulling SUBSCRIPTION report for {report_date_str}")
    sub_resp = await apple_api_get("/v1/salesReports", params={
        "filter[reportType]": "SUBSCRIPTION",
        "filter[reportSubType]": "SUMMARY",
        "filter[frequency]": "DAILY",
        "filter[reportDate]": report_date_str,
        "filter[vendorNumber]": APPLE_VENDOR_NUMBER,
    }, expect_binary=True)

    if isinstance(sub_resp, dict) and sub_resp.get("error"):
        results["reports"]["subscription"] = {"error": sub_resp}
    elif isinstance(sub_resp, bytes):
        rows = parse_apple_tsv(sub_resp)
        totals = aggregate_apple_subscription_rows(rows)
        results["reports"]["subscription"] = {"rows": len(rows), "totals": totals, "sample": rows[:3]}
        if db_pool:
            try:
                async with db_pool.acquire() as conn:
                    await store_apple_subscription_metric(conn, report_date, totals, rows)
                results["reports"]["subscription"]["stored"] = True
            except Exception as e:
                results["reports"]["subscription"]["store_error"] = str(e)

    # --- SUBSCRIPTION_EVENT report (daily activity) ---
    print(f"[Apple] Pulling SUBSCRIPTION_EVENT report for {report_date_str}")
    event_resp = await apple_api_get("/v1/salesReports", params={
        "filter[reportType]": "SUBSCRIPTION_EVENT",
        "filter[reportSubType]": "SUMMARY",
        "filter[frequency]": "DAILY",
        "filter[reportDate]": report_date_str,
        "filter[vendorNumber]": APPLE_VENDOR_NUMBER,
    }, expect_binary=True)

    if isinstance(event_resp, dict) and event_resp.get("error"):
        results["reports"]["subscription_event"] = {"error": event_resp}
    elif isinstance(event_resp, bytes):
        rows = parse_apple_tsv(event_resp)
        totals = aggregate_apple_event_rows(rows)
        results["reports"]["subscription_event"] = {"rows": len(rows), "totals": totals, "sample": rows[:5]}
        if db_pool:
            try:
                async with db_pool.acquire() as conn:
                    await store_apple_event_metric(conn, report_date, totals, rows)
                results["reports"]["subscription_event"]["stored"] = True
            except Exception as e:
                results["reports"]["subscription_event"]["store_error"] = str(e)

    return results


@app.post("/api/admin/apple-backfill")
async def apple_backfill(request: Request):
    """Pull Apple reports for a date range. Max 90 days per run.
    Body: {"start_date": "2026-01-01", "end_date": "2026-03-10"}"""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)

    if not APPLE_VENDOR_NUMBER:
        raise HTTPException(status_code=400, detail="APPLE_VENDOR_NUMBER not configured")

    body = await request.json()
    start = body.get("start_date", "")
    end = body.get("end_date", "")
    if not start or not end:
        raise HTTPException(status_code=400, detail="start_date and end_date required (YYYY-MM-DD)")

    start_dt = datetime.strptime(start, "%Y-%m-%d")
    end_dt = datetime.strptime(end, "%Y-%m-%d")
    if (end_dt - start_dt).days > 90:
        raise HTTPException(status_code=400, detail="Max 90 days per backfill run")

    results = {"days_processed": 0, "errors": [], "success": []}
    current = start_dt
    while current <= end_dt:
        date_str = current.strftime("%Y-%m-%d")
        rd = current.date()

        for report_type, metric_type in [("SUBSCRIPTION", "subscription"), ("SUBSCRIPTION_EVENT", "subscription_event")]:
            try:
                resp = await apple_api_get("/v1/salesReports", params={
                    "filter[reportType]": report_type,
                    "filter[reportSubType]": "SUMMARY",
                    "filter[frequency]": "DAILY",
                    "filter[reportDate]": date_str,
                    "filter[vendorNumber]": APPLE_VENDOR_NUMBER,
                }, expect_binary=True)

                if isinstance(resp, dict) and resp.get("error"):
                    status = resp.get("status", 0)
                    if status != 404:
                        results["errors"].append(f"{date_str} {report_type}: {status}")
                elif isinstance(resp, bytes) and db_pool:
                    rows = parse_apple_tsv(resp)
                    async with db_pool.acquire() as conn:
                        if metric_type == "subscription":
                            totals = aggregate_apple_subscription_rows(rows)
                            await store_apple_subscription_metric(conn, rd, totals, rows)
                        else:
                            totals = aggregate_apple_event_rows(rows)
                            await store_apple_event_metric(conn, rd, totals, rows)
            except Exception as e:
                results["errors"].append(f"{date_str} {report_type}: {str(e)}")

        results["days_processed"] += 1
        results["success"].append(date_str)
        current += timedelta(days=1)

    return results


@app.get("/api/admin/apple-metrics")
async def apple_metrics(request: Request):
    """Get aggregated Apple metrics for dashboard. Optional date range."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database")

    date_from = request.query_params.get("from", "")
    date_to = request.query_params.get("to", "")

    async with db_pool.acquire() as conn:
        # Latest subscription snapshot
        latest_sub = await conn.fetchrow("""
            SELECT * FROM platform_metrics
            WHERE source = 'apple' AND metric_type = 'subscription'
            ORDER BY date DESC LIMIT 1
        """)

        # Event totals for date range or last 30 days
        if date_from and date_to:
            events = await conn.fetch("""
                SELECT * FROM platform_metrics
                WHERE source = 'apple' AND metric_type = 'subscription_event'
                AND date >= $1 AND date <= $2 ORDER BY date
            """, datetime.strptime(date_from, "%Y-%m-%d").date(),
                datetime.strptime(date_to, "%Y-%m-%d").date())
        else:
            events = await conn.fetch("""
                SELECT * FROM platform_metrics
                WHERE source = 'apple' AND metric_type = 'subscription_event'
                AND date > CURRENT_DATE - INTERVAL '30 days' ORDER BY date
            """)

        total_new = total_renewals = total_conversions = total_cancellations = 0
        total_revenue = total_proceeds = 0
        daily_data = []
        for e in events:
            total_new += e["new_subscriptions"] or 0
            total_renewals += e["renewals"] or 0
            total_conversions += e["conversions"] or 0
            total_cancellations += e["cancellations"] or 0
            total_revenue += e["revenue_cents"] or 0
            total_proceeds += e["proceeds_cents"] or 0
            daily_data.append({
                "date": str(e["date"]),
                "new_subs": e["new_subscriptions"] or 0,
                "renewals": e["renewals"] or 0,
                "conversions": e["conversions"] or 0,
                "cancellations": e["cancellations"] or 0,
                "revenue_cents": e["revenue_cents"] or 0,
                "proceeds_cents": e["proceeds_cents"] or 0,
            })

    return {
        "current_snapshot": {
            "date": str(latest_sub["date"]) if latest_sub else None,
            "active_subscriptions": latest_sub["active_subscriptions"] if latest_sub else 0,
            "active_free_trials": latest_sub["active_free_trials"] if latest_sub else 0,
        } if latest_sub else None,
        "period_totals": {
            "new_subscriptions": total_new,
            "renewals": total_renewals,
            "conversions": total_conversions,
            "cancellations": total_cancellations,
            "revenue_cents": total_revenue,
            "revenue_display": f"${total_revenue / 100:,.2f}",
            "proceeds_cents": total_proceeds,
            "proceeds_display": f"${total_proceeds / 100:,.2f}",
            "apple_fee_cents": total_revenue - total_proceeds,
        },
        "daily": daily_data,
        "days_with_data": len(daily_data),
    }


# --- S26 Phase 4: Daily expiry sweep ---
# Marks pending-cancel records as truly canceled when their pending_cancel_at <= now().
# Run manually first via the endpoint; once verified, can be wired into shadow sync auto-run.

async def s26_expire_pending_cancels(conn) -> dict:
    """Find all cancel_state='pending' records whose pending_cancel_at has passed
    and mark them fully canceled. Returns count + sample for logging."""
    now_utc = datetime.now(timezone.utc)
    rows = await conn.fetch(
        """SELECT id, email, stripe_subscription_id, source, pending_cancel_at
           FROM subscriptions
           WHERE cancel_state = 'pending'
             AND pending_cancel_at IS NOT NULL
             AND pending_cancel_at <= $1
             AND status IN ('active', 'trialing')""",
        now_utc
    )
    expired = []
    for r in rows:
        try:
            await conn.execute(
                """UPDATE subscriptions
                   SET status = 'canceled',
                       canceled_at = COALESCE(canceled_at, NOW()),
                       cancel_state = 'expired',
                       updated_at = NOW()
                   WHERE id = $1""",
                r["id"]
            )
            expired.append({
                "id": r["id"],
                "email": r["email"],
                "sub_id": r["stripe_subscription_id"],
                "source": r["source"],
                "pending_cancel_at": r["pending_cancel_at"].isoformat() if r["pending_cancel_at"] else None,
            })
        except Exception as e:
            print(f"[S26 Sweep] Error expiring id={r['id']}: {e}")
    return {"expired_count": len(expired), "expired_sample": expired[:25]}


@app.post("/api/admin/s26-expire-sweep")
async def s26_expire_sweep_endpoint(request: Request):
    """Manual trigger for the period-end expiry sweep."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    async with db_pool.acquire() as conn:
        result = await s26_expire_pending_cancels(conn)
    return {"status": "ok", **result}


# --- S26 Phase 5: Backfill mis-cancelled Stripe records ---
# Reactivates Stripe records that were canceled under old click-time semantics
# but whose current_period_end is still in the future. Sets cancel_state='pending'.
# Preview/apply pattern: pass preview=false to actually apply changes.

@app.post("/api/admin/s26-backfill-pending")
async def s26_backfill_pending(request: Request):
    """Reactivate mis-cancelled Stripe records as cancel_state='pending' if still in grace.
    Pass JSON {"preview": false, "days": 60} to apply. Default is preview-only."""
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    if not STRIPE_SECRET_KEY:
        raise HTTPException(status_code=500, detail="Stripe not configured")

    try:
        body = await request.json()
    except Exception:
        body = {}
    preview = body.get("preview", True)
    days = max(1, min(int(body.get("days", 60)), 365))
    id_skip = set(int(x) for x in body.get("id_skip", []))

    batch_id = f"s26_backfill_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """SELECT id, email, stripe_subscription_id, canceled_at, import_batch
               FROM subscriptions
               WHERE status = 'canceled'
                 AND source = 'stripe'
                 AND stripe_subscription_id LIKE 'sub_%'
                 AND canceled_at IS NOT NULL
                 AND canceled_at > NOW() - ($1 || ' days')::INTERVAL
                 AND (import_batch IS NULL OR import_batch NOT LIKE '%s25_test_cancel_%')
               ORDER BY canceled_at DESC""",
            str(days)
        )
        # Build set of emails that already have an active sub elsewhere
        active_emails_rows = await conn.fetch(
            "SELECT DISTINCT lower(email) AS em FROM subscriptions WHERE status IN ('active', 'trialing') AND email IS NOT NULL AND email != ''"
        )
        active_emails = set(r["em"] for r in active_emails_rows)

    import stripe as _s26_stripe
    _s26_stripe.api_key = STRIPE_SECRET_KEY
    to_reactivate = []
    skipped_expired = []
    skipped_dispute = []
    skipped_dup_active = []
    skipped_id_list = []
    not_found = []
    errors = []

    for r in rows:
        sub_id = r["stripe_subscription_id"]
        if r["id"] in id_skip:
            skipped_id_list.append({"id": r["id"], "email": r["email"]})
            continue
        # Skip if email already has another active sub
        if r["email"] and r["email"].strip().lower() in active_emails:
            skipped_dup_active.append({"id": r["id"], "email": r["email"], "sub_id": sub_id})
            continue
        try:
            real = _s26_stripe.Subscription.retrieve(sub_id)
            pe = real.get("current_period_end")
            now_ts = int(datetime.now(timezone.utc).timestamp())
            cancellation = real.get("cancellation_details") or {}
            reason = cancellation.get("reason") or ""

            # Skip if cancelled for payment failure / dispute / fraud
            if reason in ("payment_failed", "payment_disputed"):
                skipped_dispute.append({"id": r["id"], "email": r["email"], "reason": reason})
                continue

            if pe and pe > now_ts:
                to_reactivate.append({
                    "id": r["id"],
                    "email": r["email"],
                    "sub_id": sub_id,
                    "period_end": datetime.fromtimestamp(pe, tz=timezone.utc).isoformat(),
                    "reason": reason or None,
                })
            else:
                skipped_expired.append({"id": r["id"], "email": r["email"], "period_end_ts": pe})
        except _s26_stripe.error.InvalidRequestError:
            not_found.append({"id": r["id"], "email": r["email"], "sub_id": sub_id})
        except Exception as e:
            errors.append({"id": r["id"], "email": r["email"], "error": str(e)})

    if preview:
        return {
            "status": "preview",
            "batch_id": batch_id,
            "scanned_days": days,
            "candidates_total": len(rows),
            "would_reactivate": len(to_reactivate),
            "skipped_truly_expired": len(skipped_expired),
            "skipped_dispute_or_failure": len(skipped_dispute),
            "skipped_dup_active_email": len(skipped_dup_active),
            "skipped_by_id_list": len(skipped_id_list),
            "not_found_in_stripe": len(not_found),
            "errors": len(errors),
            "reactivate_sample": to_reactivate[:50],
            "skipped_dup_active_sample": skipped_dup_active[:10],
            "skipped_dispute_sample": skipped_dispute[:10],
            "not_found_sample": not_found[:10],
            "errors_sample": errors[:10],
            "note": "Send {\"preview\": false} to apply. Use {\"id_skip\": [123, 456]} to exclude specific ids. Records with import_batch starting 's25_test_cancel_' are auto-excluded.",
        }

    # APPLY
    reactivated = 0
    apply_errors = 0
    async with db_pool.acquire() as conn:
        for item in to_reactivate:
            try:
                await conn.execute(
                    """UPDATE subscriptions
                       SET status = 'active',
                           canceled_at = NULL,
                           reactivated_at = NOW(),
                           cancel_state = 'pending',
                           pending_cancel_at = $1,
                           cancel_requested_at = COALESCE(cancel_requested_at, canceled_at, NOW()),
                           import_batch = $2,
                           updated_at = NOW()
                       WHERE id = $3 AND status = 'canceled'""",
                    datetime.fromisoformat(item["period_end"]), batch_id, item["id"]
                )
                reactivated += 1
            except Exception as e:
                print(f"[S26 Backfill] Error id={item['id']}: {e}")
                apply_errors += 1

    return {
        "status": "applied",
        "batch_id": batch_id,
        "reactivated": reactivated,
        "errors": apply_errors,
        "note": f"Reversible via batch_id={batch_id}",
    }


# --- S26: Lookup records by import_batch (read-only) ---

@app.get("/api/admin/s26-batch-lookup")
async def s26_batch_lookup(request: Request, batch: str = ""):
    """READ-ONLY. Returns id, email, plan_amount, plan_interval, status,
    cancel_state for all records with the given import_batch (exact match
    or substring)."""
    pw = request.query_params.get("pw") or request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    if not batch:
        raise HTTPException(status_code=400, detail="Pass ?batch=...")
    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """SELECT id, email, stripe_subscription_id, source, status, cancel_state,
                      plan_amount, plan_interval, pending_cancel_at, import_batch
               FROM subscriptions
               WHERE import_batch LIKE '%' || $1 || '%'
               ORDER BY id""",
            batch
        )
    # Compute normalized MRR contribution
    total_mrr_cents = 0
    records = []
    for r in rows:
        d = dict(r)
        if r["plan_interval"] == "month":
            contrib = r["plan_amount"] or 0
        elif r["plan_interval"] == "year":
            contrib = (r["plan_amount"] or 0) // 12
        else:
            contrib = 0
        if r["status"] in ("active", "trialing"):
            total_mrr_cents += contrib
        d["mrr_contribution_cents"] = contrib
        d["pending_cancel_at"] = r["pending_cancel_at"].isoformat() if r["pending_cancel_at"] else None
        records.append(d)
    return {
        "status": "ok",
        "count": len(records),
        "total_mrr_cents_added": total_mrr_cents,
        "total_mrr_display": f"${total_mrr_cents/100:,.2f}/mo",
        "records": records,
    }


# --- S26: Fuzzy email-match diagnostic for historical Stripe records ---
# Tosh's S25 response showed that some "missing from ymove" Stripe records are
# actually present in ymove under a slightly different email (e.g. amets30 vs
# amets20, or alessandracelia.volpato vs alessandraclelia.volpato).
# This is read-only: takes a list of candidate Stripe sub_ids, pulls customer
# name from Stripe, then searches our DB for active records with similar
# names or local-part. Outputs candidates for manual review. No writes.

def _s26_levenshtein(a: str, b: str) -> int:
    """Tiny Levenshtein distance, no external deps."""
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i] + [0] * len(b)
        for j, cb in enumerate(b, 1):
            curr[j] = min(
                prev[j] + 1,
                curr[j - 1] + 1,
                prev[j - 1] + (0 if ca == cb else 1),
            )
        prev = curr
    return prev[-1]


@app.post("/api/admin/s26-fuzzy-email-match")
async def s26_fuzzy_email_match(request: Request):
    """READ-ONLY: For each candidate Stripe sub_id, fetch customer name from Stripe,
    then search ALL ymove-source emails in our DB for fuzzy matches by name + local-part.
    Body: {"sub_ids": ["sub_xxx", ...]}  (defaults to S25 historical 8 if omitted)
    """
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    if not STRIPE_SECRET_KEY:
        raise HTTPException(status_code=500, detail="Stripe not configured")

    try:
        body = await request.json()
    except Exception:
        body = {}

    # Default candidates: the 8 historical Stripe emails from S25 wrap
    default_emails = [
        "abbey.e.baier@gmail.com", "ahfouch@gmail.com",
        "alessandraclelia.volpato@gmail.com", "amets30@yahoo.com",
        "andreedesrochers@gmail.com", "cassyroop@gmail.com",
        "chloe.levray@gmail.com", "hstrandness@gmail.com",
    ]
    candidate_emails = body.get("emails", default_emails)

    import stripe as _s26_stripe
    _s26_stripe.api_key = STRIPE_SECRET_KEY

    # Pull all known emails + names from our DB once
    async with db_pool.acquire() as conn:
        known = await conn.fetch(
            """SELECT lower(email) AS email, first_name, last_name, source, status,
                      stripe_subscription_id
               FROM subscriptions
               WHERE email IS NOT NULL AND email != ''"""
        )

    db_by_email = {r["email"]: dict(r) for r in known}
    all_db_emails = list(db_by_email.keys())

    results = []
    for src_email in candidate_emails:
        src_email_l = src_email.strip().lower()
        entry = {"candidate_email": src_email_l, "stripe_customer_name": None,
                 "exact_match_in_db": src_email_l in db_by_email,
                 "matches": []}

        # Find the Stripe customer for this email via the subscription record in our DB
        our_rec = db_by_email.get(src_email_l)
        if our_rec and (our_rec.get("stripe_subscription_id") or "").startswith("sub_"):
            try:
                real = _s26_stripe.Subscription.retrieve(our_rec["stripe_subscription_id"])
                cust_id = real.get("customer")
                if cust_id:
                    cust = _s26_stripe.Customer.retrieve(cust_id)
                    entry["stripe_customer_name"] = cust.get("name") or ""
            except Exception as e:
                entry["stripe_lookup_error"] = str(e)

        # Compute fuzzy matches in our DB
        local_part = src_email_l.split("@")[0]
        cand_name_lower = (entry["stripe_customer_name"] or "").lower().strip()

        for db_email, rec in db_by_email.items():
            if db_email == src_email_l:
                continue
            if rec["status"] not in ("active", "trialing"):
                continue
            db_local = db_email.split("@")[0]
            local_dist = _s26_levenshtein(local_part, db_local)
            db_name = ((rec.get("first_name") or "") + " " + (rec.get("last_name") or "")).lower().strip()
            name_match = bool(cand_name_lower and db_name and (cand_name_lower in db_name or db_name in cand_name_lower))

            if local_dist <= 3 or name_match:
                entry["matches"].append({
                    "db_email": db_email,
                    "db_name": db_name or None,
                    "source": rec["source"],
                    "local_part_distance": local_dist,
                    "name_match": name_match,
                    "sub_id": rec["stripe_subscription_id"],
                })

        # Sort matches by best (low distance, name match wins ties)
        entry["matches"].sort(key=lambda m: (0 if m["name_match"] else 1, m["local_part_distance"]))
        entry["matches"] = entry["matches"][:5]
        results.append(entry)

    return {"status": "ok", "candidates_checked": len(results), "results": results}


# --- S26 one-off: Markus reactivation + email typo fix ---
# Markus is a real ymove user (markus.zwigart@gmx.de) who entered a typo in Stripe
# checkout (markus.zwigart@gmx.dr). S25 cancelled him as junk. This endpoint
# reactivates AND fixes the email so future shadow syncs match correctly.
# Hardcoded id + emails for safety.

@app.post("/api/admin/s26-fix-markus")
async def s26_fix_markus(request: Request):
    pw = request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    target_id = 1662
    expected_old_email = "markus.zwigart@gmx.dr"
    new_email = "markus.zwigart@gmx.de"
    batch_id = f"s26_markus_fix_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"

    async with db_pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, email, status, stripe_subscription_id FROM subscriptions WHERE id = $1",
            target_id
        )
        if not row:
            raise HTTPException(status_code=404, detail=f"id {target_id} not found")
        if (row["email"] or "").lower() != expected_old_email.lower():
            raise HTTPException(
                status_code=409,
                detail=f"Safety check failed: expected email {expected_old_email}, found {row['email']}"
            )
        # Check no other active sub already exists for the new email
        dup = await conn.fetchrow(
            "SELECT id FROM subscriptions WHERE lower(email) = lower($1) AND status IN ('active', 'trialing') LIMIT 1",
            new_email
        )
        if dup:
            raise HTTPException(
                status_code=409,
                detail=f"An active sub already exists for {new_email} (id {dup['id']}). Manual review needed."
            )
        await conn.execute(
            """UPDATE subscriptions
               SET status = 'active',
                   email = $1,
                   canceled_at = NULL,
                   reactivated_at = NOW(),
                   import_batch = $2,
                   updated_at = NOW()
               WHERE id = $3""",
            new_email, batch_id, target_id
        )
    return {
        "status": "ok",
        "batch_id": batch_id,
        "id": target_id,
        "old_email": expected_old_email,
        "new_email": new_email,
        "note": f"Reversible via batch_id={batch_id}",
    }


# --- S26 Phase 1: Period-end cancellation diagnostic (read-only) ---
# Lists Stripe-source records that are currently marked 'canceled' in our DB
# but whose Stripe current_period_end is still in the future. These are records
# that Phase 4 (backfill) will reactivate as cancel_state='pending'.

@app.get("/api/admin/s26-pending-cancel-diagnostic")
async def s26_pending_cancel_diagnostic(request: Request, days: int = 60):
    """READ-ONLY: scan recently-cancelled Stripe records and check Stripe API
    for current_period_end. Records with period_end > now() were mis-cancelled
    under the old click-time semantics and should be in pending state.
    """
    pw = request.query_params.get("pw") or request.headers.get("X-Admin-Password", "")
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    if not STRIPE_SECRET_KEY:
        raise HTTPException(status_code=500, detail="Stripe not configured")

    cutoff_days = max(1, min(days, 365))
    candidates = []
    in_grace = []
    truly_expired = []
    not_found = []
    errors = []

    async with db_pool.acquire() as conn:
        rows = await conn.fetch(
            """SELECT id, email, stripe_subscription_id, source, status,
                      canceled_at, created_at
               FROM subscriptions
               WHERE status = 'canceled'
                 AND source = 'stripe'
                 AND stripe_subscription_id LIKE 'sub_%'
                 AND canceled_at IS NOT NULL
                 AND canceled_at > NOW() - ($1 || ' days')::INTERVAL
               ORDER BY canceled_at DESC""",
            str(cutoff_days)
        )

    candidates = [dict(r) for r in rows]

    import stripe as stripe_lib
    stripe_lib.api_key = STRIPE_SECRET_KEY

    for c in candidates:
        sub_id = c["stripe_subscription_id"]
        try:
            real_sub = stripe_lib.Subscription.retrieve(sub_id)
            period_end = real_sub.get("current_period_end")
            cape = real_sub.get("cancel_at_period_end")
            real_status = real_sub.get("status")
            now_ts = int(datetime.now(timezone.utc).timestamp())
            entry = {
                "id": c["id"],
                "email": c["email"],
                "sub_id": sub_id,
                "our_canceled_at": c["canceled_at"].isoformat() if c["canceled_at"] else None,
                "stripe_status": real_status,
                "stripe_period_end": datetime.fromtimestamp(period_end, tz=timezone.utc).isoformat() if period_end else None,
                "stripe_cancel_at_period_end": cape,
            }
            if period_end and period_end > now_ts:
                entry["bucket"] = "in_grace"
                in_grace.append(entry)
            else:
                entry["bucket"] = "truly_expired"
                truly_expired.append(entry)
        except stripe_lib.error.InvalidRequestError:
            not_found.append({"id": c["id"], "email": c["email"], "sub_id": sub_id})
        except Exception as e:
            errors.append({"id": c["id"], "email": c["email"], "sub_id": sub_id, "error": str(e)})

    return {
        "status": "ok",
        "scanned_days": cutoff_days,
        "candidates_total": len(candidates),
        "in_grace_count": len(in_grace),
        "truly_expired_count": len(truly_expired),
        "not_found_count": len(not_found),
        "errors_count": len(errors),
        "in_grace_sample": in_grace[:50],
        "in_grace_full": in_grace,
        "truly_expired_sample": truly_expired[:10],
        "not_found_sample": not_found[:10],
        "errors_sample": errors[:10],
        "note": "READ-ONLY. 'in_grace' records are candidates for Phase 4 backfill (reactivate as cancel_state='pending').",
    }


# --- S28: Verify only_in_ours records against source-of-truth APIs ---
# Reads the 17 only_in_ours emails from the latest ymove-diff (8 historical Stripe,
# 9 Apple/Google) plus the 9 backfill records, and checks each against its
# source-of-truth API. Stripe records hit Stripe API. Apple/Google records hit
# ymove member-lookup (single-email, no status filter).
# READ-ONLY. No DB writes. No preview/apply needed.

S28_STRIPE_HISTORICAL_EMAILS = [
    "abbey.e.baier@gmail.com",
    "ahfouch@gmail.com",
    "alessandraclelia.volpato@gmail.com",
    "amets30@yahoo.com",
    "andreedesrochers@gmail.com",
    "cassyroop@gmail.com",
    "chloe.levray@gmail.com",
    "hstrandness@gmail.com",
]

S28_APPLE_GOOGLE_ONLY_IN_OURS = [
    "andrea.nenadic@gmail.com",
    "juliette-bisot@hotmail.fr",
    "llbankert@gmail.com",
    "morganaj1022@gmail.com",
    "nienkenijp@gmail.com",
    "smb2895@optonline.net",
    "vacantpatient@gmail.com",
    "brigidcgriffin@gmail.com",
    "jennifer.miller4@gmail.com",
]

S28_BACKFILL_BATCH = "s26_backfill_20260413_011354"


@app.post("/api/admin/s28-verify-only-in-ours")
async def s28_verify_only_in_ours(request: Request):
    """READ-ONLY. For every record in only_in_ours from the latest ymove-diff,
    verify it against the appropriate source of truth.

    - Stripe-source emails: hit Stripe API. Look up customer, list subs, return active state.
    - Apple/Google-source emails: hit ymove member-lookup (single email), return found/active state.
    - Backfill batch records: hit Stripe API, confirm period_end semantics still valid.

    No DB writes. No body required. Returns categorized results.
    """
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    if not STRIPE_SECRET_KEY:
        raise HTTPException(status_code=500, detail="Stripe not configured")
    if not YMOVE_API_KEY:
        raise HTTPException(status_code=500, detail="YMOVE_API_KEY not configured")

    import stripe as _s28_stripe
    _s28_stripe.api_key = STRIPE_SECRET_KEY

    # ---------- Pull our DB records for all target emails ----------
    all_emails = (
        [e.lower() for e in S28_STRIPE_HISTORICAL_EMAILS]
        + [e.lower() for e in S28_APPLE_GOOGLE_ONLY_IN_OURS]
    )
    async with db_pool.acquire() as conn:
        our_rows = await conn.fetch(
            """SELECT id, email, source, status, stripe_subscription_id, stripe_customer_id,
                      plan_amount, plan_interval, current_period_end, created_at, import_batch
               FROM subscriptions
               WHERE lower(email) = ANY($1::text[])
                 AND status IN ('active', 'trialing')
               ORDER BY email, created_at DESC""",
            all_emails
        )
        our_by_email = {}
        for r in our_rows:
            em = r["email"].lower()
            our_by_email.setdefault(em, []).append(dict(r))

        # Backfill batch records (separate query, fetch all 9)
        backfill_rows = await conn.fetch(
            """SELECT id, email, source, status, stripe_subscription_id, stripe_customer_id,
                      plan_amount, plan_interval, current_period_end, cancel_state,
                      pending_cancel_at, created_at
               FROM subscriptions
               WHERE import_batch = $1
               ORDER BY email""",
            S28_BACKFILL_BATCH
        )

    # ---------- Helper: hit Stripe API for an email ----------
    def stripe_check(email: str, expected_sub_id: str = None) -> dict:
        result = {"email": email, "stripe_found": False, "any_active": False, "subs": []}
        try:
            customers = _s28_stripe.Customer.list(email=email, limit=10)
            if not customers.data:
                result["stripe_found"] = False
                result["note"] = "No Stripe customer with this email"
                return result
            result["stripe_found"] = True
            result["customer_count"] = len(customers.data)
            for cust in customers.data:
                subs = _s28_stripe.Subscription.list(customer=cust.id, status="all", limit=20)
                for s in subs.data:
                    sub_info = {
                        "sub_id": s.id,
                        "status": s.status,
                        "customer_id": cust.id,
                        "current_period_end": s.current_period_end,
                        "cancel_at_period_end": s.cancel_at_period_end,
                        "canceled_at": s.canceled_at,
                        "matches_our_sub_id": (s.id == expected_sub_id) if expected_sub_id else None,
                    }
                    if s.status in ("active", "trialing"):
                        result["any_active"] = True
                    result["subs"].append(sub_info)
        except Exception as e:
            result["error"] = str(e)[:200]
        return result

    # ---------- Helper: hit ymove single-email lookup ----------
    async def ymove_check(email: str) -> dict:
        result = {"email": email, "ymove_found": False}
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                resp = await client.get(
                    f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup",
                    headers={"X-Authorization": YMOVE_API_KEY},
                    params={"email": email}
                )
                if resp.status_code != 200:
                    result["error"] = f"HTTP {resp.status_code}"
                    return result
                data = resp.json()
                if not data.get("found"):
                    result["ymove_found"] = False
                    result["note"] = "ymove returned found=false"
                    return result
                result["ymove_found"] = True
                user = data.get("user", {})
                result["active_subscription"] = user.get("activeSubscription")
                result["previously_subscribed"] = user.get("previouslySubscribed")
                result["subscription_provider"] = user.get("subscriptionProvider")
                result["ymove_user_id"] = user.get("id") or user.get("userId")
                # Capture status if present
                result["status_field"] = user.get("status") or user.get("subscriptionStatus")
        except Exception as e:
            result["error"] = str(e)[:200]
        return result

    # ---------- Verify Stripe historical 8 ----------
    # Check both Stripe API (sub state) AND ymove single-email lookup (does ymove
    # have this email under any provider, regardless of status filter from bulk pull?)
    stripe_historical_results = []
    for email in S28_STRIPE_HISTORICAL_EMAILS:
        em = email.lower()
        our = our_by_email.get(em, [])
        our_rec = our[0] if our else None
        expected_sub_id = our_rec["stripe_subscription_id"] if our_rec else None
        check = stripe_check(em, expected_sub_id)
        ycheck = await ymove_check(em)
        stripe_historical_results.append({
            "email": em,
            "our_record": {
                "sub_id": our_rec["stripe_subscription_id"] if our_rec else None,
                "status": our_rec["status"] if our_rec else None,
                "created_at": our_rec["created_at"].isoformat() if our_rec and our_rec["created_at"] else None,
                "current_period_end": our_rec["current_period_end"].isoformat() if our_rec and our_rec.get("current_period_end") else None,
                "plan_amount": our_rec["plan_amount"] if our_rec else None,
            } if our_rec else None,
            "stripe_truth": check,
            "ymove_truth": ycheck,
            "verdict": _s28_verdict(our_rec, check),
        })
        await asyncio.sleep(0.3)  # rate-limit politeness for ymove

    # ---------- Verify Apple/Google 9 against ymove ----------
    apple_google_results = []
    for email in S28_APPLE_GOOGLE_ONLY_IN_OURS:
        em = email.lower()
        our = our_by_email.get(em, [])
        our_rec = our[0] if our else None
        ycheck = await ymove_check(em)
        apple_google_results.append({
            "email": em,
            "our_record": {
                "sub_id": our_rec["stripe_subscription_id"] if our_rec else None,
                "source": our_rec["source"] if our_rec else None,
                "status": our_rec["status"] if our_rec else None,
                "created_at": our_rec["created_at"].isoformat() if our_rec and our_rec["created_at"] else None,
                "import_batch": our_rec["import_batch"] if our_rec else None,
            } if our_rec else None,
            "ymove_truth": ycheck,
        })
        await asyncio.sleep(0.3)  # rate-limit politeness

    # ---------- Verify backfill batch 9 against Stripe AND ymove ----------
    backfill_results = []
    for r in backfill_rows:
        email = r["email"].lower()
        check = stripe_check(email, r["stripe_subscription_id"])
        ycheck = await ymove_check(email)
        backfill_results.append({
            "email": email,
            "our_record": {
                "id": r["id"],
                "sub_id": r["stripe_subscription_id"],
                "status": r["status"],
                "cancel_state": r.get("cancel_state"),
                "pending_cancel_at": r["pending_cancel_at"].isoformat() if r.get("pending_cancel_at") else None,
                "current_period_end": r["current_period_end"].isoformat() if r.get("current_period_end") else None,
                "plan_interval": r.get("plan_interval"),
                "plan_amount": r.get("plan_amount"),
            },
            "stripe_truth": check,
            "ymove_truth": ycheck,
            "verdict": _s28_verdict(dict(r), check),
        })
        await asyncio.sleep(0.3)

    # ---------- Summary tallies ----------
    def tally(results, key="verdict"):
        out = {}
        for r in results:
            v = r.get(key, "unknown")
            out[v] = out.get(v, 0) + 1
        return out

    return {
        "status": "ok",
        "note": "READ-ONLY verification of only_in_ours records against source-of-truth APIs.",
        "summary": {
            "stripe_historical_count": len(stripe_historical_results),
            "apple_google_count": len(apple_google_results),
            "backfill_count": len(backfill_results),
            "stripe_historical_verdicts": tally(stripe_historical_results),
            "backfill_verdicts": tally(backfill_results),
        },
        "stripe_historical": stripe_historical_results,
        "apple_google": apple_google_results,
        "backfill_batch": backfill_results,
    }


import os as _s28_os

_S28_DUMP_CACHE_PATH = "/tmp/s28_ymove_dump_cache.json"


@app.post("/api/admin/s28-full-dump")
@app.get("/api/admin/s28-full-dump")
async def s28_full_dump(request: Request):
    """READ-ONLY. Produces a complete row-by-row reconciliation between our
    active+trialing subscriptions and ymove's member-lookup/all bulk pull.

    No categorization. No verdicts. No filtering of results. One row per
    unique email from either side, with every relevant field.

    Query params:
      ymove_status: status filter for ymove API (default: empty = no filter)
      format: 'json' (default) or 'csv' (returns downloadable file)
      use_cache: 'true' uses /tmp cache if present, otherwise re-pulls
      force_refresh: 'true' forces a re-pull even if cache exists
    """
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")
    if not YMOVE_API_KEY:
        raise HTTPException(status_code=500, detail="YMOVE_API_KEY not configured")

    ymove_status = request.query_params.get("ymove_status", "")
    fmt = request.query_params.get("format", "json").lower()
    use_cache = request.query_params.get("use_cache", "").lower() == "true"
    force_refresh = request.query_params.get("force_refresh", "").lower() == "true"

    started_at = datetime.now(timezone.utc)
    pull_source = "fresh"
    ymove_raw = None

    # ---------- Cache check ----------
    if use_cache and not force_refresh and _s28_os.path.exists(_S28_DUMP_CACHE_PATH):
        try:
            with open(_S28_DUMP_CACHE_PATH, "r") as f:
                cached = json.load(f)
            if cached.get("ymove_status_param") == ymove_status:
                ymove_raw = cached.get("users", [])
                pull_source = f"cache ({cached.get('cached_at')})"
        except Exception as e:
            print(f"[s28-full-dump] cache read error: {e}")

    # ---------- Fresh pull if no cache hit ----------
    if ymove_raw is None:
        ymove_raw = []
        pages_pulled = 0
        pull_status = "starting"
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                page = 1
                while True:
                    params = {"page": str(page)}
                    if ymove_status:
                        params["status"] = ymove_status
                    try:
                        resp = await client.get(
                            f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup/all",
                            headers={"X-Authorization": YMOVE_API_KEY},
                            params=params
                        )
                    except httpx.TimeoutException:
                        pull_status = f"timeout_page_{page}"
                        break
                    if resp.status_code == 429:
                        retry_after = float(resp.headers.get("retry-after", "5"))
                        await asyncio.sleep(retry_after)
                        continue
                    if resp.status_code != 200:
                        pull_status = f"http_{resp.status_code}_page_{page}"
                        break
                    data = resp.json()
                    users = data.get("users", [])
                    ymove_raw.extend(users)
                    total_pages = data.get("totalPages", 1)
                    pages_pulled = page
                    if page >= total_pages:
                        pull_status = "success"
                        break
                    page += 1
                    await asyncio.sleep(0.5)
        except Exception as e:
            pull_status = f"error: {str(e)[:200]}"

        if pull_status != "success":
            return JSONResponse(status_code=502, content={
                "error": "ymove pull failed",
                "pull_status": pull_status,
                "pages_pulled": pages_pulled,
            })

        # Write to cache for future calls
        try:
            with open(_S28_DUMP_CACHE_PATH, "w") as f:
                json.dump({
                    "cached_at": datetime.now(timezone.utc).isoformat(),
                    "ymove_status_param": ymove_status,
                    "pages_pulled": pages_pulled,
                    "users": ymove_raw,
                }, f)
        except Exception as e:
            print(f"[s28-full-dump] cache write error: {e}")

    # ---------- Index ymove data by lowercased email ----------
    ymove_by_email = {}
    for u in ymove_raw:
        em = (u.get("email") or "").strip().lower()
        if not em:
            continue
        provider = (u.get("subscriptionProvider") or None)
        if provider:
            provider = provider.lower()
            if provider not in ("apple", "google", "stripe", "manual", "undetermined"):
                provider = "undetermined"
        ymove_by_email[em] = {
            "ymove_provider": provider,
            "ymove_active_subscription": u.get("activeSubscription"),
            "ymove_previously_subscribed": u.get("previouslySubscribed"),
            "ymove_user_id": u.get("id") or u.get("userId"),
            "ymove_first_name": u.get("firstName") or u.get("first_name") or "",
            "ymove_last_name": u.get("lastName") or u.get("last_name") or "",
        }

    # ---------- Pull our active+trialing ----------
    async with db_pool.acquire() as conn:
        our_rows = await conn.fetch("""
            SELECT id, email, source, status, stripe_subscription_id,
                   stripe_customer_id, plan_amount, plan_interval,
                   current_period_end, created_at, import_batch,
                   cancel_state, pending_cancel_at,
                   first_name, last_name
            FROM subscriptions
            WHERE status IN ('active', 'trialing')
              AND email IS NOT NULL AND email != ''
        """)

    our_by_email = {}
    for r in our_rows:
        em = r["email"].strip().lower()
        # If there's a duplicate on same email, keep the most recent (higher id)
        existing = our_by_email.get(em)
        if existing is None or r["id"] > existing["id"]:
            our_by_email[em] = {
                "id": r["id"],
                "email_raw": r["email"],
                "our_source": r["source"],
                "our_sub_id": r["stripe_subscription_id"],
                "our_customer_id": r["stripe_customer_id"],
                "our_status": r["status"],
                "our_plan_amount": r["plan_amount"],
                "our_plan_interval": r["plan_interval"],
                "our_current_period_end": r["current_period_end"].isoformat() if r["current_period_end"] else None,
                "our_created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "our_import_batch": r["import_batch"],
                "our_cancel_state": r["cancel_state"],
                "our_pending_cancel_at": r["pending_cancel_at"].isoformat() if r["pending_cancel_at"] else None,
                "our_first_name": r["first_name"],
                "our_last_name": r["last_name"],
            }

    # ---------- Full outer join ----------
    all_emails = sorted(set(our_by_email.keys()) | set(ymove_by_email.keys()))
    rows_out = []
    count_in_both = 0
    count_only_in_ours = 0
    count_only_in_ymove = 0
    for em in all_emails:
        ours = our_by_email.get(em)
        theirs = ymove_by_email.get(em)
        in_ours = ours is not None
        in_ymove = theirs is not None
        if in_ours and in_ymove:
            count_in_both += 1
        elif in_ours:
            count_only_in_ours += 1
        else:
            count_only_in_ymove += 1

        row = {
            "email": em,
            "in_ours": in_ours,
            "in_ymove": in_ymove,
            "our_source": ours["our_source"] if ours else None,
            "our_sub_id": ours["our_sub_id"] if ours else None,
            "our_status": ours["our_status"] if ours else None,
            "our_plan_amount": ours["our_plan_amount"] if ours else None,
            "our_plan_interval": ours["our_plan_interval"] if ours else None,
            "our_created_at": ours["our_created_at"] if ours else None,
            "our_current_period_end": ours["our_current_period_end"] if ours else None,
            "our_import_batch": ours["our_import_batch"] if ours else None,
            "our_cancel_state": ours["our_cancel_state"] if ours else None,
            "our_pending_cancel_at": ours["our_pending_cancel_at"] if ours else None,
            "our_id": ours["id"] if ours else None,
            "our_first_name": ours["our_first_name"] if ours else None,
            "our_last_name": ours["our_last_name"] if ours else None,
            "ymove_provider": theirs["ymove_provider"] if theirs else None,
            "ymove_active_subscription": theirs["ymove_active_subscription"] if theirs else None,
            "ymove_previously_subscribed": theirs["ymove_previously_subscribed"] if theirs else None,
            "ymove_user_id": theirs["ymove_user_id"] if theirs else None,
            "ymove_first_name": theirs["ymove_first_name"] if theirs else None,
            "ymove_last_name": theirs["ymove_last_name"] if theirs else None,
        }
        rows_out.append(row)

    runtime_seconds = (datetime.now(timezone.utc) - started_at).total_seconds()

    pull_meta = {
        "ymove_status_param": ymove_status or "(none)",
        "pull_source": pull_source,
        "total_ymove_records": len(ymove_by_email),
        "total_our_records": len(our_by_email),
        "runtime_seconds": round(runtime_seconds, 1),
    }
    counts = {
        "in_both": count_in_both,
        "only_in_ours": count_only_in_ours,
        "only_in_ymove": count_only_in_ymove,
        "total_unique_emails": len(all_emails),
    }

    # ---------- CSV output path ----------
    if fmt == "csv":
        import csv as _csv
        from io import StringIO as _StringIO
        buf = _StringIO()
        fieldnames = [
            "email", "in_ours", "in_ymove",
            "our_source", "our_sub_id", "our_status",
            "our_plan_amount", "our_plan_interval",
            "our_created_at", "our_current_period_end",
            "our_import_batch", "our_cancel_state", "our_pending_cancel_at",
            "our_id", "our_first_name", "our_last_name",
            "ymove_provider", "ymove_active_subscription",
            "ymove_previously_subscribed", "ymove_user_id",
            "ymove_first_name", "ymove_last_name",
        ]
        writer = _csv.DictWriter(buf, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows_out:
            writer.writerow({k: ("" if row.get(k) is None else row.get(k)) for k in fieldnames})
        csv_text = buf.getvalue()
        filename = f"s28_full_dump_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"
        return Response(
            content=csv_text,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    # ---------- JSON output path (default) ----------
    return {
        "status": "ok",
        "note": "READ-ONLY full dump. No categorization. Use .rows to iterate.",
        "pull_meta": pull_meta,
        "counts": counts,
        "rows": rows_out,
    }


@app.post("/api/admin/s28-test-account-scan")
async def s28_test_account_scan(request: Request):
    """READ-ONLY. Scan our active+trialing records for emails matching the
    _is_test_email filter. These are likely records imported before ymove's
    test-account filter was active, still sitting in our DB while Tosh's
    dashboard filters them out. No DB writes."""
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT id, email, source, status, stripe_subscription_id,
                   created_at, import_batch, plan_amount
            FROM subscriptions
            WHERE status IN ('active', 'trialing')
              AND email IS NOT NULL AND email != ''
            ORDER BY source, email
        """)

    matches_by_source = {}
    all_matches = []
    for r in rows:
        em = r["email"]
        if _is_test_email(em):
            entry = {
                "id": r["id"],
                "email": em,
                "source": r["source"],
                "status": r["status"],
                "sub_id": r["stripe_subscription_id"],
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "import_batch": r["import_batch"],
                "plan_amount": r["plan_amount"],
            }
            matches_by_source.setdefault(r["source"] or "null", []).append(entry)
            all_matches.append(entry)

    summary = {
        "total_active_trialing_scanned": len(rows),
        "total_test_matches": len(all_matches),
        "matches_by_source": {k: len(v) for k, v in matches_by_source.items()},
        "test_patterns_used": _TEST_EMAIL_PATTERNS,
        "test_domains_used": _TEST_EMAIL_DOMAINS,
    }

    return {
        "status": "ok",
        "note": "READ-ONLY. Finds active records matching our test-account filter.",
        "summary": summary,
        "matches_by_source": matches_by_source,
        "all_matches": all_matches,
    }


@app.post("/api/admin/s28-ymove-bulk-no-filter")
async def s28_ymove_bulk_no_filter(request: Request):
    """READ-ONLY. Hit ymove member-lookup/all with NO status filter (and also try
    status=all and status=expired) to see if records exist that the default
    status=subscribed query doesn't return.

    Compares to our active+trialing Apple records and reports the delta.
    No DB writes."""
    pw = request.headers.get("X-Admin-Password", request.query_params.get("pw", ""))
    require_admin(pw)
    if not YMOVE_API_KEY:
        raise HTTPException(status_code=500, detail="YMOVE_API_KEY not configured")
    if not db_pool:
        raise HTTPException(status_code=500, detail="No database connected")

    async def pull_with_status(status_param: str) -> dict:
        """Pull bulk list with given status param (or empty for no filter)."""
        result = {"status_param": status_param, "pull_status": "starting", "pages": 0,
                  "total_users": 0, "by_provider": {}, "emails_by_provider": {}}
        emails_seen = set()
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                page = 1
                while True:
                    params = {"page": str(page)}
                    if status_param:
                        params["status"] = status_param
                    try:
                        resp = await client.get(
                            f"{YMOVE_API_BASE}/api/site/{YMOVE_SITE_ID}/member-lookup/all",
                            headers={"X-Authorization": YMOVE_API_KEY},
                            params=params
                        )
                    except httpx.TimeoutException:
                        result["pull_status"] = f"timeout_page_{page}"
                        break
                    if resp.status_code == 429:
                        retry_after = float(resp.headers.get("retry-after", "5"))
                        await asyncio.sleep(retry_after)
                        continue
                    if resp.status_code != 200:
                        result["pull_status"] = f"http_{resp.status_code}_page_{page}"
                        break
                    data = resp.json()
                    users = data.get("users", [])
                    for u in users:
                        em = (u.get("email") or "").strip().lower()
                        if not em or em in emails_seen:
                            continue
                        emails_seen.add(em)
                        provider = (u.get("subscriptionProvider") or "undetermined").lower()
                        if provider not in ("apple", "google", "stripe", "manual", "undetermined"):
                            provider = "undetermined"
                        result["by_provider"][provider] = result["by_provider"].get(provider, 0) + 1
                        result["emails_by_provider"].setdefault(provider, []).append({
                            "email": em,
                            "active_subscription": u.get("activeSubscription"),
                            "previously_subscribed": u.get("previouslySubscribed"),
                        })
                    total_pages = data.get("totalPages", 1)
                    result["pages"] = page
                    if page >= total_pages:
                        result["pull_status"] = "success"
                        break
                    page += 1
                    await asyncio.sleep(0.5)
        except Exception as e:
            result["pull_status"] = f"error: {str(e)[:200]}"
        result["total_users"] = len(emails_seen)
        return result, emails_seen

    # Pull with three different status values to compare
    no_filter_result, no_filter_emails = await pull_with_status("")
    # Brief gap between pulls
    await asyncio.sleep(1.0)
    all_filter_result, all_filter_emails = await pull_with_status("all")

    # Pull our current Apple records
    async with db_pool.acquire() as conn:
        our_apple_rows = await conn.fetch("""
            SELECT lower(email) AS email, stripe_subscription_id, status, source
            FROM subscriptions
            WHERE source = 'apple'
              AND status IN ('active', 'trialing')
              AND email IS NOT NULL AND email != ''
        """)
    our_apple_emails = {r["email"] for r in our_apple_rows}

    # Compare: which Apple-classified emails does no-filter pull have that we don't?
    def compare_apple(pull_result, pull_emails_set):
        ymove_apple_emails = set()
        for entry in pull_result["emails_by_provider"].get("apple", []):
            if entry.get("active_subscription"):
                ymove_apple_emails.add(entry["email"])

        ymove_apple_active_count = len(ymove_apple_emails)
        in_ymove_not_ours = sorted(ymove_apple_emails - our_apple_emails)
        in_ours_not_ymove = sorted(our_apple_emails - ymove_apple_emails)
        return {
            "ymove_apple_active_count": ymove_apple_active_count,
            "our_apple_count": len(our_apple_emails),
            "in_ymove_not_ours_count": len(in_ymove_not_ours),
            "in_ymove_not_ours_sample": in_ymove_not_ours[:30],
            "in_ours_not_ymove_count": len(in_ours_not_ymove),
            "in_ours_not_ymove_sample": in_ours_not_ymove[:30],
        }

    no_filter_apple_compare = compare_apple(no_filter_result, no_filter_emails)
    all_filter_apple_compare = compare_apple(all_filter_result, all_filter_emails)

    # Strip the heavy emails_by_provider from response (keep counts only) to keep payload sane
    no_filter_clean = {k: v for k, v in no_filter_result.items() if k != "emails_by_provider"}
    all_filter_clean = {k: v for k, v in all_filter_result.items() if k != "emails_by_provider"}

    return {
        "status": "ok",
        "note": "READ-ONLY. Compares ymove bulk pulls under different status filters.",
        "no_filter": {
            "pull": no_filter_clean,
            "apple_compare": no_filter_apple_compare,
        },
        "status_all": {
            "pull": all_filter_clean,
            "apple_compare": all_filter_apple_compare,
        },
        "interpretation_hint": (
            "If no_filter or status=all returns more Apple users with activeSubscription=true "
            "than the default status=subscribed pull, our shadow sync is missing them."
        ),
    }


def _s28_verdict(our_rec, stripe_check) -> str:
    """Categorize a Stripe-source record based on our DB state vs Stripe API truth."""
    if not our_rec:
        return "no_record_in_our_db"
    if stripe_check.get("error"):
        return "stripe_api_error"
    if not stripe_check.get("stripe_found"):
        return "no_stripe_customer"
    if stripe_check.get("any_active"):
        # Find the active sub
        active_subs = [s for s in stripe_check["subs"] if s["status"] in ("active", "trialing")]
        if active_subs:
            our_sub_id = our_rec.get("stripe_subscription_id")
            matching = [s for s in active_subs if s["sub_id"] == our_sub_id]
            if matching:
                if matching[0].get("cancel_at_period_end"):
                    return "active_pending_cancel_in_stripe"
                return "active_in_stripe_matches_ours"
            return "active_in_stripe_different_sub_id"
    return "no_active_sub_in_stripe"


# --- Static + Routing ---

@app.get("/mm-admin")
async def admin_page():
    return FileResponse("static/admin.html")

@app.get("/")
async def root():
    return FileResponse("static/index.html")

app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/{path:path}")
async def catch_all(path: str):
    return FileResponse("static/index.html")
